import json
import hashlib
import os
import re
import tempfile
import urllib.error
import urllib.parse
import urllib.request
import uuid
from io import BytesIO
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

import excel_bancos
from xlsx_templates import build_import_template


REGISTRATIONS_TABLE = "cadastro_registros"
DRAFTS_TABLE = "cadastro_rascunhos"
BOM_HEADERS_TABLE = "cadastro_bom_cabecalhos"
BOM_COMPONENTS_TABLE = "cadastro_bom_componentes"
EXPORT_DIR = Path(tempfile.gettempdir()) / "modulo-cadastro-exports"
ALL_CATEGORIES_KEY = "__all__"
REVIEW_PARENT_PREFIX = "REVISAO-"
DUPLICATE_PARENT_SEPARATOR = "__BOM__"
UNIT_OPTIONS = ["pc", "un", "cj", "ch", "br", "m", "mm"]
SKU_MIGRATION_FORM_KEY = "_sku_migration"
PREVIOUS_SKU_FORM_KEY = "_sku_anterior"


class SupabaseStoreError(RuntimeError):
    pass


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_unit(value: Any) -> str:
    return clean_text(value).lower()


def unidade_options() -> list[str]:
    return list(UNIT_OPTIONS)


def status_to_active(value: Any, default: bool = True) -> bool:
    text = clean_text(value).upper()
    if not text:
        return default
    return text not in {
        "0",
        "FALSE",
        "NAO",
        "NÃO",
        "NO",
        "OFF",
        "INATIVO",
        "INATIVA",
        "INATIVADO",
        "INATIVADA",
        "DESATIVADO",
        "DESATIVADA",
        "CANCELADO",
        "CANCELADA",
        "OBSOLETO",
        "OBSOLETA",
        "BLOQUEADO",
        "BLOQUEADA",
    }


def save_mode() -> str:
    return clean_text(os.environ.get("CADASTRO_SAVE_MODE")).lower()


def enabled() -> bool:
    return save_mode() in {"supabase", "postgres", "database", "banco"}


def configured() -> bool:
    return bool(_supabase_url() and _service_key())


def status() -> dict[str, Any]:
    return {
        "enabled": enabled(),
        "configured": configured(),
        "url": _supabase_url(),
        "tables": [REGISTRATIONS_TABLE, DRAFTS_TABLE, BOM_HEADERS_TABLE, BOM_COMPONENTS_TABLE],
    }


def display_target() -> str:
    if not enabled():
        return ""
    if _supabase_url():
        return f"Modo Supabase: {_supabase_url()} ({REGISTRATIONS_TABLE})"
    return "Modo Supabase: configure SUPABASE_URL e SUPABASE_SERVICE_ROLE_KEY"


def _supabase_url() -> str:
    raw = clean_text(os.environ.get("SUPABASE_URL")) or clean_text(os.environ.get("CADASTRO_SUPABASE_URL"))
    return raw.rstrip("/")


def _service_key() -> str:
    return (
        clean_text(os.environ.get("SUPABASE_SERVICE_ROLE_KEY"))
        or clean_text(os.environ.get("CADASTRO_SUPABASE_SERVICE_ROLE_KEY"))
    )


def _ensure_configured() -> None:
    if not enabled():
        raise SupabaseStoreError("Modo Supabase não está ativo.")
    if not _supabase_url():
        raise SupabaseStoreError("Configure SUPABASE_URL no Render.")
    if not _service_key():
        raise SupabaseStoreError("Configure SUPABASE_SERVICE_ROLE_KEY no Render.")


def _headers(prefer: str = "") -> dict[str, str]:
    key = _service_key()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def _request(
    method: str,
    table: str,
    query: list[tuple[str, str]] | None = None,
    payload: Any = None,
    prefer: str = "",
) -> Any:
    _ensure_configured()
    query_string = urllib.parse.urlencode(query or [])
    url = f"{_supabase_url()}/rest/v1/{table}"
    if query_string:
        url = f"{url}?{query_string}"
    data = None
    if payload is not None:
        data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(url, data=data, headers=_headers(prefer), method=method)
    try:
        with urllib.request.urlopen(request, timeout=30) as response:
            body = response.read().decode("utf-8")
            return json.loads(body) if body else None
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise SupabaseStoreError(f"Erro Supabase {exc.code}: {body}") from exc
    except urllib.error.URLError as exc:
        raise SupabaseStoreError(f"Não foi possível conectar ao Supabase: {exc}") from exc


def _request_all(table: str, params: list[tuple[str, str]], limit: int = 10000) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    page_size = 1000
    while len(rows) < limit:
        batch = _request(
            "GET",
            table,
            [
                *params,
                ("limit", str(min(page_size, limit - len(rows)))),
                ("offset", str(len(rows))),
            ],
        ) or []
        rows.extend(batch)
        if len(batch) < page_size:
            break
    return rows


def _is_missing_column_error(exc: Exception, column: str) -> bool:
    text = str(exc).lower()
    column = column.lower()
    return column in text and (
        "does not exist" in text
        or "could not find" in text
        or "schema cache" in text
        or "pgrst204" in text
    )


def _without_filter(params: list[tuple[str, str]], key: str) -> list[tuple[str, str]]:
    return [(param_key, value) for param_key, value in params if param_key != key]


def _category(category_key: str) -> dict[str, Any]:
    catalog = excel_bancos.load_catalog()
    return excel_bancos._find_category(catalog, category_key)


def _sheet_name(category: dict[str, Any]) -> str:
    return excel_bancos._safe_sheet_title(category.get("sheet_name") or category.get("label") or "Categoria")


def _category_key_filter(column: str, category_key: str) -> tuple[str, str] | None:
    keys = excel_bancos.category_key_candidates(category_key)
    if not keys:
        return None
    if len(keys) == 1:
        return (column, f"eq.{keys[0]}")
    return (column, "in.(" + ",".join(keys) + ")")


def _initial_sku(code_prefix: str) -> str:
    return f"{code_prefix}0001"


def _next_sku(category: dict[str, Any], fields: list[dict[str, Any]], form_data: Any) -> str:
    code_prefix = excel_bancos.pn_code_prefix(category, fields, form_data)
    rows = _request(
        "GET",
        REGISTRATIONS_TABLE,
        [
            ("select", "sku"),
            _category_key_filter("category_key", category["key"]),
            ("sku", f"like.{code_prefix}%"),
            ("order", "sku.desc"),
            ("limit", "1"),
        ],
    )
    last_code = ""
    if rows:
        last_code = clean_text(rows[0].get("sku"))
    if not last_code or not last_code.isdigit():
        return _initial_sku(code_prefix)
    return str(int(last_code) + 1).zfill(len(last_code))


def _field_groups(fields: list[dict[str, Any]], form_data: Any) -> dict[str, list[str]]:
    groups: dict[str, list[str]] = {}
    group_code = excel_bancos._pn_group_code(form_data.get(excel_bancos.PN_GROUP_FORM_KEY) if hasattr(form_data, "get") else "")
    if group_code:
        groups[excel_bancos.PN_GROUP_FORM_KEY] = [group_code]
    for field in fields:
        values = excel_bancos._serialize_field_values(field, form_data)
        groups[field["key"]] = values
    return groups


def _field_values(fields: list[dict[str, Any]], groups: dict[str, list[str]]) -> dict[str, str]:
    values: dict[str, str] = {}
    for field in fields:
        selected = groups.get(field["key"]) or []
        values[field["key"]] = excel_bancos._format_field_saved_value(field, selected) if selected else ""
    return values


def _field_codes(fields: list[dict[str, Any]], groups: dict[str, list[str]]) -> dict[str, str]:
    codes: dict[str, str] = {}
    for field in fields:
        selected = groups.get(field["key"]) or []
        field_codes = [excel_bancos.option_code(value) for value in selected if excel_bancos.option_code(value)]
        codes[field["key"]] = " | ".join(field_codes)
    return codes


def _search_text(*parts: Any) -> str:
    return excel_bancos.normalize_label(" ".join(clean_text(part) for part in parts if clean_text(part)))


def _is_missing_bom_code(value: Any) -> bool:
    text = clean_text(value)
    normalized = excel_bancos.normalize_label(text)
    return not text or text == "0" or normalized in {"N D", "ND", "N A"} or text.upper() == "#N/A"


def _review_parent_key(seed: str) -> str:
    digest = hashlib.sha1(clean_text(seed).encode("utf-8")).hexdigest()[:12].upper()
    return f"{REVIEW_PARENT_PREFIX}{digest}"


def _duplicate_parent_key(parent_sku: str, seed: str) -> str:
    digest = hashlib.sha1(clean_text(seed).encode("utf-8")).hexdigest()[:10].upper()
    return f"{clean_text(parent_sku)}{DUPLICATE_PARENT_SEPARATOR}{digest}"


def _base_parent_sku(parent_sku: str) -> str:
    text = clean_text(parent_sku)
    if DUPLICATE_PARENT_SEPARATOR in text:
        return text.split(DUPLICATE_PARENT_SEPARATOR, 1)[0]
    return text


def _source_with_review(source: str, reasons: list[str]) -> str:
    base = clean_text(source) or "cadastro"
    cleaned = list(dict.fromkeys(clean_text(reason) for reason in reasons if clean_text(reason)))
    if not cleaned:
        return base
    return f"{base}|needs_review:{','.join(cleaned)}"


def _review_reasons(source: str) -> list[str]:
    marker = "needs_review:"
    text = clean_text(source)
    if marker not in text:
        return []
    return [part.strip() for part in text.split(marker, 1)[1].split("|", 1)[0].split(",") if part.strip()]


def _review_reason_label(reason: str) -> str:
    labels = {
        "parent_code": "Item pai sem codigo",
        "component_code": "Item filho sem codigo",
        "quantity_default": "Quantidade ajustada para 1",
        "empty_component": "Linha de componente incompleta",
        "duplicate_parent": "Item pai duplicado no diretorio",
    }
    return labels.get(clean_text(reason), clean_text(reason))


def _display_bom_code(value: Any) -> str:
    text = clean_text(value)
    if DUPLICATE_PARENT_SEPARATOR in text:
        return _base_parent_sku(text)
    if text.startswith(REVIEW_PARENT_PREFIX) or _is_missing_bom_code(text):
        return ""
    return text


def _full_description(row: dict[str, Any]) -> str:
    return clean_text(row.get("descricao_primaria"))


def _catalog_data_by_sku(skus: list[Any]) -> dict[str, dict[str, str]]:
    codes = list(dict.fromkeys(clean_text(sku) for sku in skus if clean_text(sku)))
    if not codes:
        return {}
    rows = _request_all(
        REGISTRATIONS_TABLE,
        [
            ("select", "sku,descricao_primaria,unidade"),
            ("sku", _in_filter(codes)),
            ("order", "sku.asc"),
        ],
        limit=max(len(codes), 1),
    )
    return {
        clean_text(row.get("sku")): {
            "descricao_primaria": clean_text(row.get("descricao_primaria")),
            "unidade": normalize_unit(row.get("unidade")),
        }
        for row in rows
        if clean_text(row.get("sku"))
    }


def _duplicate_exists(
    category_key: str,
    primaria: str,
    secundaria: str,
    exclude_id: int | str | None = None,
) -> bool:
    params = [
        ("select", "id"),
        ("descricao_primaria", f"eq.{primaria}"),
        ("descricao_secundaria", f"eq.{secundaria}"),
        ("limit", "1"),
    ]
    category_filter = _category_key_filter("category_key", category_key)
    if category_filter:
        params.append(category_filter)
    if exclude_id is not None:
        params.append(("id", f"neq.{clean_text(exclude_id)}"))
    rows = _request(
        "GET",
        REGISTRATIONS_TABLE,
        params,
    )
    return bool(rows)


def _registration_payload(
    category: dict[str, Any],
    fields: list[dict[str, Any]],
    form_data: Any,
    sku: str,
    default_active: bool,
) -> tuple[dict[str, Any], dict[str, str], bool]:
    if category["key"] == excel_bancos.DEFAULT_CATEGORY_KEY:
        excel_bancos._validate_banco_dependencies(fields, form_data)
        excel_bancos._validate_visible_field_requirements(fields, category["key"], form_data)

    descriptions = excel_bancos.build_descriptions(fields, form_data, category["key"])
    groups = _field_groups(fields, form_data)
    field_values = _field_values(fields, groups)
    field_codes = _field_codes(fields, groups)
    unidade = normalize_unit(form_data.get("unidade"))
    ativo = status_to_active(form_data.get("ativo"), default=default_active)
    possui_bom = excel_bancos.requires_component_bom(fields, form_data)
    groups[excel_bancos.BOM_FORM_KEY] = possui_bom
    payload = {
        "category_key": category["key"],
        "category_label": category["label"],
        "sheet": _sheet_name(category),
        "sku": sku,
        "descricao_primaria": descriptions["primaria"],
        "descricao_secundaria": descriptions["secundaria"],
        "sufixo": descriptions.get("sufixo") or "",
        "unidade": unidade,
        "ativo": ativo,
        "caracteres_primario": len(descriptions["primaria"]),
        "caracteres_secundario": len(descriptions["secundaria"]),
        "form_values": groups,
        "field_values": field_values,
        "field_codes": field_codes,
        "search_text": _search_text(
            sku,
            category["label"],
            descriptions["primaria"],
            descriptions["secundaria"],
            unidade,
            " ".join(field_values.values()),
        ),
    }
    return payload, descriptions, possui_bom


def save_registration(form_data: Any) -> dict[str, Any]:
    category_key = clean_text(form_data.get("categoria"))
    category = _category(category_key)
    requested_group = excel_bancos._pn_group_code(form_data.get(excel_bancos.PN_GROUP_FORM_KEY))
    if category["key"] == excel_bancos.DEFAULT_CATEGORY_KEY and not requested_group:
        raise SupabaseStoreError("Selecione o grupo do cadastro.")
    fields = excel_bancos.get_banco_fields(category["key"], requested_group)
    if category["key"] == excel_bancos.DEFAULT_CATEGORY_KEY:
        excel_bancos._validate_banco_dependencies(fields, form_data)
        excel_bancos._validate_visible_field_requirements(fields, category["key"], form_data)

    descriptions = excel_bancos.build_descriptions(fields, form_data, category["key"])
    if _duplicate_exists(category["key"], descriptions["primaria"], descriptions["secundaria"]):
        raise SupabaseStoreError("Cadastro já existe com a mesma descrição primária e secundária.")

    groups = _field_groups(fields, form_data)
    field_values = _field_values(fields, groups)
    field_codes = _field_codes(fields, groups)
    sku = _next_sku(category, fields, form_data)
    unidade = normalize_unit(form_data.get("unidade"))
    ativo = status_to_active(form_data.get("ativo"), default=True)
    possui_bom = excel_bancos.requires_component_bom(fields, form_data)
    groups[excel_bancos.BOM_FORM_KEY] = possui_bom
    payload = {
        "category_key": category["key"],
        "category_label": category["label"],
        "sheet": _sheet_name(category),
        "sku": sku,
        "descricao_primaria": descriptions["primaria"],
        "descricao_secundaria": descriptions["secundaria"],
        "sufixo": descriptions.get("sufixo") or "",
        "unidade": unidade,
        "ativo": ativo,
        "caracteres_primario": len(descriptions["primaria"]),
        "caracteres_secundario": len(descriptions["secundaria"]),
        "form_values": groups,
        "field_values": field_values,
        "field_codes": field_codes,
        "search_text": _search_text(
            sku,
            category["label"],
            descriptions["primaria"],
            descriptions["secundaria"],
            unidade,
            " ".join(field_values.values()),
        ),
    }
    rows = _request("POST", REGISTRATIONS_TABLE, payload=payload, prefer="return=representation")
    row = rows[0] if rows else payload
    return {
        "id": row.get("id"),
        "row": row.get("id") or "-",
        "category": category["label"],
        "category_key": category["key"],
        "sheet": _sheet_name(category),
        "descricao_primaria": descriptions["primaria"],
        "descricao_secundaria": descriptions["secundaria"],
        "unidade": unidade,
        "ativo": ativo,
        "possui_bom": possui_bom,
        "sku": sku,
        "path": display_target(),
    }


def _draft_groups(payload_json: str | dict[str, Any]) -> dict[str, list[str]]:
    return excel_bancos._draft_payload_groups(payload_json)


def save_draft(category_key: str, draft_payload: str, draft_id: str = "") -> dict[str, Any]:
    category = _category(category_key)
    groups = _draft_groups(draft_payload)
    if not groups:
        raise SupabaseStoreError("Rascunho vazio.")
    requested_group = excel_bancos._pn_group_code(groups.get(excel_bancos.PN_GROUP_FORM_KEY))
    fields = excel_bancos.get_banco_fields(category["key"], requested_group)
    descriptions = excel_bancos.build_descriptions(fields, groups, category["key"])
    draft_id = clean_text(draft_id) or uuid.uuid4().hex[:12]
    payload = {
        "draft_id": draft_id,
        "category_key": category["key"],
        "category_label": category["label"],
        "sheet": _sheet_name(category),
        "descricao_primaria": descriptions.get("primaria") or "(sem descrição primária)",
        "payload": {"category": category["key"], "groups": groups},
    }
    existing = _request("GET", DRAFTS_TABLE, [("select", "draft_id"), ("draft_id", f"eq.{draft_id}"), ("limit", "1")])
    if existing:
        rows = _request(
            "PATCH",
            DRAFTS_TABLE,
            [("draft_id", f"eq.{draft_id}")],
            payload=payload,
            prefer="return=representation",
        )
    else:
        rows = _request("POST", DRAFTS_TABLE, payload=payload, prefer="return=representation")
    row = rows[0] if rows else payload
    return _draft_summary(row)


def _draft_summary(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "draft_id": clean_text(row.get("draft_id")),
        "category_key": clean_text(row.get("category_key")),
        "category_label": clean_text(row.get("category_label")),
        "sheet": clean_text(row.get("sheet")),
        "saved_at": clean_text(row.get("updated_at") or row.get("created_at")),
        "descricao_primaria": clean_text(row.get("descricao_primaria")),
    }


def list_drafts() -> list[dict[str, Any]]:
    rows = _request("GET", DRAFTS_TABLE, [("select", "*"), ("order", "updated_at.desc")]) or []
    return [_draft_summary(row) for row in rows]


def get_draft(draft_id: str) -> dict[str, Any] | None:
    rows = _request("GET", DRAFTS_TABLE, [("select", "*"), ("draft_id", f"eq.{clean_text(draft_id)}"), ("limit", "1")])
    if not rows:
        return None
    row = rows[0]
    payload = row.get("payload") if isinstance(row.get("payload"), dict) else {}
    return {
        **_draft_summary(row),
        "groups": _draft_groups(payload),
    }


def delete_draft(draft_id: str) -> dict[str, Any]:
    existing = get_draft(draft_id)
    if not existing:
        raise SupabaseStoreError("Rascunho não encontrado.")
    _request("DELETE", DRAFTS_TABLE, [("draft_id", f"eq.{clean_text(draft_id)}")])
    return existing


def _safe_filter_value(value: str) -> str:
    return clean_text(value).replace("*", "").replace(",", " ")


def all_categories_key(value: str) -> bool:
    return clean_text(value).lower() in {ALL_CATEGORIES_KEY, "all", "todas", "todos", "*"}


def _group_label_map() -> dict[str, str]:
    return {group["code"]: group["label"] for group in excel_bancos.list_pn_groups()}


def _row_group_code(row: dict[str, Any]) -> str:
    form_values = row.get("form_values") if isinstance(row.get("form_values"), dict) else {}
    value = form_values.get(excel_bancos.PN_GROUP_FORM_KEY)
    if isinstance(value, list):
        for item in value:
            code = excel_bancos._pn_group_code(item)
            if code:
                return code
    else:
        code = excel_bancos._pn_group_code(value)
        if code:
            return code
    sku = clean_text(row.get("sku"))
    return sku[:2] if len(sku) >= 2 and sku[:2].isdigit() else ""


def _stored_bom_preference(row: dict[str, Any]) -> bool | None:
    form_values = row.get("form_values") if isinstance(row.get("form_values"), dict) else {}
    if excel_bancos.BOM_FORM_KEY not in form_values:
        # Registros de Veiculo P.B. anteriores ao campo explicito sempre representam
        # transformacoes que podem receber uma estrutura de produto.
        if clean_text(row.get("category_key")) == "cat_34_veiculo_p_b":
            return True
        return None
    value = form_values.get(excel_bancos.BOM_FORM_KEY)
    if isinstance(value, list):
        value = value[0] if value else ""
    if isinstance(value, bool):
        return value
    return status_to_active(value, default=False)


def _enrich_registration_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    labels = _group_label_map()
    for row in rows:
        code = _row_group_code(row)
        row["grupo_codigo"] = code
        row["grupo_label"] = labels.get(code, "")
        row["possui_bom"] = _stored_bom_preference(row)
    return rows


def list_registrations(
    category_key: str = "",
    query: str = "",
    filters: dict[str, str] | None = None,
    group_code: str = "",
    missing_unit: bool = False,
    include_inactive: bool = False,
    limit: int = 250,
    offset: int = 0,
) -> list[dict[str, Any]]:
    category_value = clean_text(category_key)
    all_categories = all_categories_key(category_value)
    selected = "" if all_categories else _category(category_value)["key"] if category_value else excel_bancos.selected_category("")["key"]
    requested_limit = max(1, min(limit, 10000))
    requested_offset = max(0, offset)
    params: list[tuple[str, str]] = [
        ("select", "*"),
        ("order", "category_label.asc,sku.asc" if all_categories else "sku.asc"),
    ]
    if selected:
        category_filter = _category_key_filter("category_key", selected)
        if category_filter:
            params.append(category_filter)
    if not include_inactive:
        params.append(("ativo", "is.true"))
    term = _search_text(query)
    if term:
        params.append(("search_text", f"ilike.*{term}*"))
    selected_group = excel_bancos._pn_group_code(group_code)
    if selected_group:
        params.append(("sku", f"like.{selected_group}%"))
    if missing_unit:
        params.append(("unidade", "eq."))
    for key, value in (filters or {}).items():
        value = _safe_filter_value(value)
        if key and value:
            params.append((f"field_values->>{key}", f"ilike.*{value}*"))
    fallback_without_active = False
    if requested_limit <= 1000:
        try:
            rows = _request(
                "GET",
                REGISTRATIONS_TABLE,
                [*params, ("limit", str(requested_limit)), ("offset", str(requested_offset))],
            ) or []
            return _enrich_registration_rows(rows)
        except SupabaseStoreError as exc:
            if include_inactive or not _is_missing_column_error(exc, "ativo"):
                raise
            fallback_without_active = True
            fallback_params = _without_filter(params, "ativo")
            rows = _request(
                "GET",
                REGISTRATIONS_TABLE,
                [*fallback_params, ("limit", str(requested_limit)), ("offset", str(requested_offset))],
            ) or []
            for row in rows:
                row.setdefault("ativo", True)
            return _enrich_registration_rows(rows)

    rows: list[dict[str, Any]] = []
    page_size = 1000
    while len(rows) < requested_limit:
        try:
            batch = _request(
                "GET",
                REGISTRATIONS_TABLE,
                [
                    *params,
                    ("limit", str(min(page_size, requested_limit - len(rows)))),
                    ("offset", str(requested_offset + len(rows))),
                ],
            ) or []
        except SupabaseStoreError as exc:
            if include_inactive or fallback_without_active or not _is_missing_column_error(exc, "ativo"):
                raise
            fallback_without_active = True
            params = _without_filter(params, "ativo")
            continue
        if fallback_without_active:
            for row in batch:
                row.setdefault("ativo", True)
        rows.extend(batch)
        if len(batch) < page_size:
            break
    return _enrich_registration_rows(rows)


def count_registrations_without_unit(category_key: str = "", include_inactive: bool = False) -> int:
    category_value = clean_text(category_key)
    all_categories = all_categories_key(category_value)
    selected = "" if all_categories else _category(category_value)["key"] if category_value else excel_bancos.selected_category("")["key"]
    params: list[tuple[str, str]] = [
        ("select", "id"),
        ("unidade", "eq."),
        ("limit", "10000"),
    ]
    if selected:
        category_filter = _category_key_filter("category_key", selected)
        if category_filter:
            params.append(category_filter)
    if not include_inactive:
        params.append(("ativo", "is.true"))
    try:
        rows = _request("GET", REGISTRATIONS_TABLE, params) or []
    except SupabaseStoreError as exc:
        if include_inactive or not _is_missing_column_error(exc, "ativo"):
            raise
        rows = _request("GET", REGISTRATIONS_TABLE, _without_filter(params, "ativo")) or []
    return len(rows)


def count_inactive_registrations(category_key: str = "") -> int:
    category_value = clean_text(category_key)
    all_categories = all_categories_key(category_value)
    selected = "" if all_categories else _category(category_value)["key"] if category_value else excel_bancos.selected_category("")["key"]
    params: list[tuple[str, str]] = [
        ("select", "id"),
        ("ativo", "is.false"),
        ("limit", "10000"),
    ]
    if selected:
        category_filter = _category_key_filter("category_key", selected)
        if category_filter:
            params.append(category_filter)
    try:
        rows = _request("GET", REGISTRATIONS_TABLE, params) or []
    except SupabaseStoreError as exc:
        if not _is_missing_column_error(exc, "ativo"):
            raise
        return 0
    return len(rows)


def get_registration(registration_id: int | str) -> dict[str, Any] | None:
    rows = _request(
        "GET",
        REGISTRATIONS_TABLE,
        [
            ("select", "*"),
            ("id", f"eq.{clean_text(registration_id)}"),
            ("limit", "1"),
        ],
    )
    return rows[0] if rows else None


def _groups_from_record(fields: list[dict[str, Any]], record: dict[str, Any]) -> dict[str, list[str]]:
    form_values = record.get("form_values") if isinstance(record.get("form_values"), dict) else {}
    field_values = record.get("field_values") if isinstance(record.get("field_values"), dict) else {}
    groups: dict[str, list[str]] = {}
    for field in fields:
        raw = form_values.get(field["key"])
        if isinstance(raw, list) and raw:
            groups[field["key"]] = [clean_text(value) for value in raw if clean_text(value)]
            continue
        saved = clean_text(field_values.get(field["key"]))
        if saved:
            groups[field["key"]] = [value.strip() for value in saved.split("|") if value.strip()]
    return groups


def editable_registration(
    registration_id: int | str,
    target_category_key: str = "",
    target_group_code: str = "",
) -> dict[str, Any]:
    record = get_registration(registration_id)
    if not record:
        raise SupabaseStoreError("Cadastro não encontrado.")
    source_category = _category(clean_text(record.get("category_key")))
    category = _category(clean_text(target_category_key)) if clean_text(target_category_key) else source_category
    possui_bom = _stored_bom_preference(record)
    if possui_bom is None:
        possui_bom = bool(_bom_header_by_parent(clean_text(record.get("sku"))))
    form_values = record.get("form_values") if isinstance(record.get("form_values"), dict) else {}
    migration = form_values.get(SKU_MIGRATION_FORM_KEY) if isinstance(form_values.get(SKU_MIGRATION_FORM_KEY), dict) else {}
    record = {
        **record,
        "possui_bom": possui_bom,
        "replacement_sku": clean_text(migration.get("replacement_sku")),
        "replacement_id": migration.get("replacement_id"),
    }
    source_group_code = _row_group_code(record)
    target_group = excel_bancos._pn_group_code(target_group_code) or source_group_code
    fields = excel_bancos.get_banco_fields(category["key"], target_group)
    groups = _groups_from_record(fields, record)
    return {
        "record": record,
        "category": category,
        "source_category": source_category,
        "current_group_code": source_group_code,
        "source_group_code": source_group_code,
        "target_group_code": target_group,
        "fields": fields,
        "groups": groups,
    }


def _registration_structure_changed(
    current: dict[str, Any],
    target_category: dict[str, Any],
    fields: list[dict[str, Any]],
    form_data: Any,
) -> bool:
    current_category_key = clean_text(current.get("category_key"))
    current_group_code = _row_group_code(current)
    target_group_code = excel_bancos.pn_group_code(fields, form_data)
    return (
        not excel_bancos.same_category_key(target_category["key"], current_category_key)
        or target_group_code != current_group_code
    )


def _bom_reference_snapshots(sku: str) -> dict[str, Any]:
    header_rows = _request_all(
        BOM_HEADERS_TABLE,
        [("select", "*"), ("parent_sku", f"eq.{sku}")],
        limit=10,
    )
    parent_rows = _request_all(
        BOM_COMPONENTS_TABLE,
        [("select", "*"), ("parent_sku", f"eq.{sku}")],
    )
    component_rows = _request_all(
        BOM_COMPONENTS_TABLE,
        [("select", "*"), ("component_sku", f"eq.{sku}")],
    )
    components = {clean_text(row.get("id")): row for row in [*parent_rows, *component_rows]}
    return {"headers": header_rows, "components": list(components.values())}


def _bom_header_restore_payload(row: dict[str, Any]) -> dict[str, Any]:
    keys = {
        "parent_sku",
        "parent_descricao",
        "parent_category_key",
        "parent_category_label",
        "registration_id",
        "source",
        "search_text",
    }
    return {key: row.get(key) for key in keys}


def _bom_component_restore_payload(row: dict[str, Any]) -> dict[str, Any]:
    keys = {
        "bom_id",
        "parent_sku",
        "component_sku",
        "component_descricao",
        "unidade",
        "quantidade",
        "ordem",
        "search_text",
    }
    return {key: row.get(key) for key in keys}


def _apply_bom_sku_migration(
    snapshots: dict[str, Any],
    old_sku: str,
    new_record: dict[str, Any],
) -> dict[str, int]:
    new_sku = clean_text(new_record.get("sku"))
    for header in snapshots["headers"]:
        source = clean_text(header.get("source"))
        header_payload = {
            "parent_sku": new_sku,
            "parent_descricao": clean_text(new_record.get("descricao_primaria")),
            "parent_category_key": clean_text(new_record.get("category_key")),
            "parent_category_label": clean_text(new_record.get("category_label")),
            "registration_id": new_record.get("id"),
            "source": source,
            "search_text": _search_text(
                new_sku,
                new_record.get("descricao_primaria"),
                new_record.get("category_label"),
                source,
            ),
        }
        _request(
            "PATCH",
            BOM_HEADERS_TABLE,
            [("id", f"eq.{header['id']}")],
            payload=header_payload,
            prefer="return=minimal",
        )

    for component in snapshots["components"]:
        payload = _bom_component_restore_payload(component)
        if clean_text(payload.get("parent_sku")) == old_sku:
            payload["parent_sku"] = new_sku
        if clean_text(payload.get("component_sku")) == old_sku:
            payload["component_sku"] = new_sku
            payload["component_descricao"] = clean_text(new_record.get("descricao_primaria"))
            payload["unidade"] = normalize_unit(new_record.get("unidade"))
        payload["search_text"] = _search_text(
            payload.get("parent_sku"),
            payload.get("component_sku"),
            payload.get("component_descricao"),
            payload.get("unidade"),
        )
        _request(
            "PATCH",
            BOM_COMPONENTS_TABLE,
            [("id", f"eq.{component['id']}")],
            payload=payload,
            prefer="return=minimal",
        )
    return {"bom_headers": len(snapshots["headers"]), "bom_components": len(snapshots["components"])}


def _restore_bom_references(snapshots: dict[str, Any]) -> None:
    for component in snapshots.get("components") or []:
        _request(
            "PATCH",
            BOM_COMPONENTS_TABLE,
            [("id", f"eq.{component['id']}")],
            payload=_bom_component_restore_payload(component),
            prefer="return=minimal",
        )
    for header in snapshots.get("headers") or []:
        _request(
            "PATCH",
            BOM_HEADERS_TABLE,
            [("id", f"eq.{header['id']}")],
            payload=_bom_header_restore_payload(header),
            prefer="return=minimal",
        )


def update_registration(registration_id: int | str, form_data: Any) -> dict[str, Any]:
    current = get_registration(registration_id)
    if not current:
        raise SupabaseStoreError("Cadastro não encontrado.")
    current_values = current.get("form_values") if isinstance(current.get("form_values"), dict) else {}
    previous_migration = current_values.get(SKU_MIGRATION_FORM_KEY)
    if isinstance(previous_migration, dict) and clean_text(previous_migration.get("replacement_sku")):
        raise SupabaseStoreError(
            f"Este cadastro foi substituido pelo SKU {clean_text(previous_migration.get('replacement_sku'))}."
        )

    requested_group = excel_bancos._pn_group_code(form_data.get(excel_bancos.PN_GROUP_FORM_KEY))
    if not requested_group:
        raise SupabaseStoreError("Selecione o grupo do cadastro.")
    target_category = _category(clean_text(form_data.get("categoria")) or clean_text(current.get("category_key")))
    fields = excel_bancos.get_banco_fields(target_category["key"], requested_group)

    old_sku = clean_text(current.get("sku"))
    structure_changed = _registration_structure_changed(current, target_category, fields, form_data)
    new_sku = _next_sku(target_category, fields, form_data) if structure_changed else old_sku
    payload, descriptions, _ = _registration_payload(target_category, fields, form_data, new_sku, False)
    if _duplicate_exists(
        target_category["key"],
        descriptions["primaria"],
        descriptions["secundaria"],
        exclude_id=registration_id,
    ):
        raise SupabaseStoreError("Ja existe outro cadastro com a mesma descricao primaria e secundaria.")

    if not structure_changed:
        rows = _request(
            "PATCH",
            REGISTRATIONS_TABLE,
            [("id", f"eq.{clean_text(registration_id)}")],
            payload=payload,
            prefer="return=representation",
        )
        return rows[0] if rows else {**current, **payload}

    if clean_text(form_data.get("confirmar_migracao")) != "1":
        raise SupabaseStoreError(
            "Confirme a geracao de um novo SKU e a inativacao do codigo anterior."
        )

    payload["form_values"] = {
        **payload["form_values"],
        PREVIOUS_SKU_FORM_KEY: old_sku,
    }
    snapshots = _bom_reference_snapshots(old_sku)
    new_rows = _request(
        "POST",
        REGISTRATIONS_TABLE,
        payload=payload,
        prefer="return=representation",
    ) or []
    if not new_rows or not new_rows[0].get("id"):
        raise SupabaseStoreError("Nao foi possivel criar o cadastro substituto.")
    new_record = new_rows[0]

    try:
        migrated = _apply_bom_sku_migration(snapshots, old_sku, new_record)
        old_form_values = dict(current_values)
        old_form_values[SKU_MIGRATION_FORM_KEY] = {
            "replacement_id": new_record["id"],
            "replacement_sku": new_sku,
            "migrated_at": datetime.now(timezone.utc).isoformat(),
        }
        old_rows = _request(
            "PATCH",
            REGISTRATIONS_TABLE,
            [("id", f"eq.{clean_text(registration_id)}")],
            payload={
                "ativo": False,
                "form_values": old_form_values,
                "search_text": _search_text(current.get("search_text"), "substituido por", new_sku),
            },
            prefer="return=representation",
        )
        if not old_rows:
            raise SupabaseStoreError("Nao foi possivel inativar o SKU anterior.")
    except Exception as exc:
        rollback_errors = []
        try:
            _restore_bom_references(snapshots)
        except Exception as rollback_exc:
            rollback_errors.append(f"B.O.M.: {rollback_exc}")
        try:
            _request(
                "PATCH",
                REGISTRATIONS_TABLE,
                [("id", f"eq.{clean_text(registration_id)}")],
                payload={
                    "ativo": current.get("ativo", True),
                    "form_values": current_values,
                    "search_text": clean_text(current.get("search_text")),
                },
                prefer="return=minimal",
            )
        except Exception as rollback_exc:
            rollback_errors.append(f"SKU anterior: {rollback_exc}")
        try:
            _request("DELETE", REGISTRATIONS_TABLE, [("id", f"eq.{new_record['id']}")])
        except Exception as rollback_exc:
            rollback_errors.append(f"SKU substituto: {rollback_exc}")
        if rollback_errors:
            raise SupabaseStoreError(
                f"Falha na migracao ({exc}). Reversao incompleta: {'; '.join(rollback_errors)}"
            ) from exc
        raise

    return {
        **new_record,
        "migrated": True,
        "previous_sku": old_sku,
        **migrated,
    }


def search_products(query: str, limit: int = 25) -> list[dict[str, str]]:
    term = _search_text(query)
    if len(term) < 1:
        return []
    params = [
        ("select", "sku,descricao_primaria,category_label,unidade,search_text"),
        ("search_text", f"ilike.*{term}*"),
        ("ativo", "is.true"),
        ("order", "sku.asc"),
        ("limit", str(limit)),
    ]
    try:
        rows = _request("GET", REGISTRATIONS_TABLE, params) or []
    except SupabaseStoreError as exc:
        if not _is_missing_column_error(exc, "ativo"):
            raise
        rows = _request("GET", REGISTRATIONS_TABLE, _without_filter(params, "ativo")) or []
    return [
        {
            "codigo": clean_text(row.get("sku")),
            "descricao": clean_text(row.get("descricao_primaria")),
            "categoria": clean_text(row.get("category_label")),
            "unidade": clean_text(row.get("unidade")) or "pc",
        }
        for row in rows
    ]


def _registration_by_sku(sku: str) -> dict[str, Any] | None:
    rows = _request(
        "GET",
        REGISTRATIONS_TABLE,
        [
            ("select", "id,category_key,category_label,sku,descricao_primaria,descricao_secundaria,sufixo,unidade"),
            ("sku", f"eq.{clean_text(sku)}"),
            ("limit", "1"),
        ],
    )
    return rows[0] if rows else None


def _bom_header_by_parent(parent_sku: str) -> dict[str, Any] | None:
    rows = _request(
        "GET",
        BOM_HEADERS_TABLE,
        [
            ("select", "*"),
            ("parent_sku", f"eq.{clean_text(parent_sku)}"),
            ("limit", "1"),
        ],
    )
    return rows[0] if rows else None


def _bom_by_parent(parent_sku: str) -> dict[str, Any] | None:
    header = _bom_header_by_parent(parent_sku)
    if not header:
        return None
    components = _request(
        "GET",
        BOM_COMPONENTS_TABLE,
        [
            ("select", "*"),
            ("bom_id", f"eq.{header['id']}"),
            ("order", "ordem.asc,component_sku.asc"),
            ("limit", "10000"),
        ],
    ) or []
    return {"header": header, "components": components}


def save_bom(
    parent_sku: str,
    parent_description: str,
    components: list[dict[str, Any]],
    category_key: str = "",
    category_label: str = "",
    registration_id: int | str | None = None,
    source: str = "cadastro",
    allow_incomplete: bool = False,
    review_reasons: list[str] | None = None,
) -> dict[str, Any]:
    parent_sku = clean_text(parent_sku)
    if not parent_sku and not allow_incomplete:
        raise SupabaseStoreError("Informe o SKU do item pai da B.O.M.")
    if not components:
        raise SupabaseStoreError("Informe pelo menos um componente para a B.O.M.")

    base_parent_sku = _base_parent_sku(parent_sku)
    registration = _registration_by_sku(base_parent_sku) if base_parent_sku else None
    if registration:
        category_key = clean_text(registration.get("category_key")) or category_key
        category_label = clean_text(registration.get("category_label")) or category_label
        parent_description = _full_description(registration) or parent_description
        registration_id = registration.get("id") or registration_id

    parent_description = clean_text(parent_description) or parent_sku or "B.O.M. pendente de revisao"
    source_value = _source_with_review(source, review_reasons or [])
    payload = {
        "parent_sku": parent_sku,
        "parent_descricao": parent_description,
        "parent_category_key": clean_text(category_key),
        "parent_category_label": clean_text(category_label),
        "registration_id": registration_id,
        "source": source_value,
        "search_text": _search_text(parent_sku, base_parent_sku, parent_description, category_label, source_value),
    }

    existing = _bom_header_by_parent(parent_sku)
    if existing:
        rows = _request(
            "PATCH",
            BOM_HEADERS_TABLE,
            [("id", f"eq.{existing['id']}")],
            payload=payload,
            prefer="return=representation",
        )
    else:
        rows = _request("POST", BOM_HEADERS_TABLE, payload=payload, prefer="return=representation")
    header = rows[0] if rows else {**payload, "id": existing.get("id") if existing else None}
    bom_id = header.get("id")
    if not bom_id:
        raise SupabaseStoreError("Nao foi possivel criar o cabecalho da B.O.M.")

    _request("DELETE", BOM_COMPONENTS_TABLE, [("bom_id", f"eq.{bom_id}")])
    component_catalog = _catalog_data_by_sku(
        [component.get("codigo") or component.get("component_sku") for component in components]
    )
    component_payloads = []
    for index, component in enumerate(components, start=1):
        component_sku = clean_text(component.get("codigo") or component.get("component_sku"))
        if not component_sku and not allow_incomplete:
            continue
        try:
            quantity = float(component.get("quantidade") or component.get("quantity") or (1 if allow_incomplete else 0))
        except Exception as exc:
            if not allow_incomplete:
                raise SupabaseStoreError(f"Quantidade invalida no componente {component_sku}.") from exc
            quantity = 1.0
        if quantity <= 0 and not allow_incomplete:
            raise SupabaseStoreError(f"Quantidade deve ser maior que zero no componente {component_sku}.")
        if quantity <= 0:
            quantity = 1.0
        catalog_item = component_catalog.get(component_sku) or {}
        description = catalog_item.get("descricao_primaria") or clean_text(component.get("descricao") or component.get("component_descricao"))
        unit = catalog_item.get("unidade") or normalize_unit(component.get("unidade") or component.get("unit")) or "pc"
        component_payloads.append(
            {
                "bom_id": bom_id,
                "parent_sku": parent_sku,
                "component_sku": component_sku,
                "component_descricao": description,
                "unidade": unit,
                "quantidade": quantity,
                "ordem": index,
                "search_text": _search_text(parent_sku, base_parent_sku, parent_description, component_sku, description, unit, source_value),
            }
        )
    if not component_payloads:
        raise SupabaseStoreError("Informe pelo menos um componente valido para a B.O.M.")
    _request("POST", BOM_COMPONENTS_TABLE, payload=component_payloads, prefer="return=minimal")
    return {"bom": header, "components_count": len(component_payloads)}


def copy_bom(source_parent_sku: str, target_parent_sku: str) -> dict[str, Any]:
    source_parent_sku = clean_text(source_parent_sku)
    target_parent_sku = clean_text(target_parent_sku)
    if not source_parent_sku:
        raise SupabaseStoreError("Informe o codigo do item pai de origem.")
    if not target_parent_sku:
        raise SupabaseStoreError("Informe o codigo do item pai de destino.")
    if source_parent_sku == target_parent_sku:
        raise SupabaseStoreError("Origem e destino devem ser codigos diferentes.")

    source = _bom_by_parent(source_parent_sku)
    if not source:
        raise SupabaseStoreError("Origem nao possui B.O.M. cadastrada.")
    target_registration = _registration_by_sku(target_parent_sku)
    if not target_registration:
        raise SupabaseStoreError("Destino nao encontrado nos cadastros.")

    components = [
        {
            "codigo": component.get("component_sku"),
            "descricao": component.get("component_descricao"),
            "unidade": component.get("unidade") or "pc",
            "quantidade": component.get("quantidade") or 1,
        }
        for component in source["components"]
    ]
    if not components:
        raise SupabaseStoreError("Origem nao possui componentes para copiar.")

    result = save_bom(
        target_parent_sku,
        _full_description(target_registration) or target_parent_sku,
        components,
        category_key=clean_text(target_registration.get("category_key")),
        category_label=clean_text(target_registration.get("category_label")),
        registration_id=target_registration.get("id"),
        source=f"copia:{source_parent_sku}",
        allow_incomplete=True,
    )
    return {
        "source_parent_sku": source_parent_sku,
        "target_parent_sku": target_parent_sku,
        "components_count": result.get("components_count") or len(components),
        "bom": result.get("bom") or {},
    }


def _in_filter(values: list[Any]) -> str:
    cleaned = [clean_text(value) for value in values if clean_text(value)]
    return "in.(" + ",".join(cleaned) + ")"


def _enrich_bom(
    header: dict[str, Any],
    components: list[dict[str, Any]],
    catalog_data: dict[str, dict[str, str]] | None = None,
) -> dict[str, Any]:
    catalog_data = catalog_data or {}
    source = clean_text(header.get("source"))
    reasons = _review_reasons(source)
    enriched_components = []
    for component in components:
        component_sku = clean_text(component.get("component_sku"))
        component_reasons = []
        if _is_missing_bom_code(component_sku):
            component_reasons.append("component_code")
        if float(component.get("quantidade") or 0) == 1 and "quantity_default" in reasons:
            component_reasons.append("quantity_default")
        component_catalog = catalog_data.get(component_sku) or {}
        component_description = component_catalog.get("descricao_primaria") or clean_text(component.get("component_descricao"))
        component_unit = component_catalog.get("unidade") or normalize_unit(component.get("unidade"))
        enriched_components.append(
            {
                **component,
                "component_descricao": component_description,
                "unidade": component_unit,
                "display_component_sku": _display_bom_code(component_sku),
                "needs_review": bool(component_reasons),
                "review_reasons": [_review_reason_label(reason) for reason in component_reasons],
            }
        )
    if header.get("parent_sku", "").startswith(REVIEW_PARENT_PREFIX) or _is_missing_bom_code(header.get("parent_sku")):
        reasons.append("parent_code")
    if DUPLICATE_PARENT_SEPARATOR in clean_text(header.get("parent_sku")):
        reasons.append("duplicate_parent")
    if any(component.get("needs_review") for component in enriched_components):
        reasons.append("component_code")
    reason_labels = [_review_reason_label(reason) for reason in dict.fromkeys(reasons)]
    parent_sku = clean_text(header.get("parent_sku"))
    parent_catalog = catalog_data.get(_base_parent_sku(parent_sku)) or {}
    parent_description = parent_catalog.get("descricao_primaria") or clean_text(header.get("parent_descricao"))
    return {
        **header,
        "parent_descricao": parent_description,
        "display_parent_sku": _display_bom_code(parent_sku),
        "needs_review": bool(reason_labels),
        "review_reasons": reason_labels,
        "components": enriched_components,
    }


def list_boms(
    category_key: str = "",
    parent_query: str = "",
    component_query: str = "",
    limit: int = 500,
) -> list[dict[str, Any]]:
    bom_ids_filter: list[str] = []
    component_term = _search_text(component_query)
    if component_term:
        component_matches = _request(
            "GET",
            BOM_COMPONENTS_TABLE,
            [
                ("select", "bom_id"),
                ("search_text", f"ilike.*{component_term}*"),
                ("limit", "5000"),
            ],
        ) or []
        bom_ids_filter = list(dict.fromkeys(clean_text(row.get("bom_id")) for row in component_matches if row.get("bom_id")))
        if not bom_ids_filter:
            return []

    params: list[tuple[str, str]] = [
        ("select", "*"),
        ("order", "parent_sku.asc"),
    ]
    if clean_text(category_key):
        category_filter = _category_key_filter("parent_category_key", category_key)
        if category_filter:
            params.append(category_filter)
    parent_term = _search_text(parent_query)
    if parent_term:
        params.append(("search_text", f"ilike.*{parent_term}*"))
    if bom_ids_filter:
        params.append(("id", _in_filter(bom_ids_filter)))
    headers = _request_all(BOM_HEADERS_TABLE, params, limit=max(1, min(limit, 5000)))
    bom_ids = [clean_text(row.get("id")) for row in headers if row.get("id")]
    components_by_bom: dict[str, list[dict[str, Any]]] = {bom_id: [] for bom_id in bom_ids}
    if bom_ids:
        component_params: list[tuple[str, str]] = [
            ("select", "*"),
            ("bom_id", _in_filter(bom_ids)),
            ("order", "parent_sku.asc,ordem.asc,component_sku.asc"),
        ]
        if component_term:
            component_params.append(("search_text", f"ilike.*{component_term}*"))
        components = _request_all(BOM_COMPONENTS_TABLE, component_params, limit=10000)
        for component in components:
            components_by_bom.setdefault(clean_text(component.get("bom_id")), []).append(component)
    description_codes: list[Any] = []
    for header in headers:
        description_codes.append(_base_parent_sku(clean_text(header.get("parent_sku"))))
    for component_list in components_by_bom.values():
        for component in component_list:
            description_codes.append(component.get("component_sku"))
    catalog_data = _catalog_data_by_sku(description_codes)
    return [
        _enrich_bom(header, components_by_bom.get(clean_text(header.get("id")), []), catalog_data)
        for header in headers
    ]


def get_bom(bom_id: int | str) -> dict[str, Any]:
    bom_id = clean_text(bom_id)
    rows = _request("GET", BOM_HEADERS_TABLE, [("select", "*"), ("id", f"eq.{bom_id}"), ("limit", "1")]) or []
    if not rows:
        raise SupabaseStoreError("B.O.M. nao encontrada.")
    components = _request(
        "GET",
        BOM_COMPONENTS_TABLE,
        [
            ("select", "*"),
            ("bom_id", f"eq.{bom_id}"),
            ("order", "ordem.asc,component_sku.asc"),
            ("limit", "10000"),
        ],
    ) or []
    description_codes = [_base_parent_sku(clean_text(rows[0].get("parent_sku")))]
    description_codes.extend(component.get("component_sku") for component in components)
    catalog_data = _catalog_data_by_sku(description_codes)
    return _enrich_bom(rows[0], components, catalog_data)


def update_bom(
    bom_id: int | str,
    parent_description: str,
    components: list[dict[str, Any]],
    parent_sku: str = "",
) -> dict[str, Any]:
    current = get_bom(bom_id)
    current_parent_sku = clean_text(current.get("parent_sku"))
    requested_parent_sku = clean_text(parent_sku)
    if requested_parent_sku and requested_parent_sku == _display_bom_code(current_parent_sku):
        effective_parent_sku = current_parent_sku
    elif requested_parent_sku:
        existing = _bom_header_by_parent(requested_parent_sku)
        if existing and clean_text(existing.get("id")) != clean_text(bom_id):
            raise SupabaseStoreError("Ja existe outra B.O.M. com esse item pai.")
        effective_parent_sku = requested_parent_sku
    else:
        effective_parent_sku = current_parent_sku
    parent_description = clean_text(parent_description) or clean_text(current.get("parent_descricao")) or effective_parent_sku
    if not components:
        raise SupabaseStoreError("Informe pelo menos um componente para a B.O.M.")
    base_parent_sku = _base_parent_sku(effective_parent_sku)
    registration = _registration_by_sku(base_parent_sku) if base_parent_sku else None
    category_key = clean_text(current.get("parent_category_key"))
    category_label = clean_text(current.get("parent_category_label"))
    registration_id = current.get("registration_id")
    if registration:
        category_key = clean_text(registration.get("category_key")) or category_key
        category_label = clean_text(registration.get("category_label")) or category_label
        parent_description = _full_description(registration) or parent_description
        registration_id = registration.get("id") or registration_id

    review_reasons = []
    if effective_parent_sku.startswith(REVIEW_PARENT_PREFIX) or _is_missing_bom_code(effective_parent_sku):
        review_reasons.append("parent_code")
    if DUPLICATE_PARENT_SEPARATOR in effective_parent_sku:
        review_reasons.append("duplicate_parent")
    for component in components:
        if _is_missing_bom_code(component.get("codigo") or component.get("component_sku")):
            review_reasons.append("component_code")
        try:
            quantity_check = float(component.get("quantidade") or component.get("quantity") or 0)
        except Exception:
            quantity_check = 0
        if quantity_check <= 0:
            review_reasons.append("quantity_default")
    source = _source_with_review("edicao", review_reasons)
    header_payload = {
        "parent_sku": effective_parent_sku,
        "parent_descricao": parent_description,
        "parent_category_key": category_key,
        "parent_category_label": category_label,
        "registration_id": registration_id,
        "source": source,
        "search_text": _search_text(effective_parent_sku, base_parent_sku, parent_description, category_label, source),
    }
    rows = _request(
        "PATCH",
        BOM_HEADERS_TABLE,
        [("id", f"eq.{clean_text(bom_id)}")],
        payload=header_payload,
        prefer="return=representation",
    )
    header = rows[0] if rows else {**current, **header_payload}

    _request("DELETE", BOM_COMPONENTS_TABLE, [("bom_id", f"eq.{clean_text(bom_id)}")])
    component_catalog = _catalog_data_by_sku(
        [component.get("codigo") or component.get("component_sku") for component in components]
    )
    component_payloads = []
    for index, component in enumerate(components, start=1):
        component_sku = clean_text(component.get("codigo") or component.get("component_sku"))
        try:
            quantity = float(component.get("quantidade") or component.get("quantity") or 1)
        except Exception:
            quantity = 1.0
        if quantity <= 0:
            quantity = 1.0
        catalog_item = component_catalog.get(component_sku) or {}
        description = catalog_item.get("descricao_primaria") or clean_text(component.get("descricao") or component.get("component_descricao"))
        unit = catalog_item.get("unidade") or normalize_unit(component.get("unidade") or component.get("unit")) or "pc"
        component_payloads.append(
            {
                "bom_id": clean_text(bom_id),
                "parent_sku": effective_parent_sku,
                "component_sku": component_sku,
                "component_descricao": description,
                "unidade": unit,
                "quantidade": quantity,
                "ordem": index,
                "search_text": _search_text(effective_parent_sku, base_parent_sku, parent_description, component_sku, description, unit, source),
            }
        )
    if not component_payloads:
        raise SupabaseStoreError("Informe pelo menos um componente valido para a B.O.M.")
    _request("POST", BOM_COMPONENTS_TABLE, payload=component_payloads, prefer="return=minimal")
    return {**header, "components": component_payloads}


def delete_bom(bom_id: int | str) -> dict[str, Any]:
    bom_id = clean_text(bom_id)
    rows = _request("GET", BOM_HEADERS_TABLE, [("select", "*"), ("id", f"eq.{bom_id}"), ("limit", "1")]) or []
    if not rows:
        raise SupabaseStoreError("B.O.M. nao encontrada.")
    _request("DELETE", BOM_COMPONENTS_TABLE, [("bom_id", f"eq.{bom_id}")])
    _request("DELETE", BOM_HEADERS_TABLE, [("id", f"eq.{bom_id}")])
    return rows[0]


def _normalize_header(value: Any) -> str:
    return excel_bancos.normalize_label(value).replace(" ", "_").lower()


def _parse_quantity_for_import(value: Any) -> tuple[float, bool]:
    text = clean_text(value)
    if _is_missing_bom_code(text):
        return 1.0, True
    try:
        parsed = float(text.replace(".", "").replace(",", ".") if "," in text else text)
    except Exception:
        return 1.0, True
    if parsed <= 0:
        return 1.0, True
    return parsed, False


def _bom_parent_description_from_filename(filename: str) -> str:
    stem = Path(clean_text(filename)).stem
    return re.sub(r"^\s*\d+[\.\-\s]+", "", stem).strip() or stem


def _parse_bom_workbook(content: bytes, filename: str = "") -> dict[str, dict[str, Any]]:
    wb = load_workbook(BytesIO(content), data_only=True)
    parents: dict[str, dict[str, Any]] = {}
    try:
        for ws in wb.worksheets:
            header_row = None
            header_map: dict[str, int] = {}
            for row_index in range(1, min(ws.max_row, 10) + 1):
                headers = {_normalize_header(ws.cell(row_index, col).value): col for col in range(1, ws.max_column + 1)}
                if "item_codigo" in headers and "componente_codigo" in headers:
                    header_row = row_index
                    header_map = headers
                    break
            if not header_row:
                continue
            item_block_col = header_map.get("item_bloco")
            parent_hint = clean_text(ws.cell(header_row, item_block_col + 1).value) if item_block_col else ""
            review_parent_key = _review_parent_key(f"{filename}:{ws.title}")
            for row_index in range(header_row + 1, ws.max_row + 1):
                raw_parent_sku = clean_text(ws.cell(row_index, header_map["item_codigo"]).value)
                row_parent_description = clean_text(
                    ws.cell(row_index, header_map.get("item_descricao", 0)).value
                ) if header_map.get("item_descricao") else ""
                raw_component_sku = clean_text(ws.cell(row_index, header_map["componente_codigo"]).value)
                description = clean_text(ws.cell(row_index, header_map.get("descricao", 3)).value)
                unit = clean_text(ws.cell(row_index, header_map.get("unidade", 4)).value) or "pc"
                quantity_value = ws.cell(row_index, header_map.get("quantidade", 5)).value
                if not any([raw_parent_sku, raw_component_sku, description, clean_text(quantity_value)]):
                    continue
                reasons: list[str] = []
                parent_missing = _is_missing_bom_code(raw_parent_sku)
                component_missing = _is_missing_bom_code(raw_component_sku)
                parent_sku = review_parent_key if parent_missing else raw_parent_sku
                component_sku = "" if component_missing else raw_component_sku
                if parent_missing:
                    reasons.append("parent_code")
                if component_missing:
                    reasons.append("component_code")
                quantity, quantity_defaulted = _parse_quantity_for_import(quantity_value)
                if quantity_defaulted:
                    reasons.append("quantity_default")
                if component_missing and not description:
                    reasons.append("empty_component")
                group = parents.setdefault(
                    parent_sku,
                    {
                        "parent_description": (
                            row_parent_description
                            or parent_hint
                            or (description if parent_missing else "")
                            or _bom_parent_description_from_filename(filename)
                            or parent_sku
                        ),
                        "components": [],
                        "review_reasons": [],
                    },
                )
                group["review_reasons"].extend(reasons)
                group["components"].append(
                    {
                        "codigo": component_sku,
                        "descricao": description,
                        "unidade": unit,
                        "quantidade": quantity,
                    }
                )
    finally:
        wb.close()
    return parents


def template_bom_xlsx() -> bytes:
    return build_import_template(
        "BOM",
        [
            {"header": "item_codigo", "required": True, "description": "SKU do item pai. Repita em todas as linhas da estrutura.", "example": "30180001", "width": 18},
            {"header": "item_descricao", "required": False, "description": "Descricao primaria do item pai.", "example": "CJ EXEMPLO", "width": 54},
            {"header": "componente_codigo", "required": True, "description": "SKU do componente.", "example": "10180001", "width": 22},
            {"header": "descricao", "required": False, "description": "Descricao do componente. O cadastro do SKU tem prioridade.", "example": "COMPONENTE EXEMPLO", "width": 64},
            {"header": "unidade", "required": False, "description": "Unidade de medida. O cadastro do SKU tem prioridade.", "example": "pc", "width": 16},
            {"header": "quantidade", "required": True, "description": "Consumo numerico maior que zero.", "example": "1", "width": 18},
        ],
        warning=(
            "ATENCAO: ao importar um item_codigo ja existente, a composicao atual desse item sera substituida. "
            "Use uma linha por componente e repita o item pai."
        ),
    )


def import_bom_workbook(content: bytes, filename: str = "") -> dict[str, Any]:
    imported = 0
    parents = _parse_bom_workbook(content, filename)
    for parent_sku, data in parents.items():
        reasons = list(dict.fromkeys(data.get("review_reasons") or []))
        save_bom(
            parent_sku,
            data["parent_description"],
            data["components"],
            source=f"import:{filename or 'xlsx'}",
            allow_incomplete=bool(reasons),
            review_reasons=reasons,
        )
        imported += 1
    return {
        "parents": imported,
        "components": sum(len(item["components"]) for item in parents.values()),
        "review_parents": sum(1 for item in parents.values() if item.get("review_reasons")),
    }


def import_bom_directory(directory: str | Path) -> dict[str, Any]:
    root = Path(directory)
    if not root.exists():
        raise SupabaseStoreError(f"Diretorio de B.O.M. nao encontrado: {root}")
    files = sorted(path for path in root.rglob("*") if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"} and not path.name.startswith("~$"))
    parsed_entries: list[dict[str, Any]] = []
    result = {"files": len(files), "parents": 0, "components": 0, "review_parents": 0, "duplicate_parents": 0, "errors": []}
    for path in files:
        relative = str(path.relative_to(root))
        try:
            parents = _parse_bom_workbook(path.read_bytes(), relative)
            for parent_sku, data in parents.items():
                parsed_entries.append({"file": relative, "parent_sku": parent_sku, "data": data})
        except Exception as exc:
            result["errors"].append({"file": relative, "error": str(exc)})

    parent_counts: dict[str, int] = {}
    for entry in parsed_entries:
        parent_counts[entry["parent_sku"]] = parent_counts.get(entry["parent_sku"], 0) + 1
    seen_parent: dict[str, int] = {}
    for entry in parsed_entries:
        original_parent_sku = entry["parent_sku"]
        data = entry["data"]
        reasons = list(dict.fromkeys(data.get("review_reasons") or []))
        storage_parent_sku = original_parent_sku
        if parent_counts.get(original_parent_sku, 0) > 1:
            seen_parent[original_parent_sku] = seen_parent.get(original_parent_sku, 0) + 1
            if seen_parent[original_parent_sku] > 1:
                storage_parent_sku = _duplicate_parent_key(original_parent_sku, entry["file"])
                result["duplicate_parents"] += 1
            reasons.append("duplicate_parent")
        save_bom(
            storage_parent_sku,
            data["parent_description"],
            data["components"],
            source=f"import:{entry['file']}",
            allow_incomplete=bool(reasons),
            review_reasons=reasons,
        )
        result["parents"] += 1
        result["components"] += len(data["components"])
        if reasons:
            result["review_parents"] += 1
    return result


def export_boms(category_key: str = "", parent_query: str = "", component_query: str = "") -> Path:
    rows = list_boms(category_key=category_key, parent_query=parent_query, component_query=component_query, limit=5000)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output = EXPORT_DIR / f"bom_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    headers = [
        "categoria",
        "item_codigo",
        "item_descricao",
        "componente_codigo",
        "descricao",
        "unidade",
        "quantidade",
    ]
    ws.append(headers)
    for bom in rows:
        for component in bom.get("components") or []:
            quantity = component.get("quantidade")
            try:
                quantity = int(quantity) if float(quantity).is_integer() else float(quantity)
            except Exception:
                pass
            ws.append(
                [
                    bom.get("parent_category_label"),
                    bom.get("display_parent_sku") if "display_parent_sku" in bom else _display_bom_code(bom.get("parent_sku")),
                    bom.get("parent_descricao"),
                    component.get("display_component_sku") if "display_component_sku" in component else _display_bom_code(component.get("component_sku")),
                    component.get("component_descricao"),
                    component.get("unidade"),
                    quantity,
                ]
            )
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for column, width in {"A": 24, "B": 16, "C": 54, "D": 20, "E": 72, "F": 12, "G": 14}.items():
        ws.column_dimensions[column].width = width
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(output)
    wb.close()
    return output


def export_registrations(
    category_key: str,
    query: str = "",
    filters: dict[str, str] | None = None,
    group_code: str = "",
    missing_unit: bool = False,
    include_inactive: bool = False,
) -> Path:
    all_categories = all_categories_key(category_key)
    category = {"key": ALL_CATEGORIES_KEY, "label": "Todas as categorias"} if all_categories else _category(category_key)
    fields = [] if all_categories else excel_bancos.get_banco_fields_for_display(category["key"], group_code)
    rows = list_registrations(
        ALL_CATEGORIES_KEY if all_categories else category["key"],
        query=query,
        filters=filters,
        group_code=group_code,
        missing_unit=missing_unit,
        include_inactive=include_inactive,
        limit=10000,
    )
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output = EXPORT_DIR / f"cadastros_{category['key']}_{stamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = ("Todos Cadastros" if all_categories else _sheet_name(category))[:31]
    headers = [
        "CATEGORIA",
        "SKU",
        "DESCRIÇÃO PRIMÁRIA",
        "DESCRIÇÃO SECUNDÁRIA",
        "UNIDADE",
        "STATUS",
        "SUFIXO",
        "CARACTERES PRIMARIO",
        "CARACTERES SECUNDARIO",
    ]
    headers.insert(2, "GRUPO")
    if all_categories:
        headers.append("CAMPOS")
    else:
        headers.extend(excel_bancos.header_for_field(field["label"], field["scope"]) for field in fields)
    ws.append(headers)
    for row in rows:
        values = row.get("field_values") if isinstance(row.get("field_values"), dict) else {}
        row_values = [
            row.get("category_label"),
            row.get("sku"),
            f"{row.get('grupo_codigo') or ''} - {row.get('grupo_label') or ''}".strip(" -"),
            row.get("descricao_primaria"),
            row.get("descricao_secundaria"),
            row.get("unidade"),
            "ATIVO" if row.get("ativo", True) else "INATIVO",
            row.get("sufixo"),
            row.get("caracteres_primario"),
            row.get("caracteres_secundario"),
        ]
        if all_categories:
            row_values.append(" | ".join(f"{key}: {value}" for key, value in values.items() if clean_text(value)))
        else:
            row_values.extend(values.get(field["key"], "") for field in fields)
        ws.append(row_values)

    header_fill = PatternFill("solid", fgColor="E2E8F0")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    widths = {1: 24, 2: 14, 3: 22, 4: 48, 5: 72, 6: 14, 7: 16, 8: 18, 9: 18, 10: 22, 11: 72}
    for index in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(1, index).column_letter].width = widths.get(index, 28)
    for row_cells in ws.iter_rows(min_row=2):
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"
    wb.save(output)
    wb.close()
    return output

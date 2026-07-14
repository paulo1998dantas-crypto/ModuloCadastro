import argparse
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import excel_bancos
import supabase_store


DEFAULT_WORKBOOK = (
    r"C:\Users\PRODUCAO-2.0\J I MONTADORA DE VEICULOS ESPECIAIS LTDA"
    r"\JI Montadora - 02 Produção\01 Controle de Produção"
    r"\01 - Projeto Cadastro\01 - Gerador cadastros.xlsx"
)


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _header_map(ws, header_row: int = 2) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        normalized = excel_bancos.normalize_label(ws.cell(header_row, column).value)
        if normalized and normalized not in mapping:
            mapping[normalized] = column
    return mapping


def _cell(ws, row: int, column: int | None) -> str:
    if not column:
        return ""
    return clean_text(ws.cell(row, column).value)


def _exists(category_key: str, sku: str) -> bool:
    rows = supabase_store._request(
        "GET",
        supabase_store.REGISTRATIONS_TABLE,
        [
            ("select", "id"),
            ("category_key", f"eq.{category_key}"),
            ("sku", f"eq.{sku}"),
            ("limit", "1"),
        ],
    )
    return bool(rows)


def _insert(payload: dict[str, Any]) -> None:
    supabase_store._request(
        "POST",
        supabase_store.REGISTRATIONS_TABLE,
        payload=payload,
        prefer="return=minimal",
    )


def import_workbook(workbook_path: Path, dry_run: bool = False) -> dict[str, int]:
    if not dry_run and not supabase_store.enabled():
        raise SystemExit("Defina CADASTRO_SAVE_MODE=supabase antes de importar.")
    if not workbook_path.exists():
        raise SystemExit(f"Planilha não encontrada: {workbook_path}")

    catalog = excel_bancos.load_catalog()
    wb = load_workbook(workbook_path, data_only=False, read_only=False)
    inserted = 0
    skipped = 0
    try:
        for category in catalog.get("categories") or []:
            sheet_name = excel_bancos._safe_sheet_title(category.get("sheet_name") or category.get("label"))
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            headers = _header_map(ws)
            sku_col = headers.get("SKU")
            primary_col = headers.get(excel_bancos.normalize_label("DESCRIÇÃO PRIMÁRIA"))
            secondary_col = headers.get(excel_bancos.normalize_label("DESCRIÇÃO SECUNDÁRIA"))
            suffix_col = headers.get(excel_bancos.normalize_label("SUFIXO"))
            if not sku_col:
                continue

            fields = excel_bancos.get_banco_fields_for_display(category["key"])
            field_columns = {
                field["key"]: headers.get(excel_bancos.normalize_label(excel_bancos.header_for_field(field["label"], field["scope"])))
                for field in fields
            }

            for row in range(3, ws.max_row + 1):
                sku = _cell(ws, row, sku_col)
                if not sku:
                    continue
                if not dry_run and _exists(category["key"], sku):
                    skipped += 1
                    continue
                field_values = {field["key"]: _cell(ws, row, field_columns.get(field["key"])) for field in fields}
                primaria = _cell(ws, row, primary_col)
                secundaria = _cell(ws, row, secondary_col)
                payload = {
                    "category_key": category["key"],
                    "category_label": category["label"],
                    "sheet": sheet_name,
                    "sku": sku,
                    "descricao_primaria": primaria,
                    "descricao_secundaria": secundaria,
                    "sufixo": _cell(ws, row, suffix_col),
                    "caracteres_primario": len(primaria),
                    "caracteres_secundario": len(secundaria),
                    "form_values": {},
                    "field_values": field_values,
                    "field_codes": {},
                    "search_text": supabase_store._search_text(
                        sku,
                        category["label"],
                        primaria,
                        secundaria,
                        " ".join(field_values.values()),
                    ),
                }
                if not dry_run:
                    _insert(payload)
                inserted += 1
    finally:
        wb.close()
    return {"inserted": inserted, "skipped": skipped}


def main() -> None:
    parser = argparse.ArgumentParser(description="Importa a planilha atual para as tabelas cadastro_* no Supabase.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Caminho da planilha Gerador cadastros.xlsx.")
    parser.add_argument("--dry-run", action="store_true", help="Conta registros sem inserir.")
    args = parser.parse_args()
    result = import_workbook(Path(args.workbook), dry_run=args.dry_run)
    print(f"Inseridos: {result['inserted']}")
    print(f"Ignorados por SKU existente: {result['skipped']}")


if __name__ == "__main__":
    main()

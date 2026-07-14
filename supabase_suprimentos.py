import json
import re
import unicodedata
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

import supabase_store


PESSOAS_TABLE = "suprimentos_pessoas"
PROCESSOS_TABLE = "suprimentos_processos"
REGRAS_TABLE = "suprimentos_regras_popup_item"
RELACOES_TABLE = "suprimentos_relacoes_processo_item"
PAGE_SIZE = 1000


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def configured() -> bool:
    return supabase_store.configured()


def _request(method: str, table: str, query=None, payload=None, prefer: str = ""):
    return supabase_store._request(method, table, query=query or [], payload=payload, prefer=prefer)


def _all_rows(table: str, select: str = "*", order: str = "") -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    offset = 0
    while True:
        query = [("select", select), ("limit", str(PAGE_SIZE)), ("offset", str(offset))]
        if order:
            query.append(("order", order))
        page = _request("GET", table, query=query) or []
        rows.extend(page)
        if len(page) < PAGE_SIZE:
            break
        offset += PAGE_SIZE
    return rows


def _bool(value: Any) -> bool:
    text = clean_text(value).lower()
    return text in {"1", "sim", "s", "true", "verdadeiro", "x", "yes"}


def _blank_dash(value: Any) -> str:
    text = clean_text(value)
    return "" if text in {"---", "-", "None", "none", "NULL", "null"} else text


def _numeric(value: Any) -> float:
    text = _blank_dash(value)
    if isinstance(value, str) and "," in text:
        text = text.replace(".", "").replace(",", ".")
    if not text:
        return 0
    try:
        return float(text)
    except Exception:
        return 0


def _date_or_none(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = _blank_dash(value)
    return text or None


def _search_text(*parts: Any) -> str:
    return " ".join(clean_text(part) for part in parts if clean_text(part))


def _key_from_pessoa(pessoa: dict[str, Any]) -> str:
    return (
        clean_text(pessoa.get("identificador"))
        or clean_text(pessoa.get("cnpj_cpf"))
        or clean_text(pessoa.get("nome_fantasia"))
        or clean_text(pessoa.get("razao_social"))
    )


def normalizar_pessoa(pessoa: dict[str, Any]) -> dict[str, Any]:
    row = dict(pessoa or {})
    for key, value in list(row.items()):
        if isinstance(value, datetime):
            row[key] = value.isoformat()
        elif isinstance(value, str):
            row[key] = _blank_dash(value)
    row["identificador"] = _key_from_pessoa(row)
    row["pessoa_fisica"] = _bool(row.get("pessoa_fisica"))
    row["cliente"] = _bool(row.get("cliente"))
    row["fornecedor"] = _bool(row.get("fornecedor"))
    row["colaborador"] = _bool(row.get("colaborador"))
    row["transportadora"] = _bool(row.get("transportadora"))
    for key in ("limite_credito", "periodicidade_venda_compra_dias", "valor_minimo_compra"):
        row[key] = _numeric(row.get(key))
    row["data_nascimento_fundacao"] = _date_or_none(row.get("data_nascimento_fundacao"))
    row["search_text"] = _search_text(
        row.get("nome_fantasia"),
        row.get("razao_social"),
        row.get("cnpj_cpf"),
        row.get("email"),
        row.get("telefone"),
        row.get("cidade"),
        row.get("uf"),
        row.get("identificador"),
    )
    row.setdefault("payload", {})
    return row


def salvar_pessoas(pessoas: list[dict[str, Any]]) -> int:
    rows = [normalizar_pessoa(pessoa) for pessoa in pessoas if _key_from_pessoa(pessoa)]
    if not rows:
        return 0
    _request(
        "POST",
        PESSOAS_TABLE,
        query=[("on_conflict", "identificador")],
        payload=rows,
        prefer="resolution=merge-duplicates,return=minimal",
    )
    return len(rows)


def listar_pessoas(limit: int = 80) -> list[dict[str, Any]]:
    return _all_rows(PESSOAS_TABLE, order="nome_fantasia.asc")[:limit]


def listar_processos() -> dict[str, dict[str, list[dict[str, str]]]]:
    rows = _all_rows(
        PROCESSOS_TABLE,
        select="conjunto,processo,ordem,atividade,responsavel",
        order="conjunto.asc,processo.asc,ordem.asc",
    )
    processos: dict[str, dict[str, list[dict[str, str]]]] = {}
    for row in rows:
        conjunto = clean_text(row.get("conjunto")) or "PADRAO"
        processo = clean_text(row.get("processo"))
        atividade = clean_text(row.get("atividade"))
        if not processo or not atividade:
            continue
        processos.setdefault(conjunto, {}).setdefault(processo, []).append(
            {"atividade": atividade, "responsavel": clean_text(row.get("responsavel"))}
        )
    return processos


def salvar_processos(processos: dict[str, dict[str, list[dict[str, str]]]]) -> int:
    rows = []
    for conjunto, por_processo in (processos or {}).items():
        conjunto = clean_text(conjunto) or "PADRAO"
        for processo, linhas in (por_processo or {}).items():
            processo = clean_text(processo)
            for ordem, linha in enumerate(linhas or [], 1):
                atividade = clean_text((linha or {}).get("atividade"))
                if not processo or not atividade:
                    continue
                rows.append(
                    {
                        "conjunto": conjunto,
                        "processo": processo,
                        "ordem": ordem,
                        "atividade": atividade,
                        "responsavel": clean_text((linha or {}).get("responsavel")),
                        "search_text": _search_text(conjunto, processo, atividade, (linha or {}).get("responsavel")),
                    }
                )
    _request("DELETE", PROCESSOS_TABLE, query=[("conjunto", "neq.__never_delete__")])
    if rows:
        _request("POST", PROCESSOS_TABLE, payload=rows, prefer="return=minimal")
    return len(rows)


def listar_regras() -> list[dict[str, Any]]:
    rows = _all_rows(REGRAS_TABLE, select="rule_id,gatilho,opcoes,quantidade,quantidade_editavel", order="rule_id.asc")
    return [
        {
            "id": clean_text(row.get("rule_id")),
            "gatilho": clean_text(row.get("gatilho")),
            "opcoes": row.get("opcoes") or [],
            "quantidade": row.get("quantidade") or 1,
            "quantidade_editavel": bool(row.get("quantidade_editavel")),
        }
        for row in rows
        if clean_text(row.get("rule_id")) and clean_text(row.get("gatilho"))
    ]


def salvar_regras(regras: list[dict[str, Any]]) -> int:
    rows = []
    for regra in regras or []:
        rule_id = clean_text(regra.get("id") or regra.get("rule_id"))
        gatilho = clean_text(regra.get("gatilho"))
        opcoes = [clean_text(opcao) for opcao in (regra.get("opcoes") or []) if clean_text(opcao)]
        if not rule_id or not gatilho or not opcoes:
            continue
        rows.append(
            {
                "rule_id": rule_id,
                "gatilho": gatilho,
                "opcoes": opcoes,
                "quantidade": regra.get("quantidade") or 1,
                "quantidade_editavel": bool(regra.get("quantidade_editavel")),
            }
        )
    _request("DELETE", REGRAS_TABLE, query=[("rule_id", "neq.__never_delete__")])
    if rows:
        _request("POST", REGRAS_TABLE, payload=rows, prefer="return=minimal")
    return len(rows)


def listar_relacoes() -> dict[str, list[str]]:
    rows = _all_rows(RELACOES_TABLE, select="item_codigo,processos", order="item_codigo.asc")
    return {
        clean_text(row.get("item_codigo")): [
            clean_text(processo) for processo in (row.get("processos") or []) if clean_text(processo)
        ]
        for row in rows
        if clean_text(row.get("item_codigo"))
    }


def salvar_relacoes(relacoes: dict[str, list[str]]) -> int:
    rows = []
    for codigo, processos in (relacoes or {}).items():
        codigo = clean_text(codigo)
        processos = [clean_text(processo) for processo in (processos or []) if clean_text(processo)]
        if codigo and processos:
            rows.append({"item_codigo": codigo, "processos": list(dict.fromkeys(processos))})
    _request("DELETE", RELACOES_TABLE, query=[("item_codigo", "neq.__never_delete__")])
    if rows:
        _request("POST", RELACOES_TABLE, payload=rows, prefer="return=minimal")
    return len(rows)


def _normalizar_header(texto: Any) -> str:
    texto = unicodedata.normalize("NFKD", clean_text(texto).lower())
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"[^a-z0-9]+", "_", texto)
    return texto.strip("_")


def _linhas_xlsx(content: bytes) -> tuple[list[str], list[list[Any]]]:
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [list(row) for row in ws.iter_rows(values_only=True) if any(row)]
    if not rows:
        return [], []
    headers = [_normalizar_header(value) for value in rows[0]]
    return headers, rows[1:]


PESSOA_FIELDS = {
    "data_registro": ("data_de_registro", "data_registro"),
    "pessoa_fisica": ("pessoafisica", "pessoa_fisica"),
    "nome_fantasia": ("nomefantasia", "nome_fantasia", "fornecedor", "cliente"),
    "razao_social": ("razaosocial", "razao_social"),
    "cnpj_cpf": ("cnpj_cpf", "cnpj", "cpf"),
    "codigo_identificador_unico": ("codigo_identificador_unico",),
    "rg": ("rg",),
    "ie": ("ie",),
    "logradouro": ("logradouro", "endereco"),
    "logradouro_numero": ("logradouronumero", "logradouro_numero", "numero"),
    "complemento": ("complemento",),
    "bairro": ("bairro",),
    "cidade": ("cidade",),
    "codigo_municipio": ("codigomunicipio", "codigo_municipio"),
    "pais": ("pais",),
    "codigo_pais": ("codigopais", "codigo_pais"),
    "cep": ("cep",),
    "uf": ("uf",),
    "codigo_uf": ("codigouf", "codigo_uf"),
    "telefone": ("telefone",),
    "whatsapp": ("whatsapp",),
    "celular": ("celular",),
    "email": ("email",),
    "site": ("site",),
    "cliente": ("cliente",),
    "fornecedor": ("fornecedor",),
    "colaborador": ("colaborador",),
    "transportadora": ("transportadora",),
    "pessoa_grupo": ("pessoagrupo", "pessoa_grupo"),
    "identificador": ("identificador",),
    "vendedor_padrao": ("vendedorpadrao", "vendedor_padrao"),
    "categoria": ("categoria",),
    "tabela_preco": ("tabelapreco", "tabela_preco"),
    "observacoes": ("observacoes", "observacao"),
    "limite_credito": ("limite_de_credito", "limite_credito"),
    "periodicidade_venda_compra_dias": ("periodicidade_venda_compra_dias", "periodicidade_venda_compra"),
    "validation": ("validation",),
    "valor_minimo_compra": ("valorminimocompra", "valor_minimo_compra"),
    "data_nascimento_fundacao": ("datanascimentofundacao", "data_nascimento_fundacao"),
}


def _valor(row: list[Any], mapa: dict[str, int], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        idx = mapa.get(alias)
        if idx is not None and idx < len(row):
            value = row[idx]
            if clean_text(value):
                return value
    return ""


def importar_pessoas_xlsx(content: bytes) -> int:
    headers, rows = _linhas_xlsx(content)
    mapa = {name: idx for idx, name in enumerate(headers) if name}
    pessoas = []
    for row in rows:
        pessoa = {field: _valor(row, mapa, aliases) for field, aliases in PESSOA_FIELDS.items()}
        if any(pessoa.get(field) for field in ("nome_fantasia", "razao_social", "cnpj_cpf", "identificador")):
            pessoa["payload"] = {
                headers[idx]: clean_text(value)
                for idx, value in enumerate(row)
                if idx < len(headers) and headers[idx] and clean_text(value)
            }
            pessoas.append(pessoa)
    return salvar_pessoas(pessoas)


def importar_processos_xlsx(content: bytes, filename: str = "") -> int:
    headers, rows = _linhas_xlsx(content)
    if not headers:
        return 0
    mapa = {name: idx for idx, name in enumerate(headers) if name}
    processos = listar_processos()
    conjunto_padrao = clean_text(filename).rsplit(".", 1)[0] or "PADRAO"
    count = 0
    if "processo" in mapa:
        staged: dict[str, dict[str, list[dict[str, str]]]] = {}
        for row in rows:
            conjunto = clean_text(_valor(row, mapa, ("conjunto", "grupo"))) or conjunto_padrao
            processo = clean_text(_valor(row, mapa, ("processo",))).upper()
            atividade = clean_text(_valor(row, mapa, ("atividade",)))
            responsavel = clean_text(_valor(row, mapa, ("responsavel",)))
            if not processo or not atividade:
                continue
            staged.setdefault(conjunto, {}).setdefault(processo, []).append(
                {"atividade": atividade, "responsavel": responsavel}
            )
            count += 1
        for conjunto, por_processo in staged.items():
            processos.setdefault(conjunto, {}).update(por_processo)
        salvar_processos(processos)
        return count
    conjunto = conjunto_padrao
    processos.setdefault(conjunto, {})
    for col_idx, processo in enumerate(headers):
        processo_nome = clean_text(processo).upper()
        if not processo_nome:
            continue
        linhas = []
        for row in rows:
            if col_idx >= len(row):
                continue
            atividade = clean_text(row[col_idx])
            if atividade:
                linhas.append({"atividade": atividade, "responsavel": ""})
        if linhas:
            processos[conjunto][processo_nome] = linhas
            count += len(linhas)
    salvar_processos(processos)
    return count


def importar_regras_xlsx(content: bytes) -> int:
    headers, rows = _linhas_xlsx(content)
    mapa = {name: idx for idx, name in enumerate(headers) if name}
    regras = listar_regras()
    current = {regra["id"]: regra for regra in regras}
    next_id = max([int(re.sub(r"\D", "", rid) or 0) for rid in current] or [0]) + 1
    for row in rows:
        gatilho = clean_text(_valor(row, mapa, ("item_gatilho", "gatilho")))
        opcoes_text = clean_text(_valor(row, mapa, ("itens_opcoes", "opcoes", "item_opcao")))
        if not gatilho or not opcoes_text:
            continue
        rule_id = clean_text(_valor(row, mapa, ("id_regra", "id"))) or f"regra-{next_id}"
        next_id += 1
        opcoes = [clean_text(part) for part in re.split(r"[;\r\n]+", opcoes_text) if clean_text(part)]
        current[rule_id] = {
            "id": rule_id,
            "gatilho": gatilho,
            "opcoes": list(dict.fromkeys(opcoes)),
            "quantidade": _numeric(_valor(row, mapa, ("quantidade",))) or 1,
            "quantidade_editavel": _bool(_valor(row, mapa, ("quantidade_editavel", "alteravel"))),
        }
    return salvar_regras(list(current.values()))


def importar_relacoes_xlsx(content: bytes) -> int:
    headers, rows = _linhas_xlsx(content)
    mapa = {name: idx for idx, name in enumerate(headers) if name}
    relacoes = listar_relacoes()
    for row in rows:
        codigo = clean_text(_valor(row, mapa, ("item_codigo", "codigo", "item")))
        processos_text = clean_text(_valor(row, mapa, ("processos", "processo", "processo_conjunto")))
        if not codigo or not processos_text:
            continue
        processos = [clean_text(part) for part in re.split(r"[;\r\n]+", processos_text) if clean_text(part)]
        relacoes[codigo] = list(dict.fromkeys(processos))
    return salvar_relacoes(relacoes)

"""Generate the auditable SQL migration for the approved VIDROS workbook.

The workbook is the source for the 156 catalogued glass SKUs.  The generated
migration only updates the descriptive payload used by ModuloCadastro; it does
not alter stock, B.O.M. structure, units, activation state or SKU numbers.
"""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


CATEGORY_KEY = "cat_12_vidros"
CATEGORY_LABEL = "12 - VIDROS"
FIELDS = (
    "prefixo",
    "descritor_base",
    "veiculo_modelo",
    "posicao_lado",
    "medida",
    "material_cor",
    "complemento_regra",
    "fornecedor_referencia",
    "espessura",
    "fornecedor",
    "especificidade",
)
OPTIONS = {
    "prefixo": ("1- JANELA", "2- VIDRO FIXO", "3- VIDRO FIXO C/ JANELA", "4- CJ"),
    "descritor_base": ("1- SELADO", "2- VISTA"),
    "veiculo_modelo": ("1- LD", "2- LE"),
    "posicao_lado": (
        "1- B/J/D", "2- E/S/J", "3- MASTER", "4- MASTER L1H1", "5- MASTER L2H2",
        "6- MASTER L2H2/L3H2", "7- MASTER L3H2", "8- SPRINTER", "9- SPRINTER 10 M",
        "10- SPRINTER 10/14 M", "11- SPRINTER 14 M", "12- TRANSIT", "13- TRANSIT L2H3",
        "14- TRANSIT L3H3",
    ),
    "medida": ("1- 1 VAO", "2- 1 VAO LD/LE / 2 VAO LE", "3- 2 VAO", "4- 3 VAO", "5- VIGIA"),
    "material_cor": ("1- VERDE",),
    "complemento_regra": tuple(f"{index}- {value}" for index, value in enumerate(
        (671, 676, 686, 743, 771, 792, 814, 815, 830, 838, 923, 976, 1078, 1082, 1089, 1096,
         1171, 1177, 1195, 1210, 1291, 1352, 1357, 1365, 1373, 1383, 1389, 1400, 1404, 1418,
         1422, 1425, 1431, 1437, 1637, 1664, 1666),
        start=1,
    )),
    "fornecedor_referencia": tuple(f"{index}- {value}" for index, value in enumerate(
        (495, 534, 556, 583, 630, 667, 686, 705, 716, 728, 730, 738, 739, 741, 761, 766, 769,
         771, 793, 798, 800, 802, 806, 814, 830, 833),
        start=1,
    )),
    "espessura": ("1- 3", "2- 4", "3- 5"),
    "fornecedor": ("1- SALT", "2- STY", "3- VF", "4- VTRX"),
    "especificidade": ("1- 2º VAO LD VIDRO FIXO", "2- N/A"),
}


def text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def identity(value: object) -> str:
    value = unicodedata.normalize("NFKD", text(value).upper())
    value = "".join(character for character in value if not unicodedata.combining(character))
    return re.sub(r"[^A-Z0-9]+", "", value)


def option_label(value: str) -> str:
    return re.sub(r"^\s*\d+\s*-\s*", "", value).strip()


def option_code(value: str) -> str:
    match = re.match(r"^\s*(\d+)\s*-", value)
    return match.group(1) if match else ""


def canonical(field: str, value: object) -> str:
    raw = text(value)
    if not raw:
        return ""
    matches = [option for option in OPTIONS[field] if identity(option_label(option)) == identity(raw)]
    if len(matches) != 1:
        raise ValueError(f"Valor sem correspondencia unica: {field}={raw!r}; candidatos={matches!r}")
    return matches[0]


def display(value: str) -> str:
    return option_label(value) if value else ""


def search_text(*parts: str) -> str:
    return " ".join(identity(part) for part in parts if text(part))


def build_row(values: tuple[object, ...]) -> dict[str, object]:
    sku, prefixo, comercial, lado, veiculo, local, cor, comprimento, largura, espessura, fornecedor, especificidade = values
    sku = text(sku)
    if not re.fullmatch(r"[13]012\d{4}", sku):
        raise ValueError(f"SKU invalido: {sku!r}")
    selected = {
        "prefixo": canonical("prefixo", prefixo),
        "descritor_base": canonical("descritor_base", comercial),
        "veiculo_modelo": canonical("veiculo_modelo", lado),
        "posicao_lado": canonical("posicao_lado", veiculo),
        "medida": canonical("medida", local),
        "material_cor": canonical("material_cor", cor),
        "complemento_regra": canonical("complemento_regra", comprimento),
        "fornecedor_referencia": canonical("fornecedor_referencia", largura),
        "espessura": canonical("espessura", espessura),
        "fornecedor": canonical("fornecedor", fornecedor),
        "especificidade": canonical("especificidade", especificidade),
    }
    dimension = "X".join(
        display(selected[field]) for field in ("complemento_regra", "fornecedor_referencia", "espessura")
        if selected[field]
    )
    primary_parts = [
        display(selected["prefixo"]),
        display(selected["descritor_base"]),
        display(selected["veiculo_modelo"]),
        dimension,
        display(selected["posicao_lado"]),
        display(selected["medida"]),
        display(selected["fornecedor"]),
        display(selected["especificidade"]),
    ]
    primary = " ".join(part for part in primary_parts if part)
    secondary_parts = [display(selected["material_cor"])]
    secondary_detail = " ".join(part for part in secondary_parts if part)
    secondary = f"{primary} {secondary_detail}".strip() if secondary_detail else primary
    group = "30" if sku.startswith("3012") else "10"
    form_values = {"pn_grupo_codigo": [group], **{key: ([value] if value else []) for key, value in selected.items()}, "possui_bom": group == "30"}
    field_values = {key: value for key, value in selected.items()}
    field_codes = {key: option_code(value) if value else "" for key, value in selected.items()}
    return {
        "sku": sku,
        "descricao_primaria": primary,
        "descricao_secundaria": secondary,
        "form_values": form_values,
        "field_values": field_values,
        "field_codes": field_codes,
        "search_text": search_text(sku, CATEGORY_LABEL, primary, secondary, "cj" if group == "30" else "pc", *field_values.values()),
    }


def sql_literal(value: object) -> str:
    return "'" + json.dumps(value, ensure_ascii=False, separators=(",", ":")).replace("'", "''") + "'"


def render_sql(rows: list[dict[str, object]], workbook_name: str) -> str:
    payload = []
    for row in rows:
        values = row["field_values"]
        payload.append(
            {
                "sku": row["sku"],
                "prefixo": display(values["prefixo"]),
                "comercial": display(values["descritor_base"]),
                "lado": display(values["veiculo_modelo"]),
                "veiculo": display(values["posicao_lado"]),
                "local": display(values["medida"]),
                "cor": display(values["material_cor"]),
                "comprimento": display(values["complemento_regra"]),
                "largura": display(values["fornecedor_referencia"]),
                "espessura": display(values["espessura"]),
                "fornecedor": display(values["fornecedor"]),
                "especificidade": display(values["especificidade"]),
            }
        )
    source = sql_literal(payload)
    return f"""-- Reconciliacao aditiva dos cadastros 12 - VIDROS a partir de {workbook_name}.
-- Cada SKU foi normalizado contra as opcoes vigentes e descrito com a mesma
-- regra usada no formulario: COMPRIMENTOXLARGURAXESPESSURA.
-- Nao altera SKU, unidade, ativo, B.O.M., estoque nem relacionamentos.
with source as (
    select * from jsonb_to_recordset({source}::jsonb) as row(
        sku text, prefixo text, comercial text, lado text, veiculo text, local text,
        cor text, comprimento text, largura text, espessura text, fornecedor text, especificidade text
    )
), normalized as (
    select
        source.*,
        case prefixo when 'JANELA' then '1- JANELA' when 'VIDRO FIXO' then '2- VIDRO FIXO'
            when 'VIDRO FIXO C/ JANELA' then '3- VIDRO FIXO C/ JANELA' when 'CJ' then '4- CJ' end as prefixo_valor,
        case comercial when 'SELADO' then '1- SELADO' when 'VISTA' then '2- VISTA' end as comercial_valor,
        case lado when 'LD' then '1- LD' when 'LE' then '2- LE' else '' end as lado_valor,
        case veiculo
            when 'B/J/D' then '1- B/J/D' when 'E/S/J' then '2- E/S/J' when 'MASTER' then '3- MASTER'
            when 'MASTER L1H1' then '4- MASTER L1H1' when 'MASTER L2H2' then '5- MASTER L2H2'
            when 'MASTER L2H2/L3H2' then '6- MASTER L2H2/L3H2' when 'MASTER L3H2' then '7- MASTER L3H2'
            when 'SPRINTER' then '8- SPRINTER' when 'SPRINTER 10 M' then '9- SPRINTER 10 M'
            when 'SPRINTER 10/14 M' then '10- SPRINTER 10/14 M' when 'SPRINTER 14 M' then '11- SPRINTER 14 M'
            when 'TRANSIT' then '12- TRANSIT' when 'TRANSIT L2H3' then '13- TRANSIT L2H3'
            when 'TRANSIT L3H3' then '14- TRANSIT L3H3' else '' end as veiculo_valor,
        case local when '1 VAO' then '1- 1 VAO' when '1 VAO LD/LE / 2 VAO LE' then '2- 1 VAO LD/LE / 2 VAO LE'
            when '2 VAO' then '3- 2 VAO' when '3 VAO' then '4- 3 VAO' when 'VIGIA' then '5- VIGIA' else '' end as local_valor,
        case cor when 'VERDE' then '1- VERDE' else '' end as cor_valor,
        case when comprimento = '' then '' else concat(array_position(array['671','676','686','743','771','792','814','815','830','838','923','976','1078','1082','1089','1096','1171','1177','1195','1210','1291','1352','1357','1365','1373','1383','1389','1400','1404','1418','1422','1425','1431','1437','1637','1664','1666'], comprimento), '- ', comprimento) end as comprimento_valor,
        case when largura = '' then '' else concat(array_position(array['495','534','556','583','630','667','686','705','716','728','730','738','739','741','761','766','769','771','793','798','800','802','806','814','830','833'], largura), '- ', largura) end as largura_valor,
        case espessura when '3' then '1- 3' when '4' then '2- 4' when '5' then '3- 5' else '' end as espessura_valor,
        case fornecedor when 'SALT' then '1- SALT' when 'STY' then '2- STY' when 'VF' then '3- VF' when 'VTRX' then '4- VTRX' else '' end as fornecedor_valor,
        case especificidade when '2º VAO LD VIDRO FIXO' then '1- 2º VAO LD VIDRO FIXO' when 'N/A' then '2- N/A' else '' end as especificidade_valor
    from source
), expected as (
    select
        normalized.sku,
        concat_ws(' ', prefixo, comercial, nullif(lado, ''), nullif(concat_ws('X', nullif(comprimento, ''), nullif(largura, ''), nullif(espessura, '')), ''), nullif(veiculo, ''), nullif(local, ''), nullif(fornecedor, ''), nullif(especificidade, '')) as descricao_primaria,
        nullif(concat_ws(' ', prefixo, comercial, nullif(lado, ''), nullif(concat_ws('X', nullif(comprimento, ''), nullif(largura, ''), nullif(espessura, '')), ''), nullif(veiculo, ''), nullif(local, ''), nullif(fornecedor, ''), nullif(especificidade, ''), nullif(cor, '')), '') as descricao_secundaria,
        jsonb_build_object(
            'pn_grupo_codigo', jsonb_build_array(case when sku like '3012%' then '30' else '10' end),
            'prefixo', jsonb_build_array(prefixo_valor), 'descritor_base', jsonb_build_array(comercial_valor),
            'veiculo_modelo', case when lado_valor = '' then '[]'::jsonb else jsonb_build_array(lado_valor) end,
            'posicao_lado', jsonb_build_array(veiculo_valor), 'medida', case when local_valor = '' then '[]'::jsonb else jsonb_build_array(local_valor) end,
            'material_cor', case when cor_valor = '' then '[]'::jsonb else jsonb_build_array(cor_valor) end,
            'complemento_regra', case when comprimento_valor = '' then '[]'::jsonb else jsonb_build_array(comprimento_valor) end,
            'fornecedor_referencia', case when largura_valor = '' then '[]'::jsonb else jsonb_build_array(largura_valor) end,
            'espessura', case when espessura_valor = '' then '[]'::jsonb else jsonb_build_array(espessura_valor) end,
            'fornecedor', jsonb_build_array(fornecedor_valor), 'especificidade', jsonb_build_array(especificidade_valor),
            'possui_bom', sku like '3012%'
        ) as form_values,
        jsonb_build_object(
            'prefixo', prefixo_valor, 'descritor_base', comercial_valor, 'veiculo_modelo', lado_valor,
            'posicao_lado', veiculo_valor, 'medida', local_valor, 'material_cor', cor_valor,
            'complemento_regra', comprimento_valor, 'fornecedor_referencia', largura_valor,
            'espessura', espessura_valor, 'fornecedor', fornecedor_valor, 'especificidade', especificidade_valor
        ) as field_values,
        jsonb_build_object(
            'prefixo', regexp_replace(prefixo_valor, '^([0-9]+)-.*$', '\\1'), 'descritor_base', regexp_replace(comercial_valor, '^([0-9]+)-.*$', '\\1'),
            'veiculo_modelo', regexp_replace(lado_valor, '^([0-9]+)-.*$', '\\1'), 'posicao_lado', regexp_replace(veiculo_valor, '^([0-9]+)-.*$', '\\1'),
            'medida', regexp_replace(local_valor, '^([0-9]+)-.*$', '\\1'), 'material_cor', regexp_replace(cor_valor, '^([0-9]+)-.*$', '\\1'),
            'complemento_regra', regexp_replace(comprimento_valor, '^([0-9]+)-.*$', '\\1'), 'fornecedor_referencia', regexp_replace(largura_valor, '^([0-9]+)-.*$', '\\1'),
            'espessura', regexp_replace(espessura_valor, '^([0-9]+)-.*$', '\\1'), 'fornecedor', regexp_replace(fornecedor_valor, '^([0-9]+)-.*$', '\\1'),
            'especificidade', regexp_replace(especificidade_valor, '^([0-9]+)-.*$', '\\1')
        ) as field_codes,
        upper(concat_ws(' ', sku, '{CATEGORY_LABEL}', prefixo, comercial, lado, veiculo, local, cor, comprimento, largura, espessura, fornecedor, especificidade)) as search_text
    from normalized
), changes as (
    select
        record.id, record.sku,
        record.descricao_primaria as before_primaria,
        record.descricao_secundaria as before_secundaria,
        record.form_values as before_form_values,
        record.field_values as before_field_values,
        record.field_codes as before_field_codes,
        record.search_text as before_search_text,
        expected.descricao_primaria,
        expected.descricao_secundaria,
        expected.form_values,
        expected.field_values,
        expected.field_codes,
        expected.search_text
    from public.cadastro_registros record
    join expected using (sku)
    where record.category_key = '{CATEGORY_KEY}'
      and (
          record.descricao_primaria,
          record.descricao_secundaria,
          record.form_values,
          record.field_values,
          record.field_codes,
          record.search_text
      ) is distinct from (
          expected.descricao_primaria,
          expected.descricao_secundaria,
          expected.form_values,
          expected.field_values,
          expected.field_codes,
          expected.search_text
      )
), audited as (
    insert into public.erp_audit_events(
        entity_type, entity_id, action, actor, origin, before_data, after_data, reason
    )
    select
        'CADASTRO_VIDROS', null, 'RECONCILIAR_PLANILHA_VIDROS', 'sistema:migration', 'MODULO_CADASTRO',
        jsonb_build_object(
            'registro_id', changes.id, 'sku', changes.sku,
            'descricao_primaria', changes.before_primaria,
            'descricao_secundaria', changes.before_secundaria,
            'form_values', changes.before_form_values,
            'field_values', changes.before_field_values,
            'field_codes', changes.before_field_codes
        ),
        jsonb_build_object(
            'registro_id', changes.id, 'sku', changes.sku,
            'descricao_primaria', changes.descricao_primaria,
            'descricao_secundaria', changes.descricao_secundaria,
            'form_values', changes.form_values,
            'field_values', changes.field_values,
            'field_codes', changes.field_codes
        ),
        'Campos e descricoes normalizados a partir de {workbook_name}; sem alteracao de estoque, B.O.M. ou status.'
    from changes
    returning id
), updated as (
    update public.cadastro_registros record
       set descricao_primaria = changes.descricao_primaria,
           descricao_secundaria = changes.descricao_secundaria,
           caracteres_primario = char_length(changes.descricao_primaria),
           caracteres_secundario = char_length(changes.descricao_secundaria),
           form_values = changes.form_values,
           field_values = changes.field_values,
           field_codes = changes.field_codes,
           search_text = changes.search_text,
           updated_at = now()
      from changes
     where record.id = changes.id
    returning record.sku
)
select
    (select count(*) from expected) as skus_planilha,
    (select count(*) from changes) as cadastros_alterados,
    (select count(*) from updated) as cadastros_atualizados,
    (select count(*) from audited) as eventos_auditoria;
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    arguments = parser.parse_args()
    sheet = load_workbook(arguments.workbook, data_only=True, read_only=True).active
    rows = [build_row(tuple(sheet.cell(row_number, column).value for column in range(2, 14))) for row_number in range(4, sheet.max_row + 1) if text(sheet.cell(row_number, 2).value)]
    if len(rows) != 156 or len({row["sku"] for row in rows}) != 156:
        raise ValueError(f"Esperados 156 SKUs unicos; encontrados {len(rows)}.")
    arguments.output.write_text(render_sql(rows, arguments.workbook.name), encoding="utf-8")
    print(f"{len(rows)} SKUs preparados em {arguments.output}")


if __name__ == "__main__":
    main()

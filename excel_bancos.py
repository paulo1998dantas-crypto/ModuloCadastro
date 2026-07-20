import json
import os
import re
import shutil
import sys
import tempfile
import unicodedata
import urllib.error
import urllib.parse
import urllib.request
import zipfile
import xml.etree.ElementTree as ET
import uuid
from copy import copy, deepcopy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.formula import ArrayFormula


DEFAULT_NEW_WORKBOOK_NAME = "Cadastro Bancos.xlsx"
DEFAULT_CATEGORY_KEY = "bancos"
DEFAULT_CATEGORY_LABEL = "Bancos"
CATEGORY_KEY_ALIASES = {
    "cat_20_bco": DEFAULT_CATEGORY_KEY,
    "cat_20_cj_bco": DEFAULT_CATEGORY_KEY,
    "20_cj_bco": DEFAULT_CATEGORY_KEY,
}
CATEGORY_CANONICAL_LABELS = {
    DEFAULT_CATEGORY_KEY: "20 - BANCOS",
}
PN_GROUP_FORM_KEY = "grupo_codigo"
FIRST_DATA_ROW = 3
REGISTRATION_SHEET_NAME = "_cadastro_app"
DRAFT_SHEET_NAME = "_rascunhos_app"
SELECTION_MODE_UNITARIA = "unitaria"
SELECTION_MODE_MULTIPLA = "multipla"
EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
IGNORED_SCAN_DIRS = {"__pycache__", "outputs", ".idea", ".venv", "backups", "browser_profile", "_internal"}
VALUE_RE = re.compile(r"^\s*(\d+)\s*[-â€“â€”]\s*(.+?)\s*$")
INVALID_SHEET_CHARS_RE = re.compile(r'[:\\/?*\[\]]')
DISTANCIA_PE_KEY = "distancia_pe"
DISTANCIA_PE_PREFIX = "ORIENTADO A ESQ:"
ORDINAL_MASCULINE = "\ufff0"
ORDINAL_FEMININE = "\ufff1"
DESCRIPTION_PRIMARY_HEADER = "DESCRICAO PRIMARIA"
DESCRIPTION_SECONDARY_HEADER = "DESCRICAO SECUNDARIA"
DESCRIPTION_SUFFIX_HEADER = "SUFIXO"
CATALOG_TABLE = "cadastro_catalogo"
CATALOG_KEY = "default"
REMOTE_CATALOG_MTIME = -1.0
_CATALOG_CACHE: dict[str, Any] | None = None
_CATALOG_CACHE_MTIME: float | None = None
_REGISTRATION_CACHE: dict[tuple[str, str, int], dict[str, Any]] = {}
_PRODUCT_CATALOG_CACHE: tuple[Path, float, list[dict[str, str]]] | None = None

DEFAULT_CONDITIONAL_RULES = [
    {
        "key": "cond_pre_fixo_veiculo",
        "source_field_key": "pre_fixo",
        "source_values": ["4- BCO CARONA ORIGINAL", "5- BCO MOTORISTA ORIGINAL", "6- BCO ORIGINAL"],
        "target_field_key": "veiculo",
        "target_field_label": "VEÃCULO",
        "target_field_scope": "secundaria",
        "action": "show",
    },
    {
        "key": "cond_encosto_reclinador",
        "source_field_key": "encosto",
        "source_values": ["1- FIXO", "3- INTERICO"],
        "target_field_key": "tipo_do_reclinador",
        "target_field_label": "TIPO DO RECLINADOR",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_encosto_grau_reclinacao",
        "source_field_key": "encosto",
        "source_values": ["1- FIXO", "3- INTERICO"],
        "target_field_key": "grau_reclinacao",
        "target_field_label": "GRAU RECLINAÃ‡ÃƒO",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_usb",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "usb",
        "target_field_label": "USB",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_resfriador",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "resfriador",
        "target_field_label": "RESFRIADOR",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_aquecedor",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "aquecedor",
        "target_field_label": "AQUECEDOR",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_massageador",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "massageador",
        "target_field_label": "MASSAGEADOR",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_mesa_snack",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "mesa_snack",
        "target_field_label": "MESA SNACK",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_isofix",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "isofix",
        "target_field_label": "ISOFIX",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_posicao_isofix",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "posicao_isofix",
        "target_field_label": "POSIÃ‡ÃƒO ISOFIX",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_linha_apoio_panturrilha",
        "source_field_key": "linha",
        "source_values": ["1- LB"],
        "target_field_key": "apoio_panturrilha",
        "target_field_label": "APOIO PANTURRILHA",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_tipo_apoio",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "tipo_apoio_braco",
        "target_field_label": "TIPO APOIO BRAÃ‡O",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_lado_braco",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "lado_braco",
        "target_field_label": "LADO BRAÃ‡O",
        "target_field_scope": "primaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_corredor",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "braco_corredor",
        "target_field_label": "BRAÃ‡O CORREDOR",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_central",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "braco_central",
        "target_field_label": "BRAÃ‡O CENTRAL",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_cor_do_braco",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "cor_do_braco",
        "target_field_label": "COR DO BRAÃ‡O",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_braco_parede",
        "source_field_key": "braco",
        "source_values": ["S/ APOIO DE BRAÃ‡O", "NA", "N/A"],
        "target_field_key": "braco_parede",
        "target_field_label": "BRAÃ‡O PAREDE",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_pega_mao_cor",
        "source_field_key": "pega_mao",
        "source_values": ["S/ PEGA MÃƒO", "NA", "N/A"],
        "target_field_key": "cor_pega_mao",
        "target_field_label": "COR PEGA MÃƒO",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_acabamento_cor",
        "source_field_key": "acabamento_lateral",
        "source_values": ["SEM ACABAMENTO LATERAL", "NA"],
        "target_field_key": "cor_acabamento",
        "target_field_label": "COR ACABAMENTO",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_isofix_posicao",
        "source_field_key": "isofix",
        "source_values": ["S/ ISOFIX", "4"],
        "target_field_key": "posicao_isofix",
        "target_field_label": "POSIÃ‡ÃƒO ISOFIX",
        "target_field_scope": "secundaria",
        "action": "hide",
    },
    {
        "key": "cond_tipo_cinto_lado",
        "source_field_key": "tipo_cinto",
        "source_values": ["3P"],
        "target_field_key": "lado_cinto",
        "target_field_label": "LADO CINTO",
        "target_field_scope": "secundaria",
        "action": "show",
    },
]


def _project_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


PROJECT_DIR = _project_dir()
RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", PROJECT_DIR))


def _runtime_data_dir() -> Path | None:
    data_dir_env = (os.environ.get("CADASTRO_DATA_DIR") or "").strip()
    save_mode_env = (os.environ.get("CADASTRO_SAVE_MODE") or "").strip().lower()
    if data_dir_env or save_mode_env in {"bridge", "ponte", "online"}:
        try:
            import bridge_store

            return bridge_store.data_dir()
        except Exception:
            if data_dir_env:
                return Path(data_dir_env).resolve()
    return None


def _runtime_file(name: str) -> Path:
    data_dir = _runtime_data_dir()
    if data_dir is not None:
        return data_dir / name
    return PROJECT_DIR / name


CONFIG_PATH = _runtime_file("config.json")
DATA_PATH = _runtime_file("cadastro_dados.json")
BACKUP_DIR = PROJECT_DIR / "backups"
BOM_OUTPUT_DIR = PROJECT_DIR / "outputs" / "bom"
_template_env = os.environ.get("CADASTRO_BANCO_TEMPLATE", "").strip()
DEFAULT_TEMPLATE_PATH = Path(_template_env).expanduser() if _template_env else PROJECT_DIR / "TEMPLATE ID BANCO.xlsx"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_label(value: Any) -> str:
    text = re.sub(r"\([^)]*\)", "", clean_text(value))
    text = unicodedata.normalize("NFKD", text.upper())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def normalize_option_label(value: Any) -> str:
    text = clean_text(value)
    text = unicodedata.normalize("NFKD", text.upper())
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def option_label(value: Any) -> str:
    text = clean_text(value)
    match = VALUE_RE.match(text)
    if match:
        return match.group(2).strip()
    return text


def option_code(value: Any) -> str:
    text = clean_text(value)
    match = VALUE_RE.match(text)
    if match:
        return match.group(1).strip()
    return ""


def option_identity(value: Any) -> str:
    return normalize_option_label(option_label(value))


def rule_option_token(value: Any) -> str:
    return normalize_option_label(option_label(value)).replace(" ", "")


def strip_accents(value: Any) -> str:
    text = clean_text(value)
    normalized = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch))


def normalize_option_text(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""

    ordinal_masculine = "Âº"
    ordinal_feminine = "Âª"

    def _preserve_ordinals(input_text: str) -> str:
        return input_text.replace(ordinal_masculine, "ï¿°").replace(ordinal_feminine, "ï¿±")

    def _restore_ordinals(input_text: str) -> str:
        return input_text.replace("ï¿°", ordinal_masculine).replace("ï¿±", ordinal_feminine)

    match = VALUE_RE.match(text)
    if match:
        code = match.group(1).strip()
        body = _preserve_ordinals(match.group(2).strip())
        body = strip_accents(body).upper()
        body = _restore_ordinals(body)
        body = re.sub(r"\s+", " ", body)
        return f"{code}- {body}"

    text = _preserve_ordinals(text)
    text = strip_accents(text).upper()
    text = _restore_ordinals(text)
    return re.sub(r"\s+", " ", text).strip()

def compose_secondary_description(primary: Any, secondary: Any) -> str:
    primary_text = clean_text(primary)
    secondary_text = clean_text(secondary)
    if not primary_text:
        return secondary_text
    if not secondary_text:
        return primary_text

    normalized_primary = normalize_label(primary_text)
    normalized_secondary = normalize_label(secondary_text)
    if normalized_secondary == normalized_primary or normalized_secondary.startswith(f"{normalized_primary} "):
        return secondary_text
    return f"{primary_text} {secondary_text}".strip()


def field_key(value: Any) -> str:
    normalized = normalize_label(value).lower().replace(" ", "_")
    return normalized or "campo"


def category_key(value: Any) -> str:
    normalized = normalize_label(value).lower().replace(" ", "_")
    return normalized or "categoria"


def canonical_category_key(value: Any) -> str:
    key = clean_text(value)
    return CATEGORY_KEY_ALIASES.get(key, key)


def category_key_candidates(value: Any) -> list[str]:
    canonical = canonical_category_key(value)
    keys = [canonical] if canonical else []
    keys.extend(alias for alias, target in CATEGORY_KEY_ALIASES.items() if target == canonical and alias not in keys)
    return keys


def same_category_key(left: Any, right: Any) -> bool:
    return canonical_category_key(left) == canonical_category_key(right)


def field_scope(value: Any) -> str:
    return "secundaria" if clean_text(value).lower() == "secundaria" else "primaria"


def field_selection_mode(value: Any) -> str:
    normalized = normalize_label(value)
    if normalized in {"MULTIPLA", "MULTIPLA SELECAO", "SELECAO MULTIPLA"}:
        return SELECTION_MODE_MULTIPLA
    return SELECTION_MODE_UNITARIA


def header_for_field(label: str, scope: str) -> str:
    suffix = "PRIMARIO" if field_scope(scope) == "primaria" else "SECUNDARIO"
    return f"{clean_text(label)} - {suffix}"


def _field_description_order(field: dict[str, Any], fallback_index: int) -> int:
    try:
        order = int(field.get("description_order"))
    except (TypeError, ValueError):
        order = 0
    return order if order > 0 else fallback_index


def _ordered_fields_for_description(fields: list[dict[str, Any]]) -> list[dict[str, Any]]:
    decorated: list[tuple[int, int, int, dict[str, Any]]] = []
    for index, field in enumerate(fields, start=1):
        scope_rank = 0 if field.get("scope") == "primaria" else 1
        decorated.append((scope_rank, _field_description_order(field, index), index, field))
    decorated.sort(key=lambda item: (item[0], item[1], item[2]))
    return [item[3] for item in decorated]


def _normalize_field_orders(category: dict[str, Any]) -> None:
    for scope in ("primaria", "secundaria"):
        scoped_fields = [field for field in category.get("fields") or [] if field.get("scope") == scope]
        scoped_fields.sort(key=lambda field: _field_description_order(field, 10**9))
        for index, field in enumerate(scoped_fields, start=1):
            field["description_order"] = index


def _read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except json.JSONDecodeError:
        return {}


def _write_json(path: Path, data: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _read_config() -> dict[str, Any]:
    return _read_json(CONFIG_PATH)


def _write_config(data: dict[str, Any]) -> None:
    _write_json(CONFIG_PATH, data)


def _seed_candidates(file_name: str) -> list[Path]:
    return [
        PROJECT_DIR / "_internal" / file_name,
        RESOURCE_DIR / file_name,
        RESOURCE_DIR / "_internal" / file_name,
    ]


def _ensure_seed_file(path: Path, file_name: str) -> None:
    if path.exists():
        return
    for candidate in _seed_candidates(file_name):
        if candidate.exists():
            shutil.copy2(candidate, path)
            return


def _copy_to_temp(path: Path) -> Path:
    suffix = path.suffix or ".xlsx"
    fd, temp_name = tempfile.mkstemp(prefix="_cadastro_app_", suffix=suffix)
    os.close(fd)
    temp_path = Path(temp_name)
    shutil.copy2(path, temp_path)
    return temp_path


def _load(path: Path):
    try:
        return load_workbook(path, keep_links=True, keep_vba=False)
    except IndexError:
        repaired_path = _repair_invalid_style_ids(path)
        workbook = load_workbook(repaired_path, keep_links=True, keep_vba=False)
        setattr(workbook, "_cadastro_cleanup_path", repaired_path)
        return workbook


def _repair_invalid_style_ids(path: Path) -> Path:
    if not path.exists():
        return path

    fd, temp_name = tempfile.mkstemp(prefix="_cadastro_app_repair_", suffix=path.suffix or ".xlsx", dir=str(path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

    try:
        with zipfile.ZipFile(path, "r") as source_zip:
            try:
                styles_root = ET.fromstring(source_zip.read("xl/styles.xml"))
                cellxfs = styles_root.find("a:cellXfs", ns)
                style_count = int(cellxfs.attrib.get("count", "0")) if cellxfs is not None else 0
            except Exception:
                style_count = 0

            max_style = max(0, style_count - 1)

            with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as output_zip:
                for info in source_zip.infolist():
                    payload = source_zip.read(info.filename)
                    if info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml") and style_count:
                        try:
                            root = ET.fromstring(payload)
                            changed = False
                            for cell in root.findall(".//a:c", ns):
                                style_value = cell.attrib.get("s")
                                if style_value is None:
                                    continue
                                try:
                                    style_id = int(style_value)
                                except ValueError:
                                    continue
                                if style_id >= style_count:
                                    cell.set("s", str(max_style))
                                    changed = True
                            if changed:
                                payload = ET.tostring(root, encoding="utf-8", xml_declaration=True)
                        except Exception:
                            pass
                    output_zip.writestr(info, payload)
        return temp_path
    except Exception:
        temp_path.unlink(missing_ok=True)
        raise


def _unique_key(label: str, used: set[str]) -> str:
    base = field_key(label)
    candidate = base
    counter = 2
    while candidate in used:
        candidate = f"{base}_{counter}"
        counter += 1
    return candidate


def _unique_category_key(label: str, used: set[str]) -> str:
    base = category_key(label)
    candidate = base
    counter = 2
    while candidate in used:
        candidate = f"{base}_{counter}"
        counter += 1
    return candidate


def _default_category() -> dict[str, Any]:
    return {
        "key": DEFAULT_CATEGORY_KEY,
        "label": DEFAULT_CATEGORY_LABEL,
        "sheet_name": DEFAULT_CATEGORY_LABEL,
        "fields": [],
    }


PN_GROUP_BY_PREFIX = {
    "ABS": "10",
    "ACABAMENTO DIFUSOR": "10",
    "ACABAMENTO LATERAL": "10",
    "ACABAMENTO PLASTICO": "10",
    "ACABAMENTO PONTEIRA": "10",
    "ACABAMENTO TAMPA": "10",
    "ACABAMENTO TELA": "10",
    "ACESSORIO": "10",
    "ALCA": "10",
    "APOIO BRACO": "10",
    "ARCO": "10",
    "ARGOLA": "10",
    "BCO": "10",
    "BCO CARONA": "10",
    "BCO CARONA ORIGINAL": "10",
    "BCO MOTORISTA": "10",
    "BCO MOTORISTA ORIGINAL": "10",
    "BOBINA": "10",
    "BOTOEIRA": "10",
    "CALCO": "10",
    "CANTONEIRA": "10",
    "CAPA": "10",
    "CHICOTE": "10",
    "CINTO": "10",
    "CM": "10",
    "COLUNA": "10",
    "CONTROLADOR": "10",
    "CP": "10",
    "CUPULA": "10",
    "DECODER": "10",
    "DIAMOND": "10",
    "ELETRICA": "10",
    "ESTROBO": "10",
    "EXAUSTOR": "10",
    "FAROL": "10",
    "FATURAMENTO DIRETO": "10",
    "FX": "10",
    "ILUMINACAO": "10",
    "INVERSOR": "10",
    "JANELA": "10",
    "LANTERNA": "10",
    "LED": "10",
    "LUMINARIA": "10",
    "LUZ": "10",
    "MINI": "10",
    "MODULO": "10",
    "MP": "10",
    "PC": "10",
    "PE": "10",
    "PEGA MAO": "10",
    "PORTA": "10",
    "POSTICO": "10",
    "QUEBRA": "10",
    "REFORCO": "10",
    "SIRENE": "10",
    "SUPORTE": "10",
    "TAMPA": "10",
    "TOMADA": "10",
    "VIDRO": "10",
    "PP": "20",
    "TETO": "20",
    "CHAPA": "30",
    "CJ": "30",
    "JI CONFORT": "40",
    "JI URBAN": "40",
    "CITROEN": "80",
    "FIAT": "80",
    "FORD": "80",
    "IVECO": "80",
    "MERCEDES BENZ SPRINTER": "80",
    "PEUGEOT": "80",
    "RENAULT": "80",
}

PN_GROUP_DEFAULT_LABELS = {
    "10": "INSUMO",
    "20": "PRODUTO PROCESSO",
    "30": "CONJUNTO / KIT",
    "40": "TRANSFORMACAO",
    "50": "MRO (MANUTENCAO, REPARO E OPERACOES)",
    "60": "EMBALAGEM",
    "70": "ATIVO FIXO",
    "80": "VEICULO",
    "90": "PROTOTIPO",
}


def _catalog_supabase_url() -> str:
    raw = clean_text(os.environ.get("SUPABASE_URL")) or clean_text(os.environ.get("CADASTRO_SUPABASE_URL"))
    return raw.rstrip("/")


def _catalog_service_key() -> str:
    return clean_text(os.environ.get("SUPABASE_SERVICE_ROLE_KEY")) or clean_text(
        os.environ.get("CADASTRO_SUPABASE_SERVICE_ROLE_KEY")
    )


def _catalog_supabase_enabled() -> bool:
    mode = clean_text(os.environ.get("CADASTRO_SAVE_MODE")).lower()
    return mode in {"supabase", "postgres", "database", "banco"} and bool(
        _catalog_supabase_url() and _catalog_service_key()
    )


def _catalog_supabase_headers(prefer: str = "") -> dict[str, str]:
    key = _catalog_service_key()
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def _catalog_supabase_request(method: str, query: list[tuple[str, str]] | None = None, payload: Any = None, prefer: str = "") -> Any:
    query_string = urllib.parse.urlencode(query or [])
    url = f"{_catalog_supabase_url()}/rest/v1/{CATALOG_TABLE}"
    if query_string:
        url = f"{url}?{query_string}"
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8") if payload is not None else None
    request = urllib.request.Request(url, data=data, headers=_catalog_supabase_headers(prefer), method=method)
    with urllib.request.urlopen(request, timeout=30) as response:
        body = response.read().decode("utf-8")
        return json.loads(body) if body else None


def _load_supabase_catalog() -> tuple[bool, dict[str, Any] | None]:
    if not _catalog_supabase_enabled():
        return False, None
    try:
        rows = _catalog_supabase_request(
            "GET",
            [
                ("select", "payload"),
                ("config_key", f"eq.{CATALOG_KEY}"),
                ("limit", "1"),
            ],
        ) or []
    except (urllib.error.HTTPError, urllib.error.URLError, TimeoutError, OSError):
        return False, None
    if not rows:
        return True, None
    payload = rows[0].get("payload")
    return True, payload if isinstance(payload, dict) else None


def _save_supabase_catalog(catalog: dict[str, Any]) -> None:
    if not _catalog_supabase_enabled():
        return
    try:
        _catalog_supabase_request(
            "POST",
            payload=[{"config_key": CATALOG_KEY, "payload": catalog}],
            prefer="resolution=merge-duplicates,return=minimal",
        )
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Erro ao salvar catalogo no Supabase: {body}") from exc
    except (urllib.error.URLError, TimeoutError, OSError) as exc:
        raise RuntimeError(f"Nao foi possivel salvar catalogo no Supabase: {exc}") from exc


def _pn_group_code(value: Any) -> str:
    match = re.search(r"\d+", clean_text(value))
    if not match:
        return ""
    return f"{int(match.group(0)):02d}"


def _pn_group_prefixes(value: Any) -> list[str]:
    if isinstance(value, list):
        raw_values = value
    else:
        raw_values = re.split(r"[,;\n]+", clean_text(value))
    prefixes: list[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        prefix = normalize_label(raw)
        if not prefix or prefix in seen:
            continue
        prefixes.append(prefix)
        seen.add(prefix)
    return prefixes


def _default_pn_groups() -> list[dict[str, Any]]:
    by_code: dict[str, list[str]] = {}
    for prefix, code in PN_GROUP_BY_PREFIX.items():
        by_code.setdefault(code, []).append(prefix)
    for code in PN_GROUP_DEFAULT_LABELS:
        by_code.setdefault(code, [])
    return [
        {
            "code": code,
            "label": PN_GROUP_DEFAULT_LABELS.get(code, f"GRUPO {code}"),
            "prefixes": sorted(prefixes),
        }
        for code, prefixes in sorted(by_code.items())
    ]


def _sanitize_pn_groups(groups: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    source = groups if groups else _default_pn_groups()
    cleaned: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    used_prefixes: set[str] = set()
    for group in source:
        code = _pn_group_code(group.get("code"))
        label = clean_text(group.get("label")).upper()
        if not code:
            continue
        if code in seen_codes:
            continue
        prefixes = []
        for prefix in _pn_group_prefixes(group.get("prefixes") or []):
            if prefix in used_prefixes:
                continue
            prefixes.append(prefix)
            used_prefixes.add(prefix)
        cleaned.append(
            {
                "code": code,
                "label": label or PN_GROUP_DEFAULT_LABELS.get(code, f"GRUPO {code}"),
                "prefixes": prefixes,
            }
        )
        seen_codes.add(code)
    return cleaned or _default_pn_groups()


def _default_catalog() -> dict[str, Any]:
    return {
        "version": 2,
        "active_category": DEFAULT_CATEGORY_KEY,
        "categories": [_default_category()],
        "pn_groups": _default_pn_groups(),
    }


def _sanitize_fields(fields: list[dict[str, Any]] | None, include_defaults: bool = False) -> list[dict[str, Any]]:
    used_keys: set[str] = set()
    cleaned: list[dict[str, Any]] = []
    for field in fields or []:
        label = clean_text(field.get("label")).upper()
        if not label:
            continue
        key = clean_text(field.get("key")) or field_key(label)
        if key in used_keys:
            key = _unique_key(label, used_keys)
        used_keys.add(key)

        scope = field_scope(field.get("scope"))
        selection_mode = field_selection_mode(field.get("selection_mode"))
        try:
            description_order = int(field.get("description_order"))
            if description_order < 1:
                description_order = None
        except (TypeError, ValueError):
            description_order = None

        options: list[str] = []
        seen_options: set[str] = set()
        for option in field.get("options") or []:
            value = normalize_option_text(option)
            if not value:
                continue
            option_key = f"{option_code(value)}|{option_identity(value)}" if option_code(value) else option_identity(value)
            if option_key in seen_options:
                continue
            seen_options.add(option_key)
            options.append(value)

        cleaned.append(
            {
                "key": key,
                "label": label,
                "scope": scope,
                "selection_mode": selection_mode,
                "description_order": description_order,
                "options": options,
            }
        )

    if include_defaults and not cleaned:
        cleaned.extend(_default_category()["fields"])
    return cleaned


def _sanitize_conditional_rules(
    rules: list[dict[str, Any]] | None,
    fields: list[dict[str, Any]],
    category_key: str,
) -> list[dict[str, Any]]:
    field_map = {field["key"]: field for field in fields}
    cleaned: list[dict[str, Any]] = []
    for rule in rules or []:
        source_key = clean_text(rule.get("source_field_key"))
        target_key = clean_text(rule.get("target_field_key"))
        if source_key not in field_map and target_key not in field_map:
            continue
        source_field = field_map.get(source_key)
        target_field = field_map.get(target_key)
        source_values = []
        for value in rule.get("source_values") or []:
            normalized = clean_text(value)
            if normalized:
                source_values.append(rule_option_token(normalized))
        if not source_values:
            continue
        action = clean_text(rule.get("action")).lower()
        if action not in {"hide", "show", "set_primary", "set_secondary"}:
            action = "hide"
        match_by = clean_text(rule.get("match_by")).lower()
        if match_by not in {"option", "prefix"}:
            match_by = "option"
        if action in {"set_primary", "set_secondary"} and (source_field is None or target_field is None):
            continue
        cleaned.append(
            {
                "key": clean_text(rule.get("key")) or uuid.uuid4().hex[:12],
                "source_field_key": source_key,
                "source_field_label": source_field["label"] if source_field else clean_text(rule.get("source_field_label")) or source_key,
                "source_field_scope": source_field["scope"] if source_field else clean_text(rule.get("source_field_scope")) or "primaria",
                "source_values": source_values,
                "target_field_key": target_key,
                "target_field_label": target_field["label"] if target_field else clean_text(rule.get("target_field_label")) or target_key,
                "target_field_scope": target_field["scope"] if target_field else clean_text(rule.get("target_field_scope")) or "secundaria",
                "action": action,
                "match_by": match_by,
            }
        )

    if not cleaned and category_key == DEFAULT_CATEGORY_KEY:
        for rule in DEFAULT_CONDITIONAL_RULES:
            cleaned.append(deepcopy(rule))
    return cleaned


def _merge_field_options(target: dict[str, Any], source: dict[str, Any]) -> None:
    seen_options = {
        f"{option_code(value)}|{option_identity(value)}" if option_code(value) else option_identity(value)
        for value in target.get("options") or []
    }
    for option in source.get("options") or []:
        option_key = f"{option_code(option)}|{option_identity(option)}" if option_code(option) else option_identity(option)
        if option_key in seen_options:
            continue
        target.setdefault("options", []).append(option)
        seen_options.add(option_key)


def _merge_category_data(target: dict[str, Any], source: dict[str, Any]) -> None:
    fields_by_key = {field["key"]: field for field in target.get("fields") or []}
    for field in source.get("fields") or []:
        existing = fields_by_key.get(field["key"])
        if existing:
            _merge_field_options(existing, field)
            continue
        new_field = deepcopy(field)
        target.setdefault("fields", []).append(new_field)
        fields_by_key[new_field["key"]] = new_field

    rules_by_key = {clean_text(rule.get("key")) for rule in target.get("conditional_rules") or []}
    for rule in source.get("conditional_rules") or []:
        rule_key = clean_text(rule.get("key"))
        if rule_key and rule_key in rules_by_key:
            continue
        target.setdefault("conditional_rules", []).append(deepcopy(rule))
        if rule_key:
            rules_by_key.add(rule_key)


def _merge_category_aliases(categories: list[dict[str, Any]]) -> list[dict[str, Any]]:
    merged: list[dict[str, Any]] = []
    by_key: dict[str, dict[str, Any]] = {}
    for raw_category in categories:
        category = deepcopy(raw_category)
        original_key = category["key"]
        canonical_key = canonical_category_key(original_key)
        category["key"] = canonical_key
        if original_key != canonical_key and canonical_key in CATEGORY_CANONICAL_LABELS:
            category["label"] = CATEGORY_CANONICAL_LABELS[canonical_key]
            category["sheet_name"] = _safe_sheet_title(category["label"])

        existing = by_key.get(canonical_key)
        if existing:
            if original_key == canonical_key:
                existing["label"] = category["label"]
                existing["sheet_name"] = category["sheet_name"]
            _merge_category_data(existing, category)
            continue
        by_key[canonical_key] = category
        merged.append(category)
    return merged


def _sanitize_catalog(catalog: dict[str, Any]) -> dict[str, Any]:
    if catalog.get("fields") is not None:
        raw_categories = [
            {
                "key": DEFAULT_CATEGORY_KEY,
                "label": DEFAULT_CATEGORY_LABEL,
                "fields": catalog.get("fields") or [],
            }
        ]
        active_category_value = DEFAULT_CATEGORY_KEY
    else:
        raw_categories = catalog.get("categories") or []
        active_category_value = clean_text(catalog.get("active_category")) or DEFAULT_CATEGORY_KEY

    if not raw_categories:
        raw_categories = [_default_category()]
        active_category_value = DEFAULT_CATEGORY_KEY

    categories: list[dict[str, Any]] = []
    used_keys: set[str] = set()
    used_labels: set[str] = set()

    for raw_category in raw_categories:
        label = clean_text(raw_category.get("label")) or "Categoria"
        label_key = normalize_label(label)
        if label_key in used_labels:
            label = f"{label} {len(categories) + 1}"
            label_key = normalize_label(label)
        used_labels.add(label_key)

        raw_key = clean_text(raw_category.get("key")) or category_key(label)
        key = raw_key if raw_key not in used_keys else _unique_category_key(label, used_keys)
        used_keys.add(key)
        fields = _sanitize_fields(raw_category.get("fields") or [], include_defaults=False)

        categories.append(
            {
                "key": key,
                "label": label,
                "sheet_name": _safe_sheet_title(raw_category.get("sheet_name") or label),
                "fields": fields,
                "conditional_rules": _sanitize_conditional_rules(raw_category.get("conditional_rules") or [], fields, key),
            }
        )

    categories = _merge_category_aliases(categories)
    active_category_value = canonical_category_key(active_category_value)
    if active_category_value not in {category["key"] for category in categories}:
        active_category_value = categories[0]["key"]

    return {
        "version": 2,
        "active_category": active_category_value,
        "categories": categories,
        "pn_groups": _sanitize_pn_groups(catalog.get("pn_groups") if isinstance(catalog, dict) else None),
    }


def load_catalog() -> dict[str, Any]:
    global _CATALOG_CACHE, _CATALOG_CACHE_MTIME
    if _CATALOG_CACHE is not None and _CATALOG_CACHE_MTIME == REMOTE_CATALOG_MTIME:
        return deepcopy(_CATALOG_CACHE)

    remote_available, remote_catalog = _load_supabase_catalog()
    if remote_available:
        if remote_catalog is not None:
            catalog = _sanitize_catalog(remote_catalog)
        else:
            _ensure_seed_file(DATA_PATH, "cadastro_dados.json")
            raw_catalog = _read_json(DATA_PATH) if DATA_PATH.exists() else _default_catalog()
            catalog = _sanitize_catalog(raw_catalog)
            _save_supabase_catalog(catalog)
        _write_json(DATA_PATH, catalog)
        _CATALOG_CACHE = catalog
        _CATALOG_CACHE_MTIME = REMOTE_CATALOG_MTIME
        return deepcopy(catalog)

    _ensure_seed_file(DATA_PATH, "cadastro_dados.json")
    if not DATA_PATH.exists():
        catalog = _default_catalog()
        save_catalog(catalog)
        return deepcopy(catalog)

    mtime = DATA_PATH.stat().st_mtime
    if _CATALOG_CACHE is not None and _CATALOG_CACHE_MTIME == mtime:
        return deepcopy(_CATALOG_CACHE)

    raw_catalog = _read_json(DATA_PATH)
    catalog = _sanitize_catalog(raw_catalog)
    if raw_catalog != catalog:
        save_catalog(catalog)
        mtime = DATA_PATH.stat().st_mtime

    _CATALOG_CACHE = catalog
    _CATALOG_CACHE_MTIME = mtime
    return deepcopy(catalog)


def save_catalog(catalog: dict[str, Any]) -> None:
    global _CATALOG_CACHE, _CATALOG_CACHE_MTIME
    sanitized = _sanitize_catalog(catalog)
    _save_supabase_catalog(sanitized)
    _write_json(DATA_PATH, sanitized)
    _CATALOG_CACHE = sanitized
    _CATALOG_CACHE_MTIME = REMOTE_CATALOG_MTIME if _catalog_supabase_enabled() else DATA_PATH.stat().st_mtime


def list_categories() -> list[dict[str, Any]]:
    categories = []
    for category in load_catalog()["categories"]:
        categories.append(
            {
                "key": category["key"],
                "label": category["label"],
                "fields_count": len(category.get("fields") or []),
            }
        )
    return categories


def list_pn_groups() -> list[dict[str, Any]]:
    groups = []
    for group in load_catalog().get("pn_groups") or _default_pn_groups():
        prefixes = list(group.get("prefixes") or [])
        groups.append(
            {
                "code": group["code"],
                "label": group["label"],
                "prefixes": prefixes,
                "prefixes_text": ", ".join(prefixes),
                "prefix_count": len(prefixes),
            }
        )
    return groups


def _find_pn_group(catalog: dict[str, Any], code_value: str) -> dict[str, Any]:
    code = _pn_group_code(code_value)
    for group in catalog.get("pn_groups") or []:
        if group.get("code") == code:
            return group
    raise ValueError("Grupo nao encontrado.")


def add_pn_group(code: str, label: str, prefixes: str = "") -> dict[str, Any]:
    catalog = load_catalog()
    group_code = _pn_group_code(code)
    group_label = clean_text(label).upper()
    if not group_code:
        raise ValueError("Informe o codigo numerico do grupo.")
    if not group_label:
        raise ValueError("Informe o nome do grupo.")
    groups = catalog.setdefault("pn_groups", _default_pn_groups())
    if any(group.get("code") == group_code for group in groups):
        raise ValueError("Esse codigo de grupo ja existe.")
    group = {
        "code": group_code,
        "label": group_label,
        "prefixes": _pn_group_prefixes(prefixes),
    }
    groups.append(group)
    groups.sort(key=lambda item: item.get("code", ""))
    save_catalog(catalog)
    return group


def update_pn_group(code: str, label: str, prefixes: str = "") -> dict[str, Any]:
    catalog = load_catalog()
    group = _find_pn_group(catalog, code)
    group_label = clean_text(label).upper()
    if not group_label:
        raise ValueError("Informe o nome do grupo.")
    group["label"] = group_label
    group["prefixes"] = _pn_group_prefixes(prefixes)
    save_catalog(catalog)
    return group


def _find_category(catalog: dict[str, Any], category_key_value: str) -> dict[str, Any]:
    requested = canonical_category_key(
        clean_text(category_key_value) or clean_text(catalog.get("active_category")) or DEFAULT_CATEGORY_KEY
    )
    for category in catalog["categories"]:
        if same_category_key(category["key"], requested):
            return category
    return catalog["categories"][0]


def selected_category(category_key_value: str) -> dict[str, Any]:
    category = _find_category(load_catalog(), category_key_value)
    return {"key": category["key"], "label": category["label"]}


def set_active_category(category_key_value: str) -> dict[str, Any]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    catalog["active_category"] = category["key"]
    save_catalog(catalog)
    sync_workbook_structure(category["key"])
    return {"key": category["key"], "label": category["label"]}


def add_category(label: str) -> dict[str, str]:
    catalog = load_catalog()
    label_clean = clean_text(label)
    if not label_clean:
        raise ValueError("Informe o nome da categoria.")
    if normalize_label(label_clean) in {normalize_label(item["label"]) for item in catalog["categories"]}:
        raise ValueError("Essa categoria jÃ¡ existe.")
    used = {item["key"] for item in catalog["categories"]}
    category = {
        "key": _unique_category_key(label_clean, used),
        "label": label_clean,
        "sheet_name": _safe_sheet_title(label_clean),
        "fields": [],
    }
    catalog["categories"].append(category)
    catalog["active_category"] = category["key"]
    save_catalog(catalog)
    sync_workbook_structure(category["key"])
    return {
        "category": category["label"],
        "category_key": category["key"],
        "path": str(DATA_PATH),
    }


def update_category(category_key_value: str, label: str) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    label_clean = clean_text(label)
    if not label_clean:
        raise ValueError("Informe o nome da categoria.")
    for existing in catalog["categories"]:
        if existing["key"] != category["key"] and normalize_label(existing["label"]) == normalize_label(label_clean):
            raise ValueError("JÃ¡ existe outra categoria com esse nome.")
    category["label"] = label_clean
    category["sheet_name"] = _safe_sheet_title(label_clean)
    save_catalog(catalog)
    sync_workbook_structure(category["key"])
    return {
        "category": category["label"],
        "category_key": category["key"],
        "path": str(DATA_PATH),
    }


def delete_category(category_key_value: str) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    if len(catalog["categories"]) == 1:
        raise ValueError("Pelo menos uma categoria deve permanecer cadastrada.")
    catalog["categories"] = [item for item in catalog["categories"] if item["key"] != category["key"]]
    if catalog.get("active_category") == category["key"]:
        catalog["active_category"] = catalog["categories"][0]["key"]
    save_catalog(catalog)
    return {"category": category["label"], "path": str(DATA_PATH)}


def active_workbook_path() -> Path:
    env_workbook = clean_text(os.environ.get("CADASTRO_ACTIVE_WORKBOOK"))
    if env_workbook:
        path = Path(env_workbook)
        if not path.is_absolute():
            path = (PROJECT_DIR / path).resolve()
        return path

    data_dir = clean_text(os.environ.get("CADASTRO_DATA_DIR"))
    save_mode = clean_text(os.environ.get("CADASTRO_SAVE_MODE")).lower()
    if data_dir and save_mode in {"bridge", "ponte", "online"}:
        return (Path(data_dir) / DEFAULT_NEW_WORKBOOK_NAME).resolve()

    _ensure_seed_file(CONFIG_PATH, "config.json")
    config = _read_config()
    raw = clean_text(config.get("active_workbook"))
    if raw:
        path = Path(raw)
        if not path.is_absolute():
            path = (PROJECT_DIR / path).resolve()
        return path
    return (PROJECT_DIR / DEFAULT_NEW_WORKBOOK_NAME).resolve()


def template_path() -> str:
    return str(active_workbook_path())


def _iter_project_paths() -> list[Path]:
    found: list[Path] = []
    for root, dirs, files in os.walk(PROJECT_DIR):
        dirs[:] = [name for name in dirs if name not in IGNORED_SCAN_DIRS]
        root_path = Path(root)
        for name in files:
            found.append(root_path / name)
    return found


def list_workbooks() -> list[dict[str, str]]:
    workbooks: list[dict[str, str]] = []
    for path in _iter_project_paths():
        if path.suffix.lower() not in EXCEL_SUFFIXES:
            continue
        workbooks.append(
            {
                "path": str(path),
                "name": path.name,
                "relative": str(path.relative_to(PROJECT_DIR)),
            }
        )
    return sorted(workbooks, key=lambda item: item["relative"].lower())


def list_folders() -> list[dict[str, str]]:
    folders: list[dict[str, str]] = []
    for root, dirs, _ in os.walk(PROJECT_DIR):
        dirs[:] = [name for name in dirs if name not in IGNORED_SCAN_DIRS]
        path = Path(root)
        if path == PROJECT_DIR:
            continue
        folders.append(
            {
                "path": str(path),
                "name": path.name,
                "relative": str(path.relative_to(PROJECT_DIR)),
            }
        )
    return sorted(folders, key=lambda item: item["relative"].lower())


def normalize_workbook_name(name_value: str) -> str:
    name = clean_text(name_value) or DEFAULT_NEW_WORKBOOK_NAME
    for char in '<>:"/\\|?*':
        name = name.replace(char, "-")
    if Path(name).suffix.lower() not in EXCEL_SUFFIXES:
        name = f"{name}.xlsx"
    return name


def set_active_workbook(path_value: str) -> Path:
    raw = clean_text(path_value)
    if not raw:
        raise ValueError("Selecione uma planilha ou informe um caminho vÃ¡lido.")
    path = Path(raw)
    if not path.is_absolute():
        path = PROJECT_DIR / path
    path = path.resolve()
    if path.suffix.lower() not in EXCEL_SUFFIXES:
        raise ValueError("Use uma planilha .xlsx ou .xlsm.")
    if not path.parent.exists():
        raise FileNotFoundError(f"Pasta nÃ£o encontrada: {path.parent}")
    config = _read_config()
    config["active_workbook"] = str(path)
    _write_config(config)
    return path


def set_active_workbook_from_folder(folder_value: str, workbook_name: str) -> Path:
    raw_folder = clean_text(folder_value)
    if not raw_folder:
        raise ValueError("Selecione uma pasta ou informe um caminho de pasta vÃ¡lido.")
    folder = Path(raw_folder)
    if not folder.is_absolute():
        folder = PROJECT_DIR / folder
    folder = folder.resolve()
    if not folder.exists() or not folder.is_dir():
        raise FileNotFoundError(f"Pasta nÃ£o encontrada: {folder}")
    path = folder / normalize_workbook_name(workbook_name)
    ensure_workbook_exists(path)
    config = _read_config()
    config["active_workbook"] = str(path)
    _write_config(config)
    return path


def _safe_sheet_title(label: str) -> str:
    title = INVALID_SHEET_CHARS_RE.sub("-", clean_text(label) or "Categoria").strip("' ")
    return (title or "Categoria")[:31]


def _sheet_title_for_category(category: dict[str, Any]) -> str:
    return _safe_sheet_title(category.get("sheet_name") or category.get("label") or "Categoria")


def _sheet_for_category(workbook, category: dict[str, Any]):
    title = _sheet_title_for_category(category)
    if title in workbook.sheetnames:
        return workbook[title]
    ws = workbook.create_sheet(title=title)
    _write_headers(ws, get_banco_fields(category["key"]))
    return ws


def _headers_from_fields(fields: list[dict[str, Any]]) -> list[str]:
    headers = ["DESCRIÃ‡ÃƒO PRIMÃRIA", "DESCRIÃ‡ÃƒO SECUNDÃRIA"]
    for field in fields:
        header = header_for_field(field["label"], field["scope"])
        headers.append(header)
    return headers


def _style_header_cell(cell, fill_color: str) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=fill_color)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def create_blank_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    default_sheet = workbook.active
    default_sheet.title = DEFAULT_CATEGORY_LABEL
    _write_headers(default_sheet, get_banco_fields(DEFAULT_CATEGORY_KEY))
    registration_sheet = workbook.create_sheet(REGISTRATION_SHEET_NAME)
    registration_sheet.sheet_state = "hidden"
    registration_sheet.append(["category_key", "category_label", "sheet", "row", "saved_at"])
    workbook.save(path)
    workbook.close()


def ensure_workbook_exists(path: Path | None = None) -> Path:
    workbook = (path or active_workbook_path()).resolve()
    if workbook.exists():
        return workbook
    if DEFAULT_TEMPLATE_PATH.exists():
        workbook.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(DEFAULT_TEMPLATE_PATH, workbook)
        return workbook
    create_blank_workbook(workbook)
    return workbook


def _write_headers(ws, fields: list[dict[str, Any]]) -> None:
    headers = _headers_from_fields(fields)
    for index, header in enumerate(headers, start=1):
        ws.cell(1, index).value = header
        _style_header_cell(ws.cell(1, index), "0F172A" if index <= 2 else "1D4ED8")
    for index in range(len(headers) + 1, ws.max_column + 1):
        ws.cell(1, index).value = None
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    for field_index, field in enumerate(fields):
        value_column = 3 + field_index
        ws.column_dimensions[get_column_letter(value_column)].width = max(20, len(field["label"]) + 4)


def _candidate_header_rows(ws) -> list[int]:
    rows: list[int] = []
    tables = getattr(ws, "tables", None) or {}
    for table in tables.values():
        ref = getattr(table, "ref", table)
        if isinstance(ref, str):
            _, min_row, _, _ = range_boundaries(ref)
            rows.append(min_row)
    rows.extend([1, 2, 3])

    unique_rows: list[int] = []
    for row in rows:
        if row >= 1 and row not in unique_rows and row <= max(ws.max_row, 3):
            unique_rows.append(row)
    return unique_rows


def _detect_header_row(ws) -> int:
    best_row = 1
    best_score = -1
    for row in _candidate_header_rows(ws):
        score = 0
        for column in range(1, ws.max_column + 1):
            if clean_text(ws.cell(row, column).value):
                score += 1
        if score > best_score:
            best_score = score
            best_row = row
    return best_row


def _table_for_sheet(ws):
    tables = getattr(ws, "tables", None) or {}
    for table in tables.values():
        return table
    return None


def _table_bounds(ws) -> tuple[int, int, int, int] | None:
    table = _table_for_sheet(ws)
    if table is None:
        return None
    ref = getattr(table, "ref", None)
    if not ref:
        return None
    return range_boundaries(ref)


def _row_has_values(ws, row: int, start_column: int, end_column: int) -> bool:
    for column in range(start_column, end_column + 1):
        if clean_text(ws.cell(row, column).value):
            return True
    return False


def _next_available_row(ws) -> int:
    primary_column, _, _ = _resolve_description_columns(ws, create_missing=False)
    if primary_column:
        bounds = _table_bounds(ws)
        if bounds is not None:
            _, min_row, _, max_row = bounds
            start_row = max(FIRST_DATA_ROW, min_row + 1)
            for row in range(start_row, max_row + 1):
                if not clean_text(ws.cell(row, primary_column).value):
                    return row
            return max_row + 1

        header_row = _detect_header_row(ws)
        start_row = max(FIRST_DATA_ROW, header_row + 1)
        row = start_row
        while clean_text(ws.cell(row, primary_column).value):
            row += 1
        return row

    header_row = _detect_header_row(ws)
    return max(FIRST_DATA_ROW, header_row + 1)


def _numeric_sku_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = clean_text(value)
    if re.fullmatch(r"\d+", text):
        return text
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    return ""


def _initial_sku_for_prefix(code_prefix: str) -> str:
    return f"{code_prefix}0001"


def _next_sequential_sku(
    ws,
    sku_column: int,
    category: dict[str, Any],
    fields: list[dict[str, Any]],
    data: Any,
) -> str:
    code_prefix = pn_code_prefix(category, fields, data)
    last_code = ""
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        code = _numeric_sku_text(ws.cell(row, sku_column).value)
        if code and code.startswith(code_prefix):
            last_code = code
    if not last_code:
        return _initial_sku_for_prefix(code_prefix)
    return str(int(last_code) + 1).zfill(len(last_code))


def _expand_table_to_row(ws, row: int) -> None:
    table = _table_for_sheet(ws)
    if table is None:
        return
    ref = getattr(table, "ref", None)
    if not ref:
        return
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    if row <= max_row:
        return
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{row}"


def _code_header_for_field(field: dict[str, Any]) -> str:
    return f"{header_for_field(field['label'], field['scope'])} CÃ“DIGO"


def _worksheet_uses_code_columns(ws) -> bool:
    header_row = _detect_header_row(ws)
    for column in range(3, ws.max_column + 1):
        if "CODIGO" in normalize_label(ws.cell(header_row, column).value):
            return True
    return False


def _header_column_map(ws) -> dict[str, int]:
    header_row = _detect_header_row(ws)
    mapping: dict[str, int] = {}
    for column in range(1, ws.max_column + 1):
        normalized = normalize_label(ws.cell(header_row, column).value)
        if normalized and normalized not in mapping:
            mapping[normalized] = column
    return mapping


def _last_header_column(ws, header_row: int | None = None) -> int:
    header_row = header_row or _detect_header_row(ws)
    for column in range(ws.max_column, 0, -1):
        if clean_text(ws.cell(header_row, column).value):
            return column
    return 0


def _append_header_cell(ws, header: str, header_row: int | None = None) -> int:
    header_row = header_row or _detect_header_row(ws)
    column = _last_header_column(ws, header_row) + 1
    ws.cell(header_row, column).value = header
    _style_header_cell(ws.cell(header_row, column), "1D4ED8")
    ws.column_dimensions[get_column_letter(column)].width = max(20, len(clean_text(header)) + 4)
    return column


def _resolve_header_column(ws, header: str, create_missing: bool = False) -> int | None:
    header_row = _detect_header_row(ws)
    header_map = _header_column_map(ws)
    column = header_map.get(normalize_label(header))
    if column is not None or not create_missing:
        return column
    return _append_header_cell(ws, header, header_row)


def _resolve_description_columns(ws, create_missing: bool = False) -> tuple[int | None, int | None, int | None]:
    primary_column = _resolve_header_column(ws, DESCRIPTION_PRIMARY_HEADER, create_missing)
    secondary_column = _resolve_header_column(ws, DESCRIPTION_SECONDARY_HEADER, create_missing)
    suffix_column = _resolve_header_column(ws, DESCRIPTION_SUFFIX_HEADER, create_missing)
    return primary_column, secondary_column, suffix_column


def _resolve_field_columns(ws, field: dict[str, Any], create_missing: bool = False) -> tuple[int | None, int | None]:
    header_map = _header_column_map(ws)
    value_header = normalize_label(header_for_field(field["label"], field["scope"]))
    code_header = normalize_label(_code_header_for_field(field))

    value_column = header_map.get(value_header)
    code_column = header_map.get(code_header)

    uses_code_columns = _worksheet_uses_code_columns(ws)
    if value_column is None and create_missing:
        value_column = _append_header_cell(ws, header_for_field(field["label"], field["scope"]))
        if uses_code_columns:
            code_column = _append_header_cell(ws, _code_header_for_field(field))

    return value_column, code_column


def _resolve_field_column_map(
    ws,
    fields: list[dict[str, Any]],
    create_missing: bool = False,
) -> dict[str, tuple[int | None, int | None]]:
    header_map = _header_column_map(ws)
    uses_code_columns = any("CODIGO" in header for header in header_map)
    column_map: dict[str, tuple[int | None, int | None]] = {}

    for field in fields:
        value_header_text = header_for_field(field["label"], field["scope"])
        code_header_text = _code_header_for_field(field)
        value_header = normalize_label(value_header_text)
        code_header = normalize_label(code_header_text)

        value_column = header_map.get(value_header)
        code_column = header_map.get(code_header)

        if value_column is None and create_missing:
            value_column = _append_header_cell(ws, value_header_text)
            header_map[value_header] = value_column
            if uses_code_columns:
                code_column = _append_header_cell(ws, code_header_text)
                header_map[code_header] = code_column

        column_map[field["key"]] = (value_column, code_column)

    return column_map


def _save_workbook_atomic(wb, target_path: Path) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temp_name = tempfile.mkstemp(prefix="_cadastro_app_save_", suffix=target_path.suffix or ".xlsx", dir=str(target_path.parent))
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        wb.save(temp_path)
        os.replace(temp_path, target_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _save_workbook_preserving_package(wb, target_path: Path) -> None:
    _save_workbook_atomic(wb, target_path)


def _preserve_workbook_parts(source_path: Path, target_path: Path, prefixes: tuple[str, ...]) -> None:
    if not source_path.exists() or not target_path.exists():
        return

    source_members: dict[str, zipfile.ZipInfo] = {}
    source_payloads: dict[str, bytes] = {}
    with zipfile.ZipFile(source_path, "r") as source_zip:
        for member in source_zip.infolist():
            if any(member.filename.startswith(prefix) for prefix in prefixes):
                source_members[member.filename] = member
                source_payloads[member.filename] = source_zip.read(member.filename)

    if not source_payloads:
        return

    fd, temp_name = tempfile.mkstemp(prefix="_cadastro_app_zip_", suffix=target_path.suffix or ".xlsx", dir=str(target_path.parent))
    os.close(fd)
    temp_path = Path(temp_name)

    try:
        with zipfile.ZipFile(target_path, "r") as target_zip, zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_DEFLATED) as output_zip:
            target_names = set(target_zip.namelist())
            for member in target_zip.infolist():
                payload = source_payloads.get(member.filename, target_zip.read(member.filename))
                info = source_members.get(member.filename, member)
                output_zip.writestr(info, payload)
            for filename, info in source_members.items():
                if filename not in target_names:
                    output_zip.writestr(info, source_payloads[filename])
        os.replace(temp_path, target_path)
    finally:
        temp_path.unlink(missing_ok=True)


def _field_response(field: dict[str, Any], index: int) -> dict[str, Any]:
    column = 3 + index
    options = list(field.get("options") or [])
    return {
        "key": field["key"],
        "label": field["label"],
        "header": header_for_field(field["label"], field["scope"]),
        "scope": field["scope"],
        "selection_mode": field.get("selection_mode", SELECTION_MODE_UNITARIA),
        "description_order": _field_description_order(field, index),
        "column": column,
        "letter": get_column_letter(column),
        "options": options,
        "option_rows": [{"row": option_index, "value": value} for option_index, value in enumerate(options, start=1)],
    }


def get_banco_fields(category_key_value: str) -> list[dict[str, Any]]:
    category = _find_category(load_catalog(), category_key_value)
    return [_field_response(field, index) for index, field in enumerate(category.get("fields") or [])]


def get_banco_fields_for_display(category_key_value: str) -> list[dict[str, Any]]:
    category = _find_category(load_catalog(), category_key_value)
    ordered_fields = _ordered_fields_for_description(category.get("fields") or [])
    return [_field_response(field, index) for index, field in enumerate(ordered_fields, start=1)]


def get_conditional_rules(category_key_value: str) -> list[dict[str, Any]]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    fields = category.get("fields") or []
    rules = category.get("conditional_rules") or []
    field_map = {field["key"]: _field_response(field, index) for index, field in enumerate(fields)}
    token_map = {
        field["key"]: {rule_option_token(option): option for option in (field.get("options") or [])}
        for field in fields
    }
    resolved: list[dict[str, Any]] = []
    for rule in rules:
        source_field = field_map.get(rule["source_field_key"])
        target_field = field_map.get(rule["target_field_key"])
        source_value_labels = [
            token_map.get(rule["source_field_key"], {}).get(value, value) for value in (rule.get("source_values") or [])
        ]
        resolved.append(
            {
                "key": rule["key"],
                "source_field_key": rule["source_field_key"],
                "source_field_label": rule["source_field_label"],
                "source_field_scope": rule["source_field_scope"],
                "source_values": list(rule.get("source_values") or []),
                "source_value_labels": source_value_labels,
                "target_field_key": rule["target_field_key"],
                "target_field_label": rule["target_field_label"],
                "target_field_scope": rule["target_field_scope"],
                "action": rule["action"],
                "match_by": rule.get("match_by", "option"),
                "source_field": source_field,
                "target_field": target_field,
            }
        )
    return resolved


def get_conditional_rules_for_form(category_key_value: str) -> list[dict[str, Any]]:
    rules = get_conditional_rules(category_key_value)
    form_rules: list[dict[str, Any]] = []
    for rule in rules:
        source_values = list(rule.get("source_values") or [])
        targets: list[dict[str, Any]] = []
        if rule.get("target_field"):
            targets.append(
                {
                    "label": rule["target_field"]["label"],
                    "scope": rule["target_field"]["scope"],
                    "requiredWhenVisible": rule["action"] == "show",
                }
            )
        elif rule.get("target_field_label"):
            targets.append(
                {
                    "label": rule["target_field_label"],
                    "scope": rule["target_field_scope"],
                    "requiredWhenVisible": rule["action"] == "show",
                }
            )
        if not targets:
            continue
        form_rules.append(
            {
                "sourceLabel": rule["source_field_label"],
                "sourceScope": rule["source_field_scope"],
                "action": rule["action"],
                "mode": "showWhen" if rule["action"] == "show" else "hideWhen",
                "matchBy": rule.get("match_by", "option"),
                "values": source_values,
                "targets": targets,
            }
        )
    return form_rules


def _sheet_for_view(workbook, category: dict[str, Any]):
    title = _sheet_title_for_category(category)
    if title in workbook.sheetnames:
        return workbook[title]
    return None


def list_saved_registrations(category_key_value: str) -> dict[str, Any]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    selected = {"key": category["key"], "label": category["label"]}
    fields = get_banco_fields(category["key"])
    workbook = active_workbook_path()
    sheet_name = _sheet_title_for_category(category)
    result = {
        "category": selected,
        "fields": fields,
        "records": [],
        "workbook_path": str(workbook),
        "workbook_exists": workbook.exists(),
        "sheet_name": sheet_name,
        "sheet_exists": False,
        "workbook_locked": False,
        "error_message": "",
    }
    if not workbook.exists():
        return result

    workbook_mtime_ns = workbook.stat().st_mtime_ns
    cache_key = (str(workbook.resolve()), category["key"], workbook_mtime_ns)
    cached = _REGISTRATION_CACHE.get(cache_key)
    if cached is not None:
        return deepcopy(cached)

    try:
        wb = load_workbook(workbook, data_only=True)
    except PermissionError:
        result["workbook_locked"] = True
        result["error_message"] = (
            "A planilha estÃ¡ aberta em outro programa e nÃ£o pode ser lida agora. "
            "Feche o Excel para visualizar os cadastros."
        )
        return result

    try:
        ws = _sheet_for_view(wb, category)
        if ws is None:
            return result

        primary_column, secondary_column, suffix_column = _resolve_description_columns(ws, create_missing=False)
        field_columns = _resolve_field_column_map(ws, fields, create_missing=False)
        result["sheet_name"] = ws.title
        result["sheet_exists"] = True
        bounds = _table_bounds(ws)
        last_row = bounds[3] if bounds else ws.max_row
        for row_index, row_values in enumerate(
            ws.iter_rows(min_row=FIRST_DATA_ROW, max_row=last_row, values_only=True),
            start=FIRST_DATA_ROW,
        ):
            descricao_primaria = clean_text(row_values[primary_column - 1]) if primary_column and primary_column <= len(row_values) else ""
            descricao_secundaria = compose_secondary_description(
                descricao_primaria,
                row_values[secondary_column - 1] if secondary_column and secondary_column <= len(row_values) else "",
            )
            sufixo = clean_text(row_values[suffix_column - 1]) if suffix_column and suffix_column <= len(row_values) else ""
            values: list[dict[str, str]] = []
            has_field_value = False

            for field_index, field in enumerate(fields):
                value_column, _ = field_columns.get(field["key"], (None, None))
                value = clean_text(row_values[value_column - 1]) if value_column and value_column <= len(row_values) else ""
                if value:
                    has_field_value = True
                values.append({"key": field["key"], "label": field["label"], "value": value})

            if not descricao_primaria and not descricao_secundaria and not has_field_value:
                continue

            result["records"].append(
                {
                    "row": row_index,
                    "descricao_primaria": descricao_primaria,
                    "descricao_secundaria": descricao_secundaria,
                    "sufixo": sufixo,
                    "values": values,
                }
            )
        _REGISTRATION_CACHE[cache_key] = deepcopy(result)
        return result
    finally:
        wb.close()

def sync_workbook_headers(workbook: Path, category_key_value: str) -> None:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    fields = get_banco_fields(category["key"])
    wb = _load(workbook)
    try:
        ws = _sheet_for_category(wb, category)
        primary_column, secondary_column, suffix_column = _resolve_description_columns(ws, create_missing=True)
        ws.freeze_panes = "A2"
        if primary_column:
            ws.column_dimensions[get_column_letter(primary_column)].width = 42
        if secondary_column:
            ws.column_dimensions[get_column_letter(secondary_column)].width = 42
        if suffix_column:
            ws.column_dimensions[get_column_letter(suffix_column)].width = 18
        _resolve_field_column_map(ws, fields, create_missing=True)
        _registration_sheet(wb)
        _save_workbook_atomic(wb, workbook)
    except PermissionError as exc:
        raise PermissionError(
            "NÃ£o consegui salvar a planilha. Feche o Excel se ela estiver aberta e tente novamente."
        ) from exc
    finally:
        wb.close()


def sync_workbook_structure(category_key_value: str) -> None:
    sync_workbook_headers(active_workbook_path(), category_key_value)


def _registration_sheet(workbook):
    if REGISTRATION_SHEET_NAME not in workbook.sheetnames:
        ws = workbook.create_sheet(REGISTRATION_SHEET_NAME)
        ws.sheet_state = "hidden"
        ws.append(["category_key", "category_label", "sheet", "row", "saved_at"])
        return ws
    ws = workbook[REGISTRATION_SHEET_NAME]
    if ws.max_row == 1 and clean_text(ws.cell(1, 1).value) != "category_key":
        ws.delete_rows(1, ws.max_row)
        ws.append(["category_key", "category_label", "sheet", "row", "saved_at"])
    ws.sheet_state = "hidden"
    return ws


def _draft_sheet(workbook):
    headers = ["draft_id", "category_key", "category_label", "sheet", "saved_at", "descricao_primaria", "data_json"]
    if DRAFT_SHEET_NAME not in workbook.sheetnames:
        ws = workbook.create_sheet(DRAFT_SHEET_NAME)
        ws.sheet_state = "hidden"
        ws.append(headers)
        return ws
    ws = workbook[DRAFT_SHEET_NAME]
    for column, header in enumerate(headers, start=1):
        if clean_text(ws.cell(1, column).value) != header:
            ws.cell(1, column).value = header
    ws.sheet_state = "hidden"
    return ws


def _draft_row(ws, draft_id: str) -> int | None:
    draft_id = clean_text(draft_id)
    if not draft_id:
        return None
    for row in range(2, ws.max_row + 1):
        if clean_text(ws.cell(row, 1).value) == draft_id:
            return row
    return None


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.protection:
            target.protection = copy(source.protection)
    if source_row in ws.row_dimensions:
        source_dimension = ws.row_dimensions[source_row]
        target_dimension = ws.row_dimensions[target_row]
        if source_dimension.height is not None:
            target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        target_dimension.thickBot = source_dimension.thickBot
        target_dimension.thickTop = source_dimension.thickTop


def _copy_row_formulas(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        value = source.value
        if isinstance(value, ArrayFormula):
            target.value = ArrayFormula(ref=target.coordinate, text=value.text)
        elif source.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
            target.value = value


def _first_code(value: str, fallback: int) -> int:
    match = VALUE_RE.match(clean_text(value))
    if match:
        return int(match.group(1))
    return fallback


def _serialize_field_values(field: dict[str, Any], data: Any) -> list[str]:
    raw_values: list[str] = []
    key = field["key"]
    if hasattr(data, "getlist"):
        raw_values.extend(data.getlist(key))
    else:
        value = data.get(key)
        if isinstance(value, list):
            raw_values.extend(value)
        elif value is not None:
            raw_values.append(value)

    if field.get("selection_mode") != SELECTION_MODE_MULTIPLA:
        raw_values = raw_values[:1]

    cleaned_values = [normalize_option_text(value) for value in raw_values if clean_text(value)]
    if not cleaned_values:
        return []

    options = field.get("options") or []
    submitted_identities = {option_identity(value): value for value in cleaned_values}
    ordered: list[str] = []
    used: set[str] = set()

    for option in options:
        identity = option_identity(option)
        if identity in submitted_identities and identity not in used:
            ordered.append(option)
            used.add(identity)

    for value in cleaned_values:
        identity = option_identity(value)
        if identity not in used:
            ordered.append(value)
            used.add(identity)

    if field.get("key") == DISTANCIA_PE_KEY:
        return _order_distancia_pe_values(ordered)
    return ordered

def pn_category_code(category: dict[str, Any]) -> str:
    for value in (category.get("label"), category.get("sheet_name")):
        match = re.search(r"\d+", clean_text(value))
        if match:
            return f"{int(match.group(0)):02d}"
    return "00"


def _selected_group_code(fields: list[dict[str, Any]], data: Any) -> str:
    if data is not None and hasattr(data, "get"):
        explicit_code = _pn_group_code(data.get(PN_GROUP_FORM_KEY))
        if explicit_code:
            return explicit_code
    if isinstance(data, dict):
        values = data.get(PN_GROUP_FORM_KEY)
        if isinstance(values, list):
            for value in values:
                explicit_code = _pn_group_code(value)
                if explicit_code:
                    return explicit_code
        else:
            explicit_code = _pn_group_code(values)
            if explicit_code:
                return explicit_code

    group_field = _find_field_by_normalized_label(fields, {"GRUPO"})
    if group_field is None:
        return ""
    for value in _serialize_field_values(group_field, data):
        match = re.match(r"^\s*(\d+)", clean_text(value))
        if match:
            return f"{int(match.group(1)):02d}"
    return ""


def _pn_group_prefix_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    for group in load_catalog().get("pn_groups") or _default_pn_groups():
        code = _pn_group_code(group.get("code"))
        if not code:
            continue
        for prefix in group.get("prefixes") or []:
            normalized = normalize_label(prefix)
            if normalized:
                mapping[normalized] = code
    return mapping or dict(PN_GROUP_BY_PREFIX)


def pn_group_code(fields: list[dict[str, Any]], data: Any) -> str:
    explicit_group = _selected_group_code(fields, data)
    if explicit_group:
        return explicit_group

    prefix_field = _find_field_by_normalized_label(fields, {"PREFIXO", "PRE FIXO", "PRÉ FIXO"})
    if prefix_field is None:
        return "10"
    prefix_map = _pn_group_prefix_map()
    for value in _selected_option_labels(prefix_field, data):
        normalized = normalize_label(value)
        if normalized in prefix_map:
            return prefix_map[normalized]
        for known_prefix, group_code in prefix_map.items():
            if normalized.startswith(f"{known_prefix} ") or normalized.startswith(known_prefix):
                return group_code
    return "10"


def pn_code_prefix(category: dict[str, Any], fields: list[dict[str, Any]], data: Any) -> str:
    return f"{pn_group_code(fields, data)}{pn_category_code(category)}"


def _distancia_pe_order(value: Any) -> tuple[int, int, str]:
    label = normalize_option_label(option_label(value))
    vao_order = {
        "PRIMEIRO": 1,
        "SEGUNDO": 2,
        "TERCEIRO": 3,
        "QUARTO": 4,
    }
    for word, order in vao_order.items():
        if word in label and "VAO" in label:
            return (0, order, label)
    return (1, 999, label)


def _order_distancia_pe_values(values: list[str]) -> list[str]:
    return sorted(values, key=_distancia_pe_order)


def _format_distancia_pe_value(values: list[str]) -> str:
    values = _order_distancia_pe_values(values)
    joined = " | ".join(values).strip()
    if not joined:
        return ""

    normalized_joined = normalize_label(joined)
    normalized_prefix = normalize_label(DISTANCIA_PE_PREFIX)
    if normalized_joined.startswith(normalized_prefix):
        return joined
    return f"{DISTANCIA_PE_PREFIX} {joined}".strip()


def _format_field_saved_value(field: dict[str, Any], values: list[str]) -> str:
    if field.get("key") == DISTANCIA_PE_KEY:
        return _format_distancia_pe_value(values)
    return " | ".join(values)


def _format_field_description(field: dict[str, Any], values: list[str]) -> str:
    if field.get("key") == DISTANCIA_PE_KEY:
        values = _order_distancia_pe_values(values)
        label = ", ".join(option_label(value) for value in values)
        if label:
            prefix = DISTANCIA_PE_PREFIX
            if normalize_label(label).startswith(normalize_label(prefix)):
                return label
            return f"{prefix} {label}".strip()
        return label
    label = " / ".join(option_label(value) for value in values)
    return label


def build_descriptions(
    fields: list[dict[str, Any]],
    data: Any,
    category_key_value: str = "",
) -> dict[str, str]:
    primary_parts: list[str] = []
    secondary_parts: list[str] = []
    secondary_codes: list[str] = []
    effective_scopes = _effective_field_scopes(fields, category_key_value, data)

    for field in _ordered_fields_for_description(fields):
        values = _serialize_field_values(field, data)
        if not values:
            continue
        label = _format_field_description(field, values)
        effective_scope = effective_scopes.get(field["key"], field["scope"])
        if effective_scope == "primaria":
            primary_parts.append(label)
        elif effective_scope == "secundaria":
            secondary_parts.append(label)
            secondary_codes.extend(code for code in (option_code(value) for value in values) if code)

    primary_description_base = " ".join(primary_parts).strip()
    suffix = ""
    if secondary_codes:
        suffix = ".".join(secondary_codes)
    secondary_description = " ".join(secondary_parts).strip()

    return {
        "primaria": primary_description_base,
        "secundaria": compose_secondary_description(primary_description_base, secondary_description),
        "sufixo": suffix,
    }


def _submitted_values(fields: list[dict[str, Any]], data: Any) -> dict[str, str]:
    return {field["key"]: " | ".join(_serialize_field_values(field, data)) for field in fields}


def _field_values_by_key(fields: list[dict[str, Any]], data: Any) -> dict[str, list[str]]:
    return {field["key"]: _serialize_field_values(field, data) for field in fields}


def _has_any_option_code(values: list[str], allowed_codes: set[str]) -> bool:
    return any(option_code(value) in allowed_codes for value in values)


def _validate_banco_dependencies(fields: list[dict[str, Any]], data: Any) -> None:
    values_by_key = _field_values_by_key(fields, data)
    pre_fixo_values = values_by_key.get("pre_fixo", [])
    veiculo_values = values_by_key.get("veiculo", [])

    if _has_any_option_code(pre_fixo_values, {"4", "5", "6"}) and not veiculo_values:
        raise ValueError(
            "O campo VEÃCULO Ã© obrigatÃ³rio quando PRE FIXO for "
            "BCO CARONA ORIGINAL, BCO MOTORISTA ORIGINAL ou BCO ORIGINAL."
        )


def _has_any_value(values: dict[str, str]) -> bool:
    return any(values.values())


def _selected_conditional_tokens(field: dict[str, Any], data: Any) -> list[str]:
    tokens: list[str] = []
    for value in _serialize_field_values(field, data):
        token = rule_option_token(value)
        if token:
            tokens.append(token)
    return tokens


def _selected_conditional_prefixes(field: dict[str, Any], data: Any) -> list[str]:
    prefixes: list[str] = []
    for value in _serialize_field_values(field, data):
        prefix = option_code(value)
        if not prefix:
            match = re.match(r"^\s*(\d+)", clean_text(value))
            prefix = match.group(1) if match else ""
        if prefix:
            prefixes.append(prefix)
    return prefixes


def _rule_match_values(rule: dict[str, Any]) -> list[str]:
    values = [clean_text(value) for value in rule.get("source_values") or []]
    if clean_text(rule.get("match_by")).lower() == "prefix":
        normalized: list[str] = []
        for value in values:
            code = option_code(value)
            if code:
                normalized.append(code)
                continue
            match = re.match(r"^\s*(\d+)", value)
            if match:
                normalized.append(match.group(1))
                continue
            token = normalize_label(value).replace(" ", "")
            if token:
                normalized.append(token)
        return normalized
    return [token for token in (rule_option_token(value) for value in values) if token]


def _rule_matches(field: dict[str, Any], data: Any, rule: dict[str, Any]) -> bool:
    match_by = clean_text(rule.get("match_by")).lower()
    selected_values = (
        _selected_conditional_prefixes(field, data)
        if match_by == "prefix"
        else _selected_conditional_tokens(field, data)
    )
    if not selected_values:
        return False

    rule_values = _rule_match_values(rule)
    if not rule_values:
        return False

    for token in selected_values:
        for value in rule_values:
            if token == value or token.startswith(value):
                return True
    return False


def _combined_conditional_rules(category_key_value: str) -> list[dict[str, Any]]:
    rules_by_key: dict[str, dict[str, Any]] = {}

    def add_rule(rule: dict[str, Any], default_match_by: str = "option") -> None:
        copied = deepcopy(rule)
        if not clean_text(copied.get("match_by")) and copied.get("source_field_key") == "pre_fixo":
            copied["match_by"] = "prefix"
        if not clean_text(copied.get("match_by")):
            copied["match_by"] = default_match_by
        key = clean_text(copied.get("key")) or "|".join(
            [
                clean_text(copied.get("source_field_key")),
                clean_text(copied.get("target_field_key")),
                clean_text(copied.get("action")),
                "|".join(clean_text(value) for value in copied.get("source_values") or []),
            ]
        )
        rules_by_key[key] = copied

    for rule in DEFAULT_CONDITIONAL_RULES:
        add_rule(rule)
    for rule in get_conditional_rules(category_key_value):
        add_rule(rule, "option")
    return list(rules_by_key.values())


def _effective_field_scopes(
    fields: list[dict[str, Any]],
    category_key_value: str,
    data: Any,
) -> dict[str, str]:
    scopes = {field["key"]: field["scope"] for field in fields}
    if not clean_text(category_key_value):
        return scopes

    field_map = {field["key"]: field for field in fields}
    for rule in _combined_conditional_rules(category_key_value):
        action = clean_text(rule.get("action")).lower()
        if action not in {"set_primary", "set_secondary"}:
            continue
        source_field = field_map.get(clean_text(rule.get("source_field_key")))
        target_key = clean_text(rule.get("target_field_key"))
        if source_field is None or target_key not in field_map:
            continue
        if _rule_matches(source_field, data, rule):
            scopes[target_key] = "primaria" if action == "set_primary" else "secundaria"
    return scopes


def _visible_field_keys(fields: list[dict[str, Any]], category_key_value: str, data: Any) -> set[str]:
    field_map = {field["key"]: field for field in fields}
    show_matches: dict[str, list[bool]] = {field["key"]: [] for field in fields}
    hide_matches: dict[str, list[bool]] = {field["key"]: [] for field in fields}

    for rule in _combined_conditional_rules(category_key_value):
        action = clean_text(rule.get("action")).lower()
        if action in {"set_primary", "set_secondary"}:
            continue
        source_key = clean_text(rule.get("source_field_key"))
        target_key = clean_text(rule.get("target_field_key"))
        source_field = field_map.get(source_key)
        target_field = field_map.get(target_key)
        if source_field is None or target_field is None:
            continue

        matches = _rule_matches(source_field, data, rule)
        if action == "show":
            show_matches[target_key].append(matches)
        else:
            hide_matches[target_key].append(matches)

    visible: set[str] = set()
    for field in fields:
        field_key = field["key"]
        show_ok = any(show_matches[field_key]) if show_matches[field_key] else True
        hide_hit = any(hide_matches[field_key])
        if show_ok and not hide_hit:
            visible.add(field_key)
    return visible


def _validate_visible_field_requirements(
    fields: list[dict[str, Any]],
    category_key_value: str,
    data: Any,
) -> None:
    visible_keys = _visible_field_keys(fields, category_key_value, data)
    missing_fields: list[str] = []
    for field in fields:
        if field["key"] not in visible_keys:
            continue
        if _serialize_field_values(field, data):
            continue
        missing_fields.append(field["label"])

    if missing_fields:
        labels = ", ".join(missing_fields)
        raise ValueError(f"Preencha os campos visÃ­veis antes de salvar: {labels}.")


def _draft_payload_groups(payload_json: str | dict[str, Any]) -> dict[str, list[str]]:
    if isinstance(payload_json, str):
        raw = json.loads(payload_json) if clean_text(payload_json) else {}
    else:
        raw = payload_json or {}
    groups = raw.get("groups") if isinstance(raw, dict) else {}
    if not isinstance(groups, dict):
        return {}
    result: dict[str, list[str]] = {}
    for key, values in groups.items():
        clean_key = clean_text(key)
        if not clean_key:
            continue
        if isinstance(values, list):
            result[clean_key] = [clean_text(value) for value in values]
        else:
            result[clean_key] = [clean_text(values)]
    return result


def save_registration_draft(category_key_value: str, payload_json: str, draft_id: str = "") -> dict[str, Any]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    fields = get_banco_fields(category["key"])
    groups = _draft_payload_groups(payload_json)
    if not groups:
        raise ValueError("Rascunho vazio.")

    descriptions = build_descriptions(fields, groups, category["key"])
    workbook = ensure_workbook_exists()
    workbook_source = _copy_to_temp(workbook)
    wb = _load(workbook_source)
    backup_path = None
    try:
        draft_id = clean_text(draft_id) or uuid.uuid4().hex[:12]
        ws = _draft_sheet(wb)
        row = _draft_row(ws, draft_id) or (ws.max_row + 1)
        saved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        data_json = json.dumps({"category": category["key"], "groups": groups}, ensure_ascii=False)
        backup_path = _backup_workbook(workbook, "rascunho")
        ws.cell(row, 1).value = draft_id
        ws.cell(row, 2).value = category["key"]
        ws.cell(row, 3).value = category["label"]
        ws.cell(row, 4).value = _safe_sheet_title(category.get("sheet_name") or category["label"])
        ws.cell(row, 5).value = saved_at
        ws.cell(row, 6).value = descriptions.get("primaria") or "(sem descrição primária)"
        ws.cell(row, 7).value = data_json
        _save_workbook_preserving_package(wb, workbook)
        return {
            "draft_id": draft_id,
            "category_key": category["key"],
            "category_label": category["label"],
            "saved_at": saved_at,
            "descricao_primaria": descriptions.get("primaria") or "",
            "path": str(workbook),
            "backup": str(backup_path) if backup_path else "",
        }
    finally:
        wb.close()
        cleanup_path = getattr(wb, "_cadastro_cleanup_path", None)
        if cleanup_path is not None:
            try:
                Path(cleanup_path).unlink(missing_ok=True)
            except OSError:
                pass
        try:
            workbook_source.unlink(missing_ok=True)
        except OSError:
            pass


def list_registration_drafts() -> list[dict[str, Any]]:
    workbook = ensure_workbook_exists()
    wb = load_workbook(workbook, data_only=True, read_only=True)
    drafts: list[dict[str, Any]] = []
    try:
        if DRAFT_SHEET_NAME not in wb.sheetnames:
            return []
        ws = wb[DRAFT_SHEET_NAME]
        for row in range(2, ws.max_row + 1):
            draft_id = clean_text(ws.cell(row, 1).value)
            if not draft_id:
                continue
            drafts.append(
                {
                    "draft_id": draft_id,
                    "category_key": clean_text(ws.cell(row, 2).value),
                    "category_label": clean_text(ws.cell(row, 3).value),
                    "sheet": clean_text(ws.cell(row, 4).value),
                    "saved_at": clean_text(ws.cell(row, 5).value),
                    "descricao_primaria": clean_text(ws.cell(row, 6).value),
                }
            )
    finally:
        wb.close()
    return sorted(drafts, key=lambda item: item.get("saved_at", ""), reverse=True)


def get_registration_draft(draft_id: str) -> dict[str, Any] | None:
    workbook = ensure_workbook_exists()
    wb = load_workbook(workbook, data_only=True, read_only=True)
    try:
        if DRAFT_SHEET_NAME not in wb.sheetnames:
            return None
        ws = wb[DRAFT_SHEET_NAME]
        for row in range(2, ws.max_row + 1):
            if clean_text(ws.cell(row, 1).value) != clean_text(draft_id):
                continue
            payload = json.loads(clean_text(ws.cell(row, 7).value) or "{}")
            return {
                "draft_id": clean_text(ws.cell(row, 1).value),
                "category_key": clean_text(ws.cell(row, 2).value),
                "category_label": clean_text(ws.cell(row, 3).value),
                "saved_at": clean_text(ws.cell(row, 5).value),
                "descricao_primaria": clean_text(ws.cell(row, 6).value),
                "groups": _draft_payload_groups(payload),
            }
    finally:
        wb.close()
    return None


def delete_registration_draft(draft_id: str) -> dict[str, Any]:
    draft_id = clean_text(draft_id)
    if not draft_id:
        raise ValueError("Rascunho não informado.")
    workbook = ensure_workbook_exists()
    workbook_source = _copy_to_temp(workbook)
    wb = _load(workbook_source)
    backup_path = None
    try:
        if DRAFT_SHEET_NAME not in wb.sheetnames:
            raise ValueError("Rascunho não encontrado.")
        ws = wb[DRAFT_SHEET_NAME]
        row = _draft_row(ws, draft_id)
        if row is None:
            raise ValueError("Rascunho não encontrado.")
        category_label = clean_text(ws.cell(row, 3).value)
        backup_path = _backup_workbook(workbook, "rascunho_excluido")
        ws.delete_rows(row, 1)
        _save_workbook_preserving_package(wb, workbook)
        return {
            "draft_id": draft_id,
            "category_label": category_label,
            "path": str(workbook),
            "backup": str(backup_path) if backup_path else "",
        }
    finally:
        wb.close()
        cleanup_path = getattr(wb, "_cadastro_cleanup_path", None)
        if cleanup_path is not None:
            try:
                Path(cleanup_path).unlink(missing_ok=True)
            except OSError:
                pass
        try:
            workbook_source.unlink(missing_ok=True)
        except OSError:
            pass


def _row_values(ws, field_columns: dict[str, tuple[int | None, int | None]], row: int) -> dict[str, str]:
    values: dict[str, str] = {}
    for field_key, (value_column, _) in field_columns.items():
        values[field_key] = clean_text(ws.cell(row, value_column).value) if value_column else ""
    return values


def _find_duplicate_registration(
    ws,
    fields: list[dict[str, Any]],
    field_columns: dict[str, tuple[int | None, int | None]],
    data: Any,
) -> int | None:
    submitted = _submitted_values(fields, data)
    if not _has_any_value(submitted):
        raise ValueError("Selecione pelo menos uma opÃ§Ã£o antes de salvar o cadastro.")

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        existing = _row_values(ws, field_columns, row)
        if not _has_any_value(existing):
            continue
        if existing == submitted:
            return row
    return None


def _find_duplicate_registration_by_description(
    ws,
    fields: list[dict[str, Any]],
    field_columns: dict[str, tuple[int | None, int | None]],
    data: Any,
    description_columns: tuple[int | None, int | None, int | None],
    category_key_value: str = "",
) -> int | None:
    submitted = _submitted_values(fields, data)
    if not _has_any_value(submitted):
        raise ValueError("Selecione pelo menos uma opÃƒÂ§ÃƒÂ£o antes de salvar o cadastro.")

    primary_column, secondary_column, suffix_column = description_columns
    if not (primary_column and secondary_column and suffix_column):
        return _find_duplicate_registration(ws, fields, field_columns, data)

    descriptions = build_descriptions(fields, data, category_key_value)
    submitted_primary = normalize_label(descriptions.get("primaria"))
    submitted_secondary = normalize_label(descriptions.get("secundaria"))
    submitted_suffix = clean_text(descriptions.get("sufixo"))

    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        existing = _row_values(ws, field_columns, row)
        if not _has_any_value(existing):
            continue
        existing_primary = normalize_label(ws.cell(row, primary_column).value)
        existing_secondary = normalize_label(ws.cell(row, secondary_column).value)
        existing_suffix = clean_text(ws.cell(row, suffix_column).value)
        if (
            existing_primary == submitted_primary
            and existing_secondary == submitted_secondary
            and existing_suffix == submitted_suffix
        ):
            return row
    return _find_duplicate_registration(ws, fields, field_columns, data)


def _backup_workbook(workbook: Path, suffix: str) -> Path:
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = BACKUP_DIR / f"{workbook.stem}_{suffix}_{stamp}{workbook.suffix}"
    shutil.copy2(workbook, backup_path)
    return backup_path


def _find_field_by_normalized_label(fields: list[dict[str, Any]], labels: set[str]) -> dict[str, Any] | None:
    normalized_labels = {normalize_label(label) for label in labels}
    for field in fields:
        if normalize_label(field.get("label")) in normalized_labels:
            return field
    return None


def _selected_option_labels(field: dict[str, Any], data: Any) -> list[str]:
    return [option_label(value) for value in _serialize_field_values(field, data)]


BOM_FORM_KEY = "possui_bom"


def requires_component_bom(fields: list[dict[str, Any]], data: Any) -> bool:
    del fields
    value = data.get(BOM_FORM_KEY) if hasattr(data, "get") else None
    if isinstance(value, list):
        value = value[0] if value else ""
    if isinstance(value, bool):
        return value
    return normalize_label(value) in {"1", "SIM", "TRUE", "COM BOM", "COM B.O.M."}


def _form_getlist(data: Any, key: str) -> list[str]:
    if hasattr(data, "getlist"):
        return [clean_text(value) for value in data.getlist(key)]
    value = data.get(key) if hasattr(data, "get") else None
    if isinstance(value, list):
        return [clean_text(item) for item in value]
    if value is None:
        return []
    return [clean_text(value)]


def parse_component_lines(data: Any, allow_incomplete: bool = False) -> list[dict[str, Any]]:
    codes = _form_getlist(data, "component_codigo")
    descriptions = _form_getlist(data, "component_descricao")
    units = _form_getlist(data, "component_unidade")
    quantities = _form_getlist(data, "component_quantidade")
    searches = _form_getlist(data, "component_search")
    max_length = max(len(codes), len(descriptions), len(units), len(quantities), len(searches), 0)

    components: list[dict[str, Any]] = []
    incomplete_rows: list[int] = []
    for index in range(max_length):
        code = codes[index] if index < len(codes) else ""
        description = descriptions[index] if index < len(descriptions) else ""
        unit = units[index] if index < len(units) else ""
        quantity_text = quantities[index] if index < len(quantities) else ""
        search_text = searches[index] if index < len(searches) else ""

        if not any([code, description, unit, quantity_text, search_text]):
            continue
        if not code and not allow_incomplete:
            incomplete_rows.append(index + 1)
            continue
        try:
            if allow_incomplete and not clean_text(quantity_text):
                quantity = 1.0
            else:
                quantity = float(quantity_text.replace(".", "").replace(",", ".") if "," in quantity_text else quantity_text)
        except ValueError:
            if not allow_incomplete:
                raise ValueError(f"Quantidade invÃ¡lida na linha de componente {index + 1}.")
            quantity = 1.0
        if quantity <= 0:
            raise ValueError(f"Quantidade deve ser maior que zero na linha de componente {index + 1}.")

        components.append(
            {
                "codigo": code,
                "descricao": description or search_text,
                "unidade": unit or "pc",
                "quantidade": quantity,
            }
        )

    if incomplete_rows:
        rows = ", ".join(str(row) for row in incomplete_rows)
        raise ValueError(f"Selecione um produto vÃ¡lido nas linhas de componente: {rows}.")
    return components


def _safe_filename(value: str) -> str:
    text = strip_accents(clean_text(value)).upper()
    text = re.sub(r"[^A-Z0-9._ -]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return (text or "BOM")[:140]


def _product_catalog_from_workbook(max_rows_per_sheet: int | None = None) -> list[dict[str, str]]:
    global _PRODUCT_CATALOG_CACHE
    workbook = ensure_workbook_exists()
    workbook_mtime = workbook.stat().st_mtime
    if max_rows_per_sheet is None and _PRODUCT_CATALOG_CACHE is not None:
        cached_path, cached_mtime, cached_products = _PRODUCT_CATALOG_CACHE
        if cached_path == workbook and cached_mtime == workbook_mtime:
            return deepcopy(cached_products)

    catalog = load_catalog()
    category_by_sheet = {
        _safe_sheet_title(category.get("sheet_name") or category.get("label")): category
        for category in catalog.get("categories") or []
    }
    wb = load_workbook(workbook, data_only=True, read_only=False)
    products: list[dict[str, str]] = []
    seen_codes: set[str] = set()
    try:
        for ws in wb.worksheets:
            if ws.title.startswith("_") or ws.title not in category_by_sheet:
                continue
            header_row = _detect_header_row(ws)
            headers = {normalize_label(ws.cell(header_row, column).value): column for column in range(1, ws.max_column + 1)}
            sku_col = headers.get("SKU") or 1
            primary_col = headers.get(normalize_label(DESCRIPTION_PRIMARY_HEADER)) or headers.get("DESCRICAO PRIMARIA") or 2
            secondary_col = headers.get(normalize_label(DESCRIPTION_SECONDARY_HEADER)) or headers.get("DESCRICAO SECUNDARIA") or 3
            unit_col = (
                headers.get("UN MEDI INTERNA")
                or headers.get("UN INTERNA")
                or headers.get("UNIDADE")
                or headers.get("UN MEDIDA")
            )
            count = 0
            for row in range(header_row + 1, ws.max_row + 1):
                code = clean_text(ws.cell(row, sku_col).value)
                primary = clean_text(ws.cell(row, primary_col).value)
                secondary = clean_text(ws.cell(row, secondary_col).value)
                if not code or not primary:
                    continue
                if code in seen_codes:
                    continue
                seen_codes.add(code)
                products.append(
                    {
                        "codigo": code,
                        "descricao": primary,
                        "descricao_secundaria": secondary,
                        "unidade": clean_text(ws.cell(row, unit_col).value) if unit_col else "pc",
                        "categoria": ws.title,
                    }
                )
                count += 1
                if max_rows_per_sheet and count >= max_rows_per_sheet:
                    break
    finally:
        wb.close()
    if max_rows_per_sheet is None:
        _PRODUCT_CATALOG_CACHE = (workbook, workbook_mtime, deepcopy(products))
    return products


def search_products(query: str, limit: int = 25) -> list[dict[str, str]]:
    term = clean_text(query)
    if len(term) < 1:
        return []
    normalized_term = normalize_label(term)
    compact_term = normalized_term.replace(" ", "")
    matches: list[tuple[int, dict[str, str]]] = []
    for product in _product_catalog_from_workbook():
        code = clean_text(product.get("codigo"))
        description = clean_text(product.get("descricao"))
        haystack = normalize_label(f"{code} {description} {product.get('categoria')}")
        compact_haystack = haystack.replace(" ", "")
        if compact_term not in compact_haystack:
            continue
        score = 0
        if code.startswith(term):
            score -= 30
        if normalize_label(description).startswith(normalized_term):
            score -= 15
        score += len(description)
        matches.append((score, product))
    matches.sort(key=lambda item: (item[0], item[1]["codigo"]))
    return [product for _, product in matches[:limit]]


def generate_bom_workbook(
    item_code: str,
    item_description: str,
    components: list[dict[str, Any]],
) -> Path:
    item_code = clean_text(item_code)
    item_description = clean_text(item_description)
    if not item_code:
        raise ValueError("Informe o cÃ³digo do conjunto/PP para gerar a planilha BOM.")
    if not components:
        raise ValueError("Informe pelo menos um componente para gerar a planilha BOM.")

    BOM_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_name = f"{stamp} - BOM - {item_code} - {_safe_filename(item_description)}.xlsx"
    output_path = BOM_OUTPUT_DIR / file_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws.append(["item_codigo", "componente_codigo", "descricao", "unidade", "quantidade", None, "ITEM_BLOCO", item_description])
    for component in components:
        quantity = component["quantidade"]
        if float(quantity).is_integer():
            quantity = int(quantity)
        ws.append(
            [
                item_code,
                component["codigo"],
                component["descricao"],
                component.get("unidade") or "pc",
                quantity,
                None,
                None,
                None,
            ]
        )

    for column, width in {"A": 16, "B": 20, "C": 72, "D": 12, "E": 14, "G": 16, "H": 56}.items():
        ws.column_dimensions[column].width = width
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)
    wb.close()
    return output_path


def save_banco_registration(form_data: Any) -> dict[str, str]:
    category_key_value = clean_text(form_data.get("categoria"))
    category = selected_category(category_key_value)
    catalog = load_catalog()
    raw_category = _find_category(catalog, category["key"])
    fields = get_banco_fields(category["key"])
    if category["key"] == DEFAULT_CATEGORY_KEY:
        _validate_banco_dependencies(fields, form_data)
        _validate_visible_field_requirements(fields, category["key"], form_data)
    workbook = ensure_workbook_exists()
    workbook_source = _copy_to_temp(workbook)
    wb = _load(workbook_source)
    backup_path = None

    try:
        ws = _sheet_for_category(wb, raw_category)
        primary_column, secondary_column, suffix_column = _resolve_description_columns(ws, create_missing=True)
        # Column identity always follows the field's structural/original scope.
        # Conditional scope changes affect only descriptions for this record.
        field_columns = _resolve_field_column_map(ws, fields, create_missing=True)
        duplicate_row = _find_duplicate_registration_by_description(
            ws,
            fields,
            field_columns,
            form_data,
            (primary_column, secondary_column, suffix_column),
            category["key"],
        )
        if duplicate_row:
            raise ValueError(f"Cadastro jÃ¡ existe na linha {duplicate_row}.")

        backup_path = _backup_workbook(workbook, "cadastro")
        row = _next_available_row(ws)
        if row > FIRST_DATA_ROW:
            _copy_row_style(ws, row - 1, row)
            _copy_row_formulas(ws, row - 1, row)
        _expand_table_to_row(ws, row)

        descriptions = build_descriptions(fields, form_data, category["key"])
        sku_column = _resolve_header_column(ws, "SKU", create_missing=True)
        sku_value = _next_sequential_sku(ws, sku_column, raw_category, fields, form_data)
        if sku_column:
            ws.cell(row, sku_column).value = sku_value
        if primary_column:
            ws.cell(row, primary_column).value = descriptions["primaria"]
        if secondary_column:
            ws.cell(row, secondary_column).value = descriptions["secundaria"]
        if suffix_column:
            ws.cell(row, suffix_column).value = descriptions["sufixo"] or None

        for field in fields:
            values = _serialize_field_values(field, form_data)
            value_column, code_column = field_columns.get(field["key"], (None, None))
            ws.cell(row, value_column).value = _format_field_saved_value(field, values) if values else None
            if code_column:
                codes = [option_code(value) for value in values if option_code(value)]
                ws.cell(row, code_column).value = " | ".join(codes) if codes else None

        control_ws = _registration_sheet(wb)
        control_ws.append(
            [
                raw_category["key"],
                raw_category["label"],
                ws.title,
                row,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ]
        )

        _save_workbook_preserving_package(wb, workbook)
        _REGISTRATION_CACHE.clear()
        global _PRODUCT_CATALOG_CACHE
        _PRODUCT_CATALOG_CACHE = None
        set_active_category(category["key"])
        return {
            "row": row,
            "category": raw_category["label"],
            "category_key": raw_category["key"],
            "sheet": ws.title,
            "descricao_primaria": descriptions["primaria"],
            "descricao_secundaria": descriptions["secundaria"],
            "sku": sku_value,
            "path": str(workbook),
            "backup": str(backup_path) if backup_path else "",
        }
    except PermissionError as exc:
        raise PermissionError(
            "NÃ£o consegui salvar a planilha. Feche o Excel se ela estiver aberta e tente novamente."
        ) from exc
    finally:
        wb.close()
        cleanup_path = getattr(wb, "_cadastro_cleanup_path", None)
        if cleanup_path is not None:
            try:
                Path(cleanup_path).unlink(missing_ok=True)
            except OSError:
                pass
        try:
            workbook_source.unlink(missing_ok=True)
        except OSError:
            pass


def _find_field(catalog: dict[str, Any], category_key_value: str, field_key_value: str) -> dict[str, Any]:
    category = _find_category(catalog, category_key_value)
    for field in category.get("fields") or []:
        if field["key"] == field_key_value:
            return field
    raise ValueError("Campo nÃ£o encontrado.")


def _next_option_code(options: list[str]) -> int:
    max_code = 0
    for option in options:
        match = VALUE_RE.match(clean_text(option))
        if match:
            max_code = max(max_code, int(match.group(1)))
    return max_code + 1


def _format_option_value(value: str, options: list[str], existing_value: str = "") -> str:
    text = clean_text(value)
    if not text:
        return ""
    if VALUE_RE.match(text):
        return normalize_option_text(text)
    code = _first_code(existing_value, _next_option_code(options))
    return normalize_option_text(f"{code}- {text}")


def add_field_option(category_key_value: str, field_key_value: str, option_value: str) -> dict[str, str]:
    catalog = load_catalog()
    field = _find_field(catalog, category_key_value, field_key_value)
    raw_option = clean_text(option_value)
    if not raw_option:
        raise ValueError("Informe a nova opÃ§Ã£o.")

    options = field.setdefault("options", [])
    if option_identity(raw_option) in {option_identity(value) for value in options}:
        raise ValueError("Essa opÃ§Ã£o jÃ¡ existe para o campo selecionado.")

    option = _format_option_value(raw_option, options)
    options.append(option)
    save_catalog(catalog)
    return {
        "field": field["label"],
        "option": option,
        "row": len(options),
        "path": str(DATA_PATH),
        "backup": "",
    }


def update_field_option(category_key_value: str, field_key_value: str, row_value: int, option_value: str) -> dict[str, str]:
    catalog = load_catalog()
    field = _find_field(catalog, category_key_value, field_key_value)
    options = field.setdefault("options", [])
    index = int(row_value) - 1
    if index < 0 or index >= len(options):
        raise ValueError("OpÃ§Ã£o nÃ£o encontrada.")

    current = options[index]
    option = _format_option_value(option_value, options, current)
    other_options = [value for pos, value in enumerate(options) if pos != index]
    if option_identity(option) in {option_identity(value) for value in other_options}:
        raise ValueError("Essa opÃ§Ã£o jÃ¡ existe para o campo selecionado.")

    options[index] = option
    save_catalog(catalog)
    return {
        "field": field["label"],
        "option": option,
        "row": row_value,
        "path": str(DATA_PATH),
        "backup": "",
    }


def add_conditional_rule(
    category_key_value: str,
    source_field_key_value: str,
    source_value_value: str,
    target_field_key_value: str,
    target_field_label_value: str = "",
    action_value: str = "hide",
) -> dict[str, Any]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    fields = category.get("fields") or []
    source_field = _find_field(catalog, category["key"], source_field_key_value)
    target_field = None
    target_key = clean_text(target_field_key_value)
    target_label = clean_text(target_field_label_value)
    if target_key:
        target_field = _find_field(catalog, category["key"], target_key)
    elif not target_label:
        raise ValueError("Informe o campo alvo da regra.")
    source_value = rule_option_token(clean_text(source_value_value))
    if not source_value:
        raise ValueError("Informe a opÃ§Ã£o condicional.")
    action = clean_text(action_value).lower()
    if action not in {"hide", "show", "set_primary", "set_secondary"}:
        action = "hide"
    if action in {"set_primary", "set_secondary"} and target_field is None:
        raise ValueError("Selecione um campo alvo existente para alterar a descri\u00e7\u00e3o.")

    rules = category.setdefault("conditional_rules", [])
    for rule in rules:
        same_trigger_and_target = (
            rule.get("source_field_key") == source_field["key"]
            and rule.get("target_field_key") == (target_field["key"] if target_field else "")
            and rule_option_token(clean_text(rule.get("target_field_label"))) == rule_option_token(
                target_label or (target_field["label"] if target_field else "")
            )
            and rule_option_token(rule.get("source_values", [""])[0] if rule.get("source_values") else "") == source_value
        )
        if not same_trigger_and_target:
            continue
        if rule.get("action") == action:
            raise ValueError("Essa regra jÃ¡ existe.")
        if action in {"set_primary", "set_secondary"} and rule.get("action") in {"set_primary", "set_secondary"}:
            raise ValueError("JÃ¡ existe uma regra de classificaÃ§Ã£o para esse campo, opÃ§Ã£o e alvo.")

    rule = {
        "key": uuid.uuid4().hex[:12],
        "source_field_key": source_field["key"],
        "source_field_label": source_field["label"],
        "source_field_scope": source_field["scope"],
        "source_values": [source_value],
        "target_field_key": target_field["key"] if target_field else "",
        "target_field_label": target_field["label"] if target_field else target_label,
        "target_field_scope": target_field["scope"] if target_field else "secundaria",
        "action": action,
    }
    rules.append(rule)
    save_catalog(catalog)
    return {
        "rule": rule,
        "source_field": source_field["label"],
        "target_field": target_field["label"] if target_field else target_label,
        "path": str(DATA_PATH),
        "backup": "",
    }


def delete_conditional_rule(category_key_value: str, rule_key_value: str) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    rules = category.setdefault("conditional_rules", [])
    rule_index = next((index for index, rule in enumerate(rules) if rule.get("key") == rule_key_value), None)
    if rule_index is None:
        raise ValueError("Regra nÃ£o encontrada.")
    rule = rules.pop(rule_index)
    save_catalog(catalog)
    return {
        "rule": rule.get("key", ""),
        "path": str(DATA_PATH),
        "backup": "",
    }


def update_field_options(
    category_key_value: str,
    field_key_value: str,
    row_values: list[int],
    option_values: list[str],
) -> dict[str, Any]:
    catalog = load_catalog()
    field = _find_field(catalog, category_key_value, field_key_value)
    options = field.setdefault("options", [])
    if len(row_values) != len(option_values):
        raise ValueError("Quantidade de alteraÃ§Ãµes invÃ¡lida.")

    updates: list[tuple[int, str, int]] = []
    for row_value, raw_option in zip(row_values, option_values):
        index = int(row_value) - 1
        if index < 0 or index >= len(options):
            raise ValueError("OpÃ§Ã£o nÃ£o encontrada.")
        raw_text = clean_text(raw_option)
        if not raw_text:
            continue
        current = options[index]
        option = _format_option_value(raw_text, options, current)
        updates.append((index, option, row_value))

    if not updates:
        return {
            "field": field["label"],
            "updated": [],
            "count": 0,
            "path": str(DATA_PATH),
            "backup": "",
        }

    new_options = list(options)
    for index, option, _ in updates:
        new_options[index] = option

    seen_identities: set[str] = set()
    for option in new_options:
        identity = option_identity(option)
        if identity in seen_identities:
            raise ValueError("Essa opÃ§Ã£o jÃ¡ existe para o campo selecionado.")
        seen_identities.add(identity)

    options[:] = new_options
    save_catalog(catalog)
    return {
        "field": field["label"],
        "updated": [option for _, option, _ in updates],
        "count": len(updates),
        "path": str(DATA_PATH),
        "backup": "",
    }


def delete_field_option(category_key_value: str, field_key_value: str, row_value: int) -> dict[str, str]:
    catalog = load_catalog()
    field = _find_field(catalog, category_key_value, field_key_value)
    options = field.setdefault("options", [])
    index = int(row_value) - 1
    if index < 0 or index >= len(options):
        raise ValueError("OpÃ§Ã£o nÃ£o encontrada.")
    option = options.pop(index)
    save_catalog(catalog)
    return {
        "field": field["label"],
        "option": option,
        "row": row_value,
        "path": str(DATA_PATH),
        "backup": "",
    }


def add_field(category_key_value: str, label: str, scope: str, selection_mode: str) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    label_clean = clean_text(label).upper()
    if not label_clean:
        raise ValueError("Informe o nome do campo.")
    if normalize_label(label_clean) in {normalize_label(field["label"]) for field in category["fields"]}:
        raise ValueError("Esse campo jÃ¡ existe nesta categoria.")

    used = {field["key"] for field in category["fields"]}
    scope_value = field_scope(scope)
    next_order = max(
        (_field_description_order(field, 0) for field in category["fields"] if field.get("scope") == scope_value),
        default=0,
    ) + 1
    field = {
        "key": _unique_key(label_clean, used),
        "label": label_clean,
        "scope": scope_value,
        "selection_mode": field_selection_mode(selection_mode),
        "description_order": next_order,
        "options": [],
    }
    category["fields"].append(field)
    _normalize_field_orders(category)
    save_catalog(catalog)
    sync_workbook_structure(category["key"])
    return {
        "field": field["label"],
        "scope": field["scope"],
        "selection_mode": field["selection_mode"],
        "path": str(DATA_PATH),
        "backup": "",
    }


def update_field(
    category_key_value: str,
    field_key_value: str,
    label: str,
    scope: str,
    selection_mode: str,
) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    field = _find_field(catalog, category["key"], field_key_value)
    label_clean = clean_text(label).upper()
    if not label_clean:
        raise ValueError("Informe o nome do campo.")

    for existing in category["fields"]:
        if existing["key"] != field_key_value and normalize_label(existing["label"]) == normalize_label(label_clean):
            raise ValueError("JÃ¡ existe outro campo com esse nome nesta categoria.")

    previous_scope = field.get("scope")
    new_scope = field_scope(scope)
    field["label"] = label_clean
    field["scope"] = new_scope
    field["selection_mode"] = field_selection_mode(selection_mode)
    if previous_scope != new_scope:
        field["description_order"] = max(
            (_field_description_order(existing_field, 0) for existing_field in category["fields"] if existing_field.get("scope") == new_scope),
            default=0,
        ) + 1
    _normalize_field_orders(category)
    save_catalog(catalog)
    sync_workbook_structure(category["key"])
    return {
        "field": field["label"],
        "scope": field["scope"],
        "selection_mode": field["selection_mode"],
        "path": str(DATA_PATH),
        "backup": "",
    }


def delete_field(category_key_value: str, field_key_value: str) -> dict[str, str]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    field = _find_field(catalog, category["key"], field_key_value)
    category["fields"] = [item for item in category["fields"] if item["key"] != field_key_value]
    _normalize_field_orders(category)
    save_catalog(catalog)
    return {"field": field["label"], "path": str(DATA_PATH), "backup": ""}


def reorder_fields_by_description(
    category_key_value: str,
    scope_value: str,
    ordered_field_keys: list[str],
) -> dict[str, Any]:
    catalog = load_catalog()
    category = _find_category(catalog, category_key_value)
    scope_clean = field_scope(scope_value)
    fields = category.get("fields") or []
    scoped_fields = [field for field in fields if field.get("scope") == scope_clean]
    scoped_keys = {field["key"] for field in scoped_fields}
    ordered_keys = [clean_text(key) for key in ordered_field_keys if clean_text(key)]
    if set(ordered_keys) != scoped_keys:
        raise ValueError("A ordenaÃ§Ã£o recebida nÃ£o corresponde aos campos da descriÃ§Ã£o selecionada.")

    by_key = {field["key"]: field for field in scoped_fields}
    for index, key in enumerate(ordered_keys, start=1):
        by_key[key]["description_order"] = index

    _normalize_field_orders(category)
    save_catalog(catalog)
    return {
        "category": category["label"],
        "scope": scope_clean,
        "count": len(ordered_keys),
        "path": str(DATA_PATH),
        "backup": "",
    }

"""Importacao controlada dos cadastros ativos de 18 - REVESTIMENTO.

A planilha operacional nao altera a identidade do cadastro. Ela apenas informa
os campos tecnicos que compoem as descricoes. Todo o arquivo e validado antes
de qualquer gravacao para que SKU, unidade, status e B.O.M. permaneçam
intactos.
"""

from __future__ import annotations

import re
from copy import deepcopy
from io import BytesIO
from typing import Any, Callable

from openpyxl import load_workbook

import excel_bancos


CATEGORY_KEY = "cat_18_revestimento"
CATEGORY_LABEL = "18 - REVESTIMENTO"
MAX_FILE_SIZE = 8 * 1024 * 1024

WORKBOOK_COLUMNS = {
    "ESTAGIO": "prefixo",
    "IDENTIFICACAO": "descritor_base",
    "NIVEL": "veiculo_modelo",
    "LOCAL": "posicao_lado",
    "LADO": "medida",
    "VEICULO": "material_cor",
    "FORNECEDOR": "complemento_regra",
    "TIPO": "tipo1",
    "MATERIAL": "material1",
    "ACABAMENTO": "acabamento1",
    "COR": "cor1",
    "ESPECIFICIDADE": "especificidade1",
}
REQUIRED_COLUMNS = ("COD", "CATEGORIA", "GRUPO", *WORKBOOK_COLUMNS)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _header(value: Any) -> str:
    return excel_bancos.normalize_label(_text(value)).replace(" ", "")


def _group_code(value: Any) -> str:
    match = re.match(r"\s*(\d+)", _text(value))
    return match.group(1) if match else ""


def _find_header_row(worksheet: Any) -> tuple[int, dict[str, int]]:
    for row_number in range(1, min(12, worksheet.max_row) + 1):
        headers = {
            _header(cell.value): cell.column
            for cell in worksheet[row_number]
            if _header(cell.value)
        }
        if "COD" in headers and "CATEGORIA" in headers:
            missing = [column for column in REQUIRED_COLUMNS if column not in headers]
            if missing:
                raise ValueError(
                    "A planilha de Revestimento nao possui as colunas obrigatorias: "
                    + ", ".join(missing)
                    + "."
                )
            return row_number, headers
    raise ValueError("Nao foi localizada a linha de cabecalho da planilha de Revestimento.")


def load_workbook_rows(content: bytes) -> list[dict[str, str]]:
    if not content:
        raise ValueError("O arquivo enviado esta vazio.")
    if len(content) > MAX_FILE_SIZE:
        raise ValueError("A planilha excede o limite de 8 MB.")
    try:
        workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # pragma: no cover - depends on openpyxl parser details
        raise ValueError("Nao foi possivel ler a planilha XLSX de Revestimento.") from exc
    worksheet = workbook.active
    header_row, headers = _find_header_row(worksheet)
    rows: list[dict[str, str]] = []
    seen_skus: set[str] = set()
    try:
        for row_number, source_row in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            row = {
                column: _text(source_row[headers[column] - 1])
                for column in REQUIRED_COLUMNS
            }
            if not any(row.values()):
                continue
            sku = row["COD"]
            if not sku:
                raise ValueError(f"Linha {row_number}: COD e obrigatorio.")
            if sku in seen_skus:
                raise ValueError(f"SKU duplicado na planilha: {sku}.")
            seen_skus.add(sku)
            if excel_bancos.normalize_label(row["CATEGORIA"]) != excel_bancos.normalize_label(CATEGORY_LABEL):
                raise ValueError(
                    f"SKU {sku}: categoria informada deve ser {CATEGORY_LABEL}, "
                    f"nao {row['CATEGORIA'] or '(vazia)'}.")
            group = _group_code(row["GRUPO"])
            if group not in {"10", "20", "30"}:
                raise ValueError(f"SKU {sku}: GRUPO deve ser 10, 20 ou 30.")
            row["GRUPO"] = group
            rows.append(row)
    finally:
        workbook.close()
    if not rows:
        raise ValueError("A planilha nao contem SKUs para atualizar.")
    return rows


def _matching_options(field: dict[str, Any], raw_value: str) -> list[str]:
    raw = _text(raw_value)
    if not raw:
        return []
    expected = excel_bancos.normalize_label(raw)
    matches = [
        option
        for option in field.get("options", [])
        if excel_bancos.normalize_label(excel_bancos.option_label(option)) == expected
    ]
    # O catálogo já possui valores históricos com pequenas diferenças visuais
    # (por exemplo, "ARO JAN" e "ARO JAN."). Essas diferenças não justificam
    # criar uma segunda opção nem deixar a importação ambígua.
    if not matches:
        compact = re.sub(r"[^A-Z0-9]+", "", expected)
        matches = [
            option
            for option in field.get("options", [])
            if re.sub(
                r"[^A-Z0-9]+",
                "",
                excel_bancos.normalize_label(excel_bancos.option_label(option)),
            )
            == compact
        ]
    return matches


def _canonical_option(field: dict[str, Any], raw_value: str, sku: str) -> str:
    raw = _text(raw_value)
    if not raw:
        return ""
    matches = _matching_options(field, raw)
    if len(matches) != 1:
        label = field.get("label") or field.get("key") or "campo"
        raise ValueError(
            f"SKU {sku}: valor {raw!r} nao possui correspondencia unica "
            f"no campo {label}. Corrija a planilha ou cadastre a opcao antes de importar.")
    return matches[0]


def missing_field_options(content: bytes, fields: list[dict[str, Any]]) -> dict[str, list[str]]:
    """Return only genuinely new values required by the controlled workbook.

    The workbook is validated before the category catalogue is altered. Values
    that already have an exact or punctuation-only equivalent are intentionally
    excluded, preserving one canonical option per field.
    """
    source_rows = load_workbook_rows(content)
    fields_by_key = {field.get("key"): field for field in fields}
    missing_catalog_fields = [key for key in WORKBOOK_COLUMNS.values() if key not in fields_by_key]
    if missing_catalog_fields:
        raise ValueError(
            "O catálogo de Revestimento está incompleto: " + ", ".join(missing_catalog_fields) + "."
        )

    additions: dict[str, list[str]] = {}
    for source in source_rows:
        for workbook_column, field_key in WORKBOOK_COLUMNS.items():
            raw = _text(source[workbook_column])
            if not raw:
                continue
            matches = _matching_options(fields_by_key[field_key], raw)
            if len(matches) != 1:
                # A missing option is a supported, controlled extension of the
                # category. Ambiguous historical options must be corrected
                # first; importing through them would be non-deterministic.
                if len(matches) > 1:
                    _canonical_option(fields_by_key[field_key], raw, source["COD"])
                pending = additions.setdefault(field_key, [])
                identity = excel_bancos.option_identity(raw)
                if identity not in {excel_bancos.option_identity(value) for value in pending}:
                    pending.append(raw)
    return additions


def fields_with_pending_options(
    fields: list[dict[str, Any]],
    additions: dict[str, list[str]],
) -> list[dict[str, Any]]:
    """Create an in-memory catalogue for full validation before it is saved."""
    preview = deepcopy(fields)
    by_key = {field.get("key"): field for field in preview}
    for field_key, values in additions.items():
        field = by_key[field_key]
        options = field.setdefault("options", [])
        next_code = 1
        for option in options:
            raw_code = excel_bancos.option_code(option)
            if raw_code.isdigit():
                next_code = max(next_code, int(raw_code) + 1)
        for value in values:
            options.append(excel_bancos.normalize_option_text(f"{next_code}- {value}"))
            next_code += 1
    return preview


def prepare_reconciliation(
    content: bytes,
    fields: list[dict[str, Any]],
    active_rows: list[dict[str, Any]],
    payload_builder: Callable[[dict[str, Any], dict[str, list[str]]], dict[str, Any]],
) -> dict[str, Any]:
    """Validate the whole workbook and return safe update payloads only."""
    source_rows = load_workbook_rows(content)
    fields_by_key = {field.get("key"): field for field in fields}
    missing_catalog_fields = [key for key in WORKBOOK_COLUMNS.values() if key not in fields_by_key]
    if missing_catalog_fields:
        raise ValueError(
            "O catalogo de Revestimento esta incompleto: " + ", ".join(missing_catalog_fields) + ".")

    rows_by_sku: dict[str, dict[str, Any]] = {}
    for registration in active_rows:
        sku = _text(registration.get("sku"))
        if not sku:
            continue
        if sku in rows_by_sku:
            raise ValueError(f"Base possui mais de um cadastro ativo com o SKU {sku}.")
        rows_by_sku[sku] = registration

    source_skus = {row["COD"] for row in source_rows}
    active_skus = set(rows_by_sku)
    absent_in_workbook = sorted(active_skus - source_skus)
    absent_in_database = sorted(source_skus - active_skus)
    if absent_in_workbook or absent_in_database:
        details: list[str] = []
        if absent_in_workbook:
            details.append("ativos ausentes na planilha: " + ", ".join(absent_in_workbook[:12]))
        if absent_in_database:
            details.append("SKUs da planilha inexistentes/nao ativos: " + ", ".join(absent_in_database[:12]))
        raise ValueError(
            "A planilha deve representar exatamente os cadastros ativos de Revestimento ("
            + "; ".join(details)
            + "). Nenhuma alteracao foi gravada.")

    payloads: list[dict[str, Any]] = []
    changed_skus: list[str] = []
    for source in source_rows:
        registration = rows_by_sku[source["COD"]]
        original_groups = registration.get("form_values")
        if not isinstance(original_groups, dict):
            original_groups = {}
        # GRUPO is an identity attribute, not a technical description field.
        # Keep the value already persisted in Cadastro even when an older or
        # manually edited workbook contains a different group. Deliberate
        # group migrations must use the dedicated cadastro edit workflow.
        selected = {**original_groups}
        for workbook_column, field_key in WORKBOOK_COLUMNS.items():
            canonical = _canonical_option(fields_by_key[field_key], source[workbook_column], source["COD"])
            selected[field_key] = [canonical] if canonical else []
        payload = payload_builder(registration, selected)
        payloads.append(payload)
        if any(
            payload.get(key) != registration.get(key)
            for key in ("descricao_primaria", "descricao_secundaria", "form_values", "field_values", "field_codes")
        ):
            changed_skus.append(source["COD"])
    return {
        "total": len(source_rows),
        "payloads": payloads,
        "changed": len(changed_skus),
        "changed_skus": changed_skus,
        "source_skus": sorted(source_skus),
    }

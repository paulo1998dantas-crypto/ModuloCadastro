from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="0F172A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="E2E8F0")
THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def build_import_template(
    sheet_name: str,
    fields: list[dict[str, Any]],
    warning: str = "",
) -> bytes:
    wb = Workbook()
    wb.properties.creator = "ModuloCadastro"
    wb.properties.title = f"Template de importacao - {sheet_name}"

    ws = wb.active
    ws.title = sheet_name[:31]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}1"
    ws.row_dimensions[1].height = 32

    for column, field in enumerate(fields, start=1):
        header = str(field["header"])
        cell = ws.cell(1, column, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        details = [str(field.get("description") or "").strip()]
        example = str(field.get("example") or "").strip()
        if example:
            details.append(f"Exemplo: {example}")
        details.append("Obrigatorio" if field.get("required") else "Opcional")
        cell.comment = Comment("\n".join(part for part in details if part), "ModuloCadastro")
        ws.column_dimensions[get_column_letter(column)].width = max(
            int(field.get("width") or 18),
            min(len(header) + 3, 32),
        )

    instructions = wb.create_sheet("INSTRUCOES")
    instructions.sheet_view.showGridLines = False
    instructions.freeze_panes = "A2"
    instructions.column_dimensions["A"].width = 30
    instructions.column_dimensions["B"].width = 15
    instructions.column_dimensions["C"].width = 68
    instructions.column_dimensions["D"].width = 42

    if warning:
        instructions.merge_cells("A1:D1")
        warning_cell = instructions["A1"]
        warning_cell.value = warning
        warning_cell.fill = PatternFill("solid", fgColor="FEF3C7")
        warning_cell.font = Font(color="92400E", bold=True)
        warning_cell.alignment = Alignment(vertical="center", wrap_text=True)
        instructions.row_dimensions[1].height = 48
        header_row = 3
    else:
        header_row = 1

    instruction_headers = ["CAMPO", "OBRIGATORIO", "ORIENTACAO", "EXEMPLO"]
    for column, value in enumerate(instruction_headers, start=1):
        cell = instructions.cell(header_row, column, value)
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True, color="0F172A")
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_index, field in enumerate(fields, start=header_row + 1):
        values = [
            field["header"],
            "SIM" if field.get("required") else "NAO",
            field.get("description") or "",
            field.get("example") or "",
        ]
        for column, value in enumerate(values, start=1):
            cell = instructions.cell(row_index, column, value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output = BytesIO()
    wb.save(output)
    wb.close()
    return output.getvalue()

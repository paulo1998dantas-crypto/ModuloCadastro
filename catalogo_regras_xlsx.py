from __future__ import annotations

import io
import uuid
from copy import deepcopy
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

import excel_bancos


SHEET_FIELDS = "CAMPOS_E_OPCOES"
SHEET_RULES = "REGRAS_CONDICIONAIS"

FIELD_HEADERS = [
    "ACAO_REGISTRO",
    "CATEGORIA",
    "CHAVE_CATEGORIA",
    "CAMPO",
    "CHAVE_CAMPO",
    "ESCOPO",
    "MODO_SELECAO",
    "OBRIGATORIO",
    "TIPO_ENTRADA",
    "OPCOES_DISPONIVEIS",
    "ORDEM_DESCRICAO",
]

RULE_HEADERS = [
    "ACAO_REGISTRO",
    "CHAVE_REGRA",
    "CATEGORIA",
    "CHAVE_CATEGORIA",
    "ACAO",
    "CAMPO_ORIGEM",
    "CHAVE_CAMPO_ORIGEM",
    "VALORES_GATILHO",
    "CAMPO_DESTINO",
    "CHAVE_CAMPO_DESTINO",
    "ESCOPO_DESTINO",
    "COMPARACAO",
    "ORIGEM_ATUAL",
    "SEPARADOR",
    "CAMPOS_ADICIONAIS",
    "TEXTO_LITERAL",
]

HEADER_FILL = PatternFill("solid", fgColor="0B2948")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL = PatternFill("solid", fgColor="DDF3F8")


def _text(value: Any) -> str:
    return excel_bancos.clean_text(value)


def _normalized(value: Any) -> str:
    return excel_bancos.normalize_label(_text(value))


def _header_map(ws) -> dict[str, int]:
    return {_normalized(cell.value): index for index, cell in enumerate(ws[1], start=1) if _text(cell.value)}


def _row_dict(
    ws,
    row_number: int,
    headers: list[str],
    optional_headers: set[str] | None = None,
) -> dict[str, Any]:
    columns = _header_map(ws)
    optional = {_normalized(header) for header in (optional_headers or set())}
    missing = [
        header for header in headers
        if _normalized(header) not in columns and _normalized(header) not in optional
    ]
    if missing:
        raise ValueError(f"A aba {ws.title} não contém as colunas obrigatórias: {', '.join(missing)}.")
    return {
        header: (
            ws.cell(row=row_number, column=columns[_normalized(header)]).value
            if _normalized(header) in columns
            else None
        )
        for header in headers
    }


def _split_values(value: Any) -> list[str]:
    raw = _text(value).replace("\r", "\n")
    values: list[str] = []
    seen: set[str] = set()
    for line in raw.replace("\n", ";").split(";"):
        item = excel_bancos.normalize_option_text(line)
        identity = excel_bancos.option_identity(item)
        if not item or identity in seen:
            continue
        seen.add(identity)
        values.append(item)
    return values


def _bool_value(value: Any, default: bool) -> bool:
    if value in {None, ""}:
        return default
    if isinstance(value, bool):
        return value
    normalized = _normalized(value)
    if normalized in {"SIM", "S", "TRUE", "1", "OBRIGATORIO"}:
        return True
    if normalized in {"NAO", "N", "FALSE", "0", "OPCIONAL"}:
        return False
    raise ValueError(f"Valor booleano inválido: {_text(value)}. Use SIM ou NÃO.")


def _scope(value: Any, default: str = "secundaria") -> str:
    if value in {None, ""}:
        return default
    normalized = _normalized(value)
    if normalized in {"PRIMARIA", "PRIMARIO"}:
        return "primaria"
    if normalized in {"SECUNDARIA", "SECUNDARIO"}:
        return "secundaria"
    raise ValueError(f"Escopo inválido: {_text(value)}. Use PRIMARIA ou SECUNDARIA.")


def _selection_mode(value: Any, default: str = "unitaria") -> str:
    if value in {None, ""}:
        return default
    normalized = _normalized(value)
    if normalized in {"UNITARIA", "UNITARIO", "UNICA"}:
        return excel_bancos.SELECTION_MODE_UNITARIA
    if normalized in {"MULTIPLA", "MULTIPLO"}:
        return excel_bancos.SELECTION_MODE_MULTIPLA
    raise ValueError(f"Modo de seleção inválido: {_text(value)}. Use UNITARIA ou MULTIPLA.")


def _action(value: Any) -> str:
    normalized = _normalized(value).replace(" ", "_")
    aliases = {
        "HIDE": "hide",
        "OCULTAR": "hide",
        "SHOW": "show",
        "EXIBIR": "show",
        "SET_PRIMARY": "set_primary",
        "TURN_PRIMARY": "set_primary",
        "TURN_PRIMARIO": "set_primary",
        "TURN_PRIMARIA": "set_primary",
        "TORNAR_PRIMARIO": "set_primary",
        "TORNAR_PRIMARIA": "set_primary",
        "SET_SECONDARY": "set_secondary",
        "TURN_SECONDARY": "set_secondary",
        "TURN_SECUNDARIO": "set_secondary",
        "TURN_SECUNDARIA": "set_secondary",
        "TORNAR_SECUNDARIO": "set_secondary",
        "TORNAR_SECUNDARIA": "set_secondary",
        "JOIN_FIELDS": "join_fields",
        "JUNTAR_CAMPOS": "join_fields",
        "CONCATENAR_CAMPOS": "join_fields",
        "PREPEND_LITERAL": "prepend_literal",
        "PREFIXO_LITERAL": "prepend_literal",
        "ADICIONAR_PREFIXO": "prepend_literal",
    }
    action = aliases.get(normalized)
    if action is None:
        raise ValueError(
            f"Ação inválida: {_text(value)}. Use HIDE, SHOW, TURN_PRIMARY, "
            "TURN_SECONDARY, JOIN_FIELDS ou PREPEND_LITERAL."
        )
    return action


def _find_category(catalog: dict[str, Any], key_value: Any, label_value: Any) -> dict[str, Any]:
    key = _text(key_value)
    label = _text(label_value)
    if key:
        for category in catalog.get("categories") or []:
            if category.get("key") == key:
                return category
        if not label:
            raise ValueError(f"Categoria não encontrada para a chave {key}.")
    if label:
        matches = [
            category
            for category in (catalog.get("categories") or [])
            if _normalized(category.get("label")) == _normalized(label)
        ]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise ValueError(f"A categoria {label} é ambígua.")
        used = {category.get("key") for category in catalog.get("categories") or []}
        new_key = key or excel_bancos._unique_category_key(label, used)
        category = {
            "key": new_key,
            "label": label,
            "sheet_name": excel_bancos._safe_sheet_title(label),
            "fields": [],
            "field_overrides": {},
            "conditional_rules": [],
        }
        catalog.setdefault("categories", []).append(category)
        return category
    raise ValueError("Informe CATEGORIA ou CHAVE_CATEGORIA.")


def _effective_fields(category: dict[str, Any]) -> list[dict[str, Any]]:
    return excel_bancos._fields_for_category(category)


def _find_effective_field(category: dict[str, Any], key_value: Any, label_value: Any) -> dict[str, Any] | None:
    key = _text(key_value)
    label = _text(label_value)
    fields = _effective_fields(category)
    if key:
        match = next((field for field in fields if field.get("key") == key), None)
        if match is not None:
            return match
    if label:
        matches = [field for field in fields if _normalized(field.get("label")) == _normalized(label)]
        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            raise ValueError(f"O campo {label} é ambíguo na categoria {category.get('label')}.")
    return None


def _upsert_field(category: dict[str, Any], row: dict[str, Any]) -> tuple[str, str]:
    label = _text(row["CAMPO"]).upper()
    key = _text(row["CHAVE_CAMPO"])
    if not label:
        raise ValueError("CAMPO não pode ficar vazio.")

    existing = _find_effective_field(category, key, label)
    if existing is None:
        used = {field.get("key") for field in _effective_fields(category)}
        key = key or excel_bancos._unique_key(label, used)
        existing = {
            "key": key,
            "label": label,
            "scope": _scope(row["ESCOPO"], "secundaria"),
            "selection_mode": _selection_mode(row["MODO_SELECAO"]),
            "description_order": None,
            "options": [],
            "required": True,
            "free_text": False,
        }
        category.setdefault("fields", []).append(existing)
        target = existing
        operation = "inserted"
    else:
        key = existing["key"]
        target = next(
            (field for field in category.get("fields") or [] if field.get("key") == key),
            None,
        )
        if target is None:
            target = category.setdefault("field_overrides", {}).setdefault(key, {})
        operation = "updated"

    target["label"] = label
    target["scope"] = _scope(row["ESCOPO"], existing.get("scope", "secundaria"))
    target["selection_mode"] = _selection_mode(
        row["MODO_SELECAO"], existing.get("selection_mode", "unitaria")
    )
    target["required"] = _bool_value(row["OBRIGATORIO"], bool(existing.get("required", True)))
    input_type = _normalized(row["TIPO_ENTRADA"])
    if input_type:
        if input_type not in {"LISTA", "TEXTO_LIVRE", "TEXTO LIVRE", "LIVRE"}:
            raise ValueError(f"TIPO_ENTRADA inválido para {label}. Use LISTA ou TEXTO_LIVRE.")
        target["free_text"] = input_type in {"TEXTO_LIVRE", "TEXTO LIVRE", "LIVRE"}
    if row["OPCOES_DISPONIVEIS"] not in {None, ""}:
        target["options"] = _split_values(row["OPCOES_DISPONIVEIS"])
    if row["ORDEM_DESCRICAO"] not in {None, ""}:
        try:
            target["description_order"] = max(1, int(row["ORDEM_DESCRICAO"]))
        except (TypeError, ValueError) as exc:
            raise ValueError(f"ORDEM_DESCRICAO inválida para {label}.") from exc
    return operation, key


def _resolve_rule_field(category: dict[str, Any], key_value: Any, label_value: Any, role: str) -> dict[str, Any]:
    field = _find_effective_field(category, key_value, label_value)
    if field is None:
        informed = _text(key_value) or _text(label_value)
        raise ValueError(f"Campo {role} não encontrado: {informed}.")
    return field


def _upsert_rule(category: dict[str, Any], row: dict[str, Any]) -> str:
    source_key_value = _text(row["CHAVE_CAMPO_ORIGEM"])
    source_label_value = _text(row["CAMPO_ORIGEM"])
    action = _action(row["ACAO"])
    if action == "join_fields":
        source = _resolve_rule_field(
            category, row["CHAVE_CAMPO_ORIGEM"], row["CAMPO_ORIGEM"], "de origem"
        )
        target = _resolve_rule_field(
            category, row["CHAVE_CAMPO_DESTINO"], row["CAMPO_DESTINO"], "de destino"
        )
        if source["key"] == target["key"]:
            raise ValueError("JOIN_FIELDS exige dois campos diferentes.")
        separator = _text(row.get("SEPARADOR")) or "X"
        if len(separator) > 8:
            raise ValueError("SEPARADOR deve ter no máximo 8 caracteres.")
        additional_targets: list[str] = []
        for additional_value in _split_values(row.get("CAMPOS_ADICIONAIS")):
            additional_target = _resolve_rule_field(category, additional_value, additional_value, "adicional")
            if additional_target["key"] in {source["key"], target["key"]}:
                raise ValueError("JOIN_FIELDS não pode repetir o campo de origem ou destino.")
            if additional_target["key"] not in additional_targets:
                additional_targets.append(additional_target["key"])
        rule_key = _text(row["CHAVE_REGRA"])
        rules = category.setdefault("description_rules", [])
        existing = next(
            (rule for rule in rules if rule_key and _text(rule.get("key")) == rule_key),
            None,
        )
        if existing is None:
            existing = next(
                (
                    rule for rule in rules
                    if rule.get("source_field_key") == source["key"]
                    and rule.get("target_field_key") == target["key"]
                    and _text(rule.get("action")).lower() == "join_fields"
                ),
                None,
            )
        operation = "updated" if existing is not None else "inserted"
        if existing is None:
            existing = {}
            rules.append(existing)
        existing.update(
            {
                "key": rule_key or existing.get("key") or uuid.uuid4().hex[:12],
                "action": "join_fields",
                "source_field_key": source["key"],
                "source_field_label": source["label"],
                "target_field_key": target["key"],
                "target_field_label": target["label"],
                "additional_target_field_keys": additional_targets,
                "separator": separator,
            }
        )
        return operation

    if action == "prepend_literal":
        source_type = (
            "group"
            if source_key_value == excel_bancos.PN_GROUP_FORM_KEY
            or _normalized(source_label_value) in {"GRUPO", "GRUPO DO SKU"}
            else "field"
        )
        if source_type != "group":
            raise ValueError("PREPEND_LITERAL exige GRUPO DO SKU como CAMPO_ORIGEM.")
        values = [
            excel_bancos._pn_group_code(value)
            for value in _split_values(row["VALORES_GATILHO"])
        ]
        values = [value for value in values if value]
        literal = _text(row.get("TEXTO_LITERAL"))
        if not values or not literal:
            raise ValueError("PREPEND_LITERAL exige VALORES_GATILHO e TEXTO_LITERAL.")
        rule_key = _text(row["CHAVE_REGRA"])
        rules = category.setdefault("description_rules", [])
        existing = next(
            (rule for rule in rules if rule_key and _text(rule.get("key")) == rule_key),
            None,
        )
        if existing is None:
            existing = next(
                (
                    rule for rule in rules
                    if _text(rule.get("action")).lower() == "prepend_literal"
                    and sorted(rule.get("source_values") or []) == sorted(values)
                    and _text(rule.get("literal")) == literal
                ),
                None,
            )
        operation = "updated" if existing is not None else "inserted"
        if existing is None:
            existing = {}
            rules.append(existing)
        existing.update(
            {
                "key": rule_key or existing.get("key") or uuid.uuid4().hex[:12],
                "action": "prepend_literal",
                "source_type": "group",
                "source_field_key": excel_bancos.PN_GROUP_FORM_KEY,
                "source_field_label": "GRUPO DO SKU",
                "source_values": values,
                "literal": literal,
            }
        )
        return operation

    source_type = (
        "group"
        if source_key_value == excel_bancos.PN_GROUP_FORM_KEY
        or _normalized(source_label_value) in {"GRUPO", "GRUPO DO SKU"}
        else "field"
    )
    source = (
        {
            "key": excel_bancos.PN_GROUP_FORM_KEY,
            "label": "GRUPO DO SKU",
            "scope": "estrutura",
        }
        if source_type == "group"
        else _resolve_rule_field(category, source_key_value, source_label_value, "de origem")
    )
    target_key = _text(row["CHAVE_CAMPO_DESTINO"])
    target_label = _text(row["CAMPO_DESTINO"])
    target = _find_effective_field(category, target_key, target_label)
    if action in {"set_primary", "set_secondary"} and target is None:
        raise ValueError("SET_PRIMARY e SET_SECONDARY exigem um campo destino existente.")
    if target is None and not target_label:
        raise ValueError("Informe CAMPO_DESTINO ou CHAVE_CAMPO_DESTINO.")

    values = [
        excel_bancos._pn_group_code(value)
        if source_type == "group"
        else excel_bancos.rule_option_token(value)
        for value in _split_values(row["VALORES_GATILHO"])
    ]
    values = [value for value in values if value]
    if not values:
        raise ValueError("VALORES_GATILHO não pode ficar vazio.")

    match_by = _normalized(row["COMPARACAO"])
    if match_by in {"", "OPCAO", "OPTION"}:
        match_by = "option"
    elif match_by in {"PREFIXO", "PREFIX"}:
        match_by = "prefix"
    else:
        raise ValueError("COMPARACAO inválida. Use OPCAO ou PREFIXO.")

    rule_key = _text(row["CHAVE_REGRA"])
    rules = category.setdefault("conditional_rules", [])
    existing = next((rule for rule in rules if rule_key and _text(rule.get("key")) == rule_key), None)
    if existing is None:
        target_identity = target.get("key") if target else _normalized(target_label)
        existing = next(
            (
                rule
                for rule in rules
                if (
                    _text(rule.get("source_type")).lower()
                    or ("group" if rule.get("source_field_key") == excel_bancos.PN_GROUP_FORM_KEY else "field")
                ) == source_type
                and rule.get("source_field_key") == source["key"]
                and (rule.get("target_field_key") or _normalized(rule.get("target_field_label"))) == target_identity
                and _text(rule.get("action")).lower() == action
                and sorted(rule.get("source_values") or []) == sorted(values)
            ),
            None,
        )
    operation = "updated" if existing is not None else "inserted"
    if existing is None:
        existing = {}
        rules.append(existing)
    existing.update(
        {
            "key": rule_key or existing.get("key") or uuid.uuid4().hex[:12],
            "source_type": source_type,
            "source_field_key": source["key"],
            "source_field_label": source["label"],
            "source_field_scope": source.get("scope", "primaria"),
            "source_values": values,
            "target_field_key": target["key"] if target else "",
            "target_field_label": target["label"] if target else target_label,
            "target_field_scope": (
                target.get("scope", "secundaria") if target else _scope(row["ESCOPO_DESTINO"])
            ),
            "action": action,
            "match_by": match_by,
        }
    )
    return operation


def import_catalog_workbook(content: bytes) -> dict[str, int]:
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise ValueError("O arquivo enviado não é um XLSX válido.") from exc
    try:
        if SHEET_FIELDS not in workbook.sheetnames or SHEET_RULES not in workbook.sheetnames:
            raise ValueError(f"O arquivo deve conter exatamente as abas {SHEET_FIELDS} e {SHEET_RULES}.")
        extra_sheets = [name for name in workbook.sheetnames if name not in {SHEET_FIELDS, SHEET_RULES}]
        if extra_sheets:
            raise ValueError(f"Remova abas extras antes de importar: {', '.join(extra_sheets)}.")

        fields_ws = workbook[SHEET_FIELDS]
        rules_ws = workbook[SHEET_RULES]
        catalog = deepcopy(excel_bancos.load_catalog())
        result = {
            "fields_inserted": 0,
            "fields_updated": 0,
            "rules_inserted": 0,
            "rules_updated": 0,
            "rows_ignored": 0,
        }

        for row_number in range(2, fields_ws.max_row + 1):
            row = _row_dict(fields_ws, row_number, FIELD_HEADERS)
            if not any(_text(value) for value in row.values()):
                continue
            operation = _normalized(row["ACAO_REGISTRO"])
            if operation in {"IGNORAR", "IGNORE"}:
                result["rows_ignored"] += 1
                continue
            if operation not in {"", "UPSERT", "ATUALIZAR", "INCLUIR"}:
                raise ValueError(f"Linha {row_number} de {SHEET_FIELDS}: ACAO_REGISTRO inválida.")
            try:
                category = _find_category(catalog, row["CHAVE_CATEGORIA"], row["CATEGORIA"])
                status, _ = _upsert_field(category, row)
                result[f"fields_{status}"] += 1
            except Exception as exc:
                raise ValueError(f"Linha {row_number} de {SHEET_FIELDS}: {exc}") from exc

        for row_number in range(2, rules_ws.max_row + 1):
            row = _row_dict(
                rules_ws,
                row_number,
                RULE_HEADERS,
                optional_headers={"SEPARADOR", "CAMPOS_ADICIONAIS", "TEXTO_LITERAL"},
            )
            if not any(_text(value) for value in row.values()):
                continue
            operation = _normalized(row["ACAO_REGISTRO"])
            if operation in {"IGNORAR", "IGNORE"}:
                result["rows_ignored"] += 1
                continue
            if operation not in {"", "UPSERT", "ATUALIZAR", "INCLUIR"}:
                raise ValueError(f"Linha {row_number} de {SHEET_RULES}: ACAO_REGISTRO inválida.")
            # Linhas PADRAO_SISTEMA são documentais: a aplicação as fornece
            # pelo código, não permite sobrescrever seu comportamento por XLSX
            # e mantém o relatório seguro para ser baixado e reenviado.
            if _normalized(row.get("ORIGEM_ATUAL")) == "PADRAO SISTEMA":
                result["rows_ignored"] += 1
                continue
            try:
                category = _find_category(catalog, row["CHAVE_CATEGORIA"], row["CATEGORIA"])
                status = _upsert_rule(category, row)
                result[f"rules_{status}"] += 1
            except Exception as exc:
                raise ValueError(f"Linha {row_number} de {SHEET_RULES}: {exc}") from exc

        excel_bancos.save_catalog(catalog)
        return result
    finally:
        workbook.close()


def _style_sheet(ws, widths: list[int], table_name: str) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_number in range(2, ws.max_row + 1):
        if row_number % 2 == 0:
            for cell in ws[row_number]:
                cell.fill = ALT_FILL
        for cell in ws[row_number]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    if ws.max_row >= 2:
        table = Table(displayName=table_name, ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)


def export_catalog_workbook() -> bytes:
    catalog = excel_bancos.load_catalog()
    workbook = Workbook()
    fields_ws = workbook.active
    fields_ws.title = SHEET_FIELDS
    rules_ws = workbook.create_sheet(SHEET_RULES)
    fields_ws.append(FIELD_HEADERS)
    rules_ws.append(RULE_HEADERS)

    fields_ws["A1"].comment = Comment(
        "Use UPSERT para incluir ou atualizar. Linhas ausentes não apagam dados. "
        "Separe as opções por ponto e vírgula (;).",
        "JI Montadora",
    )
    rules_ws["A1"].comment = Comment(
        "Ações aceitas: HIDE, SHOW, TURN_PRIMARY, TURN_SECONDARY e JOIN_FIELDS. "
        "Linhas PADRAO_SISTEMA são apenas consulta e são ignoradas na reimportação. "
        "Separe múltiplos valores gatilho por ponto e vírgula (;). Para JOIN_FIELDS, "
        "informe os campos de origem, destino, adicionais (se houver) e o SEPARADOR.",
        "JI Montadora",
    )

    for category in catalog.get("categories") or []:
        if category.get("key") == excel_bancos.LEGACY_CJ_BCO_CATEGORY_KEY:
            continue
        fields = _effective_fields(category)
        ordered = sorted(
            enumerate(fields, start=1),
            key=lambda item: (
                0 if item[1].get("scope") == "primaria" else 1,
                int(item[1].get("description_order") or item[0]),
                item[1].get("label") or "",
            ),
        )
        for fallback_order, field in ordered:
            fields_ws.append(
                [
                    "UPSERT",
                    category.get("label"),
                    category.get("key"),
                    field.get("label"),
                    field.get("key"),
                    (field.get("scope") or "secundaria").upper(),
                    (field.get("selection_mode") or "unitaria").upper(),
                    "SIM" if field.get("required", True) else "NAO",
                    "TEXTO_LIVRE" if field.get("free_text") else "LISTA",
                    "; ".join(field.get("options") or []),
                    int(field.get("description_order") or fallback_order),
                ]
            )

        for rule in excel_bancos.get_conditional_rules(category.get("key")):
            rules_ws.append(
                [
                    "UPSERT",
                    rule.get("key"),
                    category.get("label"),
                    category.get("key"),
                    {
                        "set_primary": "TURN_PRIMARY",
                        "set_secondary": "TURN_SECONDARY",
                        "omit_description": "OMITIR_DA_DESCRICAO",
                    }.get(rule.get("action"), (rule.get("action") or "hide").upper()),
                    rule.get("source_field_label"),
                    rule.get("source_field_key"),
                    "; ".join(rule.get("source_value_labels") or rule.get("source_values") or []),
                    rule.get("target_field_label"),
                    rule.get("target_field_key"),
                    (rule.get("target_field_scope") or "secundaria").upper(),
                    (rule.get("match_by") or "option").upper(),
                    "PADRAO_SISTEMA" if rule.get("origin") == "system" else "CATALOGO",
                    "",
                    "",
                ]
            )

        for rule in excel_bancos.get_description_rules(category.get("key")):
            if rule.get("action") == "prepend_literal":
                rules_ws.append(
                    [
                        "UPSERT",
                        rule.get("key"),
                        category.get("label"),
                        category.get("key"),
                        "PREPEND_LITERAL",
                        "GRUPO DO SKU",
                        excel_bancos.PN_GROUP_FORM_KEY,
                        "; ".join(rule.get("source_values") or []),
                        "",
                        "",
                        "",
                        "",
                        "CATALOGO",
                        "",
                        "",
                        rule.get("literal") or "",
                    ]
                )
                continue
            rules_ws.append(
                [
                    "UPSERT",
                    rule.get("key"),
                    category.get("label"),
                    category.get("key"),
                    "JOIN_FIELDS",
                    rule.get("source_field_label"),
                    rule.get("source_field_key"),
                    "",
                    rule.get("target_field_label"),
                    rule.get("target_field_key"),
                    "",
                    "",
                    "CATALOGO",
                    rule.get("separator") or "X",
                    "; ".join(rule.get("additional_target_field_keys") or []),
                    "",
                ]
            )

    _style_sheet(fields_ws, [17, 28, 23, 28, 25, 15, 18, 16, 18, 70, 18], "CatalogoCamposOpcoes")
    _style_sheet(
        rules_ws,
        [17, 20, 28, 23, 18, 28, 25, 55, 28, 25, 18, 16, 18, 14, 35, 26],
        "CatalogoRegrasCondicionais",
    )

    validations = [
        (fields_ws, "A", '"UPSERT,IGNORAR"'),
        (fields_ws, "F", '"PRIMARIA,SECUNDARIA"'),
        (fields_ws, "G", '"UNITARIA,MULTIPLA"'),
        (fields_ws, "H", '"SIM,NAO"'),
        (fields_ws, "I", '"LISTA,TEXTO_LIVRE"'),
        (rules_ws, "A", '"UPSERT,IGNORAR"'),
        (rules_ws, "E", '"HIDE,SHOW,TURN_PRIMARY,TURN_SECONDARY,JOIN_FIELDS,PREPEND_LITERAL"'),
        (rules_ws, "K", '"PRIMARIA,SECUNDARIA"'),
        (rules_ws, "L", '"OPCAO,PREFIXO"'),
    ]
    for ws, column, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(validation)
        validation.add(f"{column}2:{column}1048576")

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()

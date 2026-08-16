"""Importador idempotente dos tempos padrão do arquivo leadtime.xlsx.

O modo padrão é dry-run. ``--apply`` grava apenas SKUs ainda não
parametrizados; ``--overwrite-existing`` precisa ser informado explicitamente
para substituir uma parametrização existente.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

import supabase_store


SHEET_NAME = "Planilha1"
HEADER_ROW = 6
DATA_START_ROW = 7


@dataclass(frozen=True)
class LeadTimeRow:
    row_number: int
    sku: str
    group: str
    description: str
    origin: str
    values: dict[str, str]
    source_hash: str


def _sku(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _number(value: Any) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal("0")
    raw = str(value).strip().replace(" ", "")
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    else:
        raw = raw.replace(",", ".")
    try:
        number = Decimal(raw)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"valor numérico inválido: {value}") from exc
    if not number.is_finite() or number < 0:
        raise ValueError(f"valor negativo ou inválido: {value}")
    return number


def _decimal_text(value: Decimal) -> str:
    return format(value.quantize(Decimal("0.001")), "f")


def parse_workbook(path: str | Path) -> tuple[list[LeadTimeRow], list[dict[str, Any]]]:
    workbook_path = Path(path).expanduser().resolve()
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    if SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"A aba obrigatória {SHEET_NAME!r} não foi encontrada.")
    sheet = workbook[SHEET_NAME]
    rows: list[LeadTimeRow] = []
    rejected: list[dict[str, Any]] = []
    seen: set[str] = set()

    source_rows = sheet.iter_rows(
        min_row=DATA_START_ROW,
        max_row=sheet.max_row,
        min_col=2,
        max_col=13,
        values_only=True,
    )
    for row_number, source_row in enumerate(source_rows, start=DATA_START_ROW):
        sku = _sku(source_row[0])
        if not sku:
            continue
        try:
            if sku in seen:
                raise ValueError("PN duplicado no arquivo")
            seen.add(sku)
            group = str(source_row[1] or "").strip()
            description = str(source_row[2] or "").strip()
            external = [_number(value) for value in source_row[3:10]]
            internal = [_number(value) for value in source_row[10:12]]
            has_external = any(value > 0 for value in external)
            has_internal = any(value > 0 for value in internal)
            if has_external and has_internal:
                raise ValueError("linha possui simultaneamente tempos externos e internos")
            if not has_external and not has_internal:
                rejected.append(
                    {
                        "row": row_number,
                        "sku": sku,
                        "status": "SEM_PARAMETRO",
                        "reason": "Nenhum tempo padrão informado; origem não foi inferida.",
                    }
                )
                continue

            if has_external:
                origin = "EXTERNA"
                values = {
                    "fornecimento_dias": _decimal_text(external[0]),
                    "transporte_dias": _decimal_text(external[1]),
                    "recebimento_dias": _decimal_text(external[2]),
                    "inspecao_recebimento_dias": _decimal_text(external[3]),
                    "estocagem_dias": _decimal_text(external[4]),
                    "expedicao_dias": _decimal_text(external[5]),
                    "montagem_kit_dias": _decimal_text(external[6]),
                }
            else:
                origin = "INTERNA"
                values = {
                    "setup_dias": "0.000",
                    "producao_dias": _decimal_text(internal[0]),
                    "liberacao_dias": _decimal_text(internal[1]),
                }
            canonical = json.dumps(
                {"sku": sku, "origin": origin, "values": values},
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            rows.append(
                LeadTimeRow(
                    row_number=row_number,
                    sku=sku,
                    group=group,
                    description=description,
                    origin=origin,
                    values=values,
                    source_hash=hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
                )
            )
        except Exception as exc:
            rejected.append(
                {
                    "row": row_number,
                    "sku": sku,
                    "status": "REJEITADO",
                    "reason": str(exc),
                }
            )
    workbook.close()
    return rows, rejected


def apply_rows(
    rows: list[LeadTimeRow],
    workbook_path: Path,
    *,
    overwrite_existing: bool = False,
) -> dict[str, Any]:
    os.environ.setdefault("CADASTRO_SAVE_MODE", "supabase")
    registrations = supabase_store.list_registrations(
        supabase_store.ALL_CATEGORIES_KEY,
        include_inactive=False,
        limit=10000,
    )
    by_sku: dict[str, list[dict[str, Any]]] = {}
    for registration in registrations:
        by_sku.setdefault(str(registration.get("sku") or "").strip().upper(), []).append(registration)
    existing = supabase_store.item_parameter_summary()
    report = {"inserted": 0, "updated": 0, "unchanged": 0, "skipped": []}

    for row in rows:
        matches = by_sku.get(row.sku.upper(), [])
        if len(matches) != 1:
            report["skipped"].append(
                {
                    "row": row.row_number,
                    "sku": row.sku,
                    "reason": "SKU ativo não encontrado" if not matches else "SKU ativo ambíguo",
                }
            )
            continue
        registration = matches[0]
        current = existing.get(str(registration["id"]))
        if current and str(current.get("source_hash") or "") == row.source_hash:
            report["unchanged"] += 1
            continue
        if current and not overwrite_existing:
            report["skipped"].append(
                {
                    "row": row.row_number,
                    "sku": row.sku,
                    "reason": "Já parametrizado; use --overwrite-existing para substituir",
                }
            )
            continue
        supabase_store.save_item_parameter(
            registration,
            {"origem_fabricacao": row.origin, **row.values},
            "sistema:importador-leadtime",
            source_type="PLANILHA_LEADTIME",
            source_file=workbook_path.name,
            source_sheet=SHEET_NAME,
            source_row=row.row_number,
            source_hash=row.source_hash,
        )
        report["updated" if current else "inserted"] += 1
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Importa tempos padrão por PN.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--apply", action="store_true", help="Grava no Supabase; sem a flag faz dry-run.")
    parser.add_argument("--overwrite-existing", action="store_true")
    parser.add_argument("--report", type=Path)
    args = parser.parse_args()

    workbook_path = args.workbook.expanduser().resolve()
    rows, rejected = parse_workbook(workbook_path)
    result: dict[str, Any] = {
        "mode": "apply" if args.apply else "dry-run",
        "source": str(workbook_path),
        "sheet": SHEET_NAME,
        "parsed": len(rows),
        "external": sum(1 for row in rows if row.origin == "EXTERNA"),
        "internal": sum(1 for row in rows if row.origin == "INTERNA"),
        "not_parameterized_or_rejected": len(rejected),
        "issues": rejected,
    }
    if args.apply:
        result["database"] = apply_rows(
            rows,
            workbook_path,
            overwrite_existing=args.overwrite_existing,
        )
    if args.report:
        args.report.parent.mkdir(parents=True, exist_ok=True)
        args.report.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

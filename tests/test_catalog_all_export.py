from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import supabase_store


class CatalogAllExportTests(unittest.TestCase):
    def test_average_costs_uses_latest_confirmed_receipts_per_sku(self):
        def request_all(table, _params, limit=10000):
            if table == "erp_goods_receipts":
                return [
                    {"id": "r-old", "data_recebimento": "2026-01-02", "created_at": "2026-01-02T10:00:00Z"},
                    {"id": "r-new", "data_recebimento": "2026-02-02", "created_at": "2026-02-02T10:00:00Z"},
                ]
            self.assertEqual("erp_goods_receipt_lines", table)
            return [
                {"id": "l-old", "goods_receipt_id": "r-old", "sku_codigo": "101", "valor_unitario_real": "10", "quantidade_aprovada": "2"},
                {"id": "l-new", "goods_receipt_id": "r-new", "sku_codigo": "101", "valor_unitario_real": "16", "quantidade_aprovada": "3"},
                {"id": "l-other", "goods_receipt_id": "r-new", "sku_codigo": "999", "valor_unitario_real": "99", "quantidade_aprovada": "1"},
                {"id": "l-ignored", "goods_receipt_id": "not-confirmed", "sku_codigo": "101", "valor_unitario_real": "1", "quantidade_aprovada": "100"},
            ]

        with patch.object(supabase_store, "_request_all", side_effect=request_all):
            result = supabase_store.item_average_costs_for_export(["101"], limit_per_sku=10)

        self.assertEqual(13.6, result["101"]["preco_medio"])
        self.assertEqual(2, result["101"]["entradas_consideradas"])
        self.assertEqual("2026-02-02", result["101"]["ultima_entrada"])

    def test_export_all_categories_writes_parameters_and_one_column_per_technical_key(self):
        registrations = [
            {
                "id": 1,
                "category_label": "20 - BANCOS",
                "sku": "10200001",
                "grupo_codigo": "10",
                "grupo_label": "INSUMO",
                "descricao_primaria": "BANCO TESTE",
                "descricao_secundaria": "",
                "unidade": "pc",
                "ativo": True,
                "possui_bom": False,
                "sufixo": "BCO",
                "caracteres_primario": 10,
                "caracteres_secundario": 0,
                "field_values": {"fornecedor": ["1- MC", "2- CS"], "largura": "410"},
            }
        ]
        parameters = {
            "1": {
                "origem_fabricacao": "EXTERNA",
                "unidade_tempo": "DIA_UTIL",
                "fornecimento_dias": "5.000",
                "preco_compra": "99.9000",
            }
        }
        averages = {"10200001": {"preco_medio": 87.5, "entradas_consideradas": 2, "ultima_entrada": "2026-09-02"}}
        technical_fields = [{"key": "fornecedor", "label": "FORNECEDOR"}, {"key": "largura", "label": "LARGURA"}]

        with TemporaryDirectory() as temp_dir, \
            patch.object(supabase_store, "EXPORT_DIR", Path(temp_dir)), \
            patch.object(supabase_store, "list_registrations", return_value=registrations), \
            patch.object(supabase_store, "item_parameters_for_export", return_value=parameters), \
            patch.object(supabase_store, "item_average_costs_for_export", return_value=averages), \
            patch.object(supabase_store, "all_export_technical_fields", return_value=technical_fields):
            output = supabase_store.export_registrations("")
            workbook = load_workbook(output, data_only=True)
            sheet = workbook.active
            headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
            row = [sheet.cell(2, col).value for col in range(1, sheet.max_column + 1)]
            workbook.close()

        self.assertIn("FABRICAÇÃO", headers)
        self.assertIn("PREÇO MÉDIO ATUALIZADO (ÚLTIMAS 10 ENTRADAS)", headers)
        self.assertIn("CAMPO TÉCNICO — FORNECEDOR", headers)
        self.assertIn("CAMPO TÉCNICO — LARGURA", headers)
        self.assertNotIn("CAMPOS", headers)
        self.assertEqual("EXTERNA", row[headers.index("FABRICAÇÃO")])
        self.assertEqual(87.5, row[headers.index("PREÇO MÉDIO ATUALIZADO (ÚLTIMAS 10 ENTRADAS)")])
        self.assertEqual("1- MC | 2- CS", row[headers.index("CAMPO TÉCNICO — FORNECEDOR")])


if __name__ == "__main__":
    unittest.main()

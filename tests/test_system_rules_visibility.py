import io
import unittest
from copy import deepcopy
from unittest.mock import patch

from openpyxl import load_workbook

import catalogo_regras_xlsx
import excel_bancos


def _catalog():
    return {
        "version": 2,
        "active_category": excel_bancos.REVESTIMENTO_CATEGORY_KEY,
        "pn_groups": [],
        "categories": [
            {
                "key": excel_bancos.REVESTIMENTO_CATEGORY_KEY,
                "label": "18 - REVESTIMENTO",
                "sheet_name": "18 - REVESTIMENTO",
                "fields": [
                    {
                        "key": "fornecedor",
                        "label": "FORNECEDOR",
                        "scope": "primaria",
                        "selection_mode": "unitaria",
                        "description_order": 1,
                        "required": True,
                        "free_text": False,
                        "options": ["1- N/A", "2- FORNECEDOR A"],
                    }
                ],
                "conditional_rules": [],
            }
        ],
    }


class SystemRulesVisibilityTests(unittest.TestCase):
    def test_regra_global_de_na_e_visivel_como_padrao_do_sistema(self):
        catalog = _catalog()
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)):
            rules = excel_bancos.get_conditional_rules(excel_bancos.REVESTIMENTO_CATEGORY_KEY)

        rule = next(rule for rule in rules if rule["key"] == "cond_sistema_na_omite_fornecedor")
        self.assertEqual(rule["origin"], "system")
        self.assertEqual(rule["action"], "omit_description")
        self.assertEqual(rule["source_value_labels"], ["1- N/A"])
        self.assertEqual(rule["source_field_key"], "fornecedor")
        self.assertEqual(rule["target_field_key"], "fornecedor")

    def test_regra_global_de_na_nao_remove_o_campo_do_formulario(self):
        catalog = _catalog()
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)):
            form_rules = excel_bancos.get_conditional_rules_for_form(
                excel_bancos.REVESTIMENTO_CATEGORY_KEY
            )

        self.assertFalse(
            any(rule["key"] == "cond_sistema_na_omite_fornecedor" for rule in form_rules)
        )

    def test_bancos_exibe_perfis_internos_como_regras_de_sistema(self):
        catalog = excel_bancos._default_catalog()
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)):
            rules = excel_bancos.get_conditional_rules(excel_bancos.DEFAULT_CATEGORY_KEY)

        group_profile = next(
            rule
            for rule in rules
            if rule["key"] == "cond_bancos_grupo_conjunto_perfil_campos"
        )
        prefix_profile = next(
            rule
            for rule in rules
            if rule["key"] == "cond_bancos_prefixo_cj_perfil_conjunto"
        )
        self.assertEqual(group_profile["origin"], "system")
        self.assertTrue(group_profile["documentation_only"])
        self.assertIn("grupo 30", group_profile["description"].lower())
        self.assertEqual(prefix_profile["origin"], "system")
        self.assertTrue(prefix_profile["documentation_only"])

    def test_perfis_internos_nao_sao_enviados_para_o_javascript_do_formulario(self):
        catalog = excel_bancos._default_catalog()
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)):
            form_rules = excel_bancos.get_conditional_rules_for_form(excel_bancos.DEFAULT_CATEGORY_KEY)

        self.assertFalse(any(rule["action"] == "system_profile" for rule in form_rules))

    def test_exporta_regra_padrao_e_a_reimportacao_a_mantem_somente_para_consulta(self):
        catalog = _catalog()
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)):
            content = catalogo_regras_xlsx.export_catalog_workbook()

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        try:
            rules_ws = workbook[catalogo_regras_xlsx.SHEET_RULES]
            headers = [cell.value for cell in rules_ws[1]]
            key_column = headers.index("CHAVE_REGRA") + 1
            action_column = headers.index("ACAO") + 1
            origin_column = headers.index("ORIGEM_ATUAL") + 1
            row_number = next(
                row
                for row in range(2, rules_ws.max_row + 1)
                if rules_ws.cell(row, key_column).value == "cond_sistema_na_omite_fornecedor"
            )
            self.assertEqual(rules_ws.cell(row_number, action_column).value, "OMITIR_DA_DESCRICAO")
            self.assertEqual(rules_ws.cell(row_number, origin_column).value, "PADRAO_SISTEMA")
        finally:
            workbook.close()

        saved = []
        with (
            patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)),
            patch.object(excel_bancos, "save_catalog", side_effect=lambda value: saved.append(deepcopy(value))),
        ):
            result = catalogo_regras_xlsx.import_catalog_workbook(content)

        self.assertGreaterEqual(result["rows_ignored"], 1)
        self.assertEqual(saved[0]["categories"][0]["conditional_rules"], [])


if __name__ == "__main__":
    unittest.main()

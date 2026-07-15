import unittest
from unittest.mock import patch

from openpyxl import Workbook

import excel_bancos


class ConditionalScopeTests(unittest.TestCase):
    def setUp(self):
        self.fields = [
            {
                "key": "origem",
                "label": "ORIGEM",
                "scope": "primaria",
                "selection_mode": "unitaria",
                "description_order": 1,
                "options": ["1- ATIVA"],
            },
            {
                "key": "alvo_primario",
                "label": "ALVO PRIMARIO",
                "scope": "primaria",
                "selection_mode": "unitaria",
                "description_order": 2,
                "options": ["7- VALOR P"],
            },
            {
                "key": "alvo_secundario",
                "label": "ALVO SECUNDARIO",
                "scope": "secundaria",
                "selection_mode": "unitaria",
                "description_order": 1,
                "options": ["8- VALOR S"],
            },
        ]
        self.data = {
            "origem": "1- ATIVA",
            "alvo_primario": "7- VALOR P",
            "alvo_secundario": "8- VALOR S",
        }

    def _rule(self, action, target_key):
        target = next(field for field in self.fields if field["key"] == target_key)
        return {
            "source_field_key": "origem",
            "source_values": ["ATIVA"],
            "target_field_key": target_key,
            "target_field_label": target["label"],
            "target_field_scope": target["scope"],
            "action": action,
            "match_by": "option",
        }

    def test_primary_field_can_become_secondary_for_description_and_suffix(self):
        rules = [self._rule("set_secondary", "alvo_primario")]
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=rules):
            descriptions = excel_bancos.build_descriptions(self.fields, self.data, "teste")

        self.assertEqual(descriptions["primaria"], "ATIVA")
        self.assertEqual(descriptions["secundaria"], "ATIVA VALOR P VALOR S")
        self.assertEqual(descriptions["sufixo"], "7.8")

    def test_secondary_field_can_become_primary(self):
        rules = [self._rule("set_primary", "alvo_secundario")]
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=rules):
            descriptions = excel_bancos.build_descriptions(self.fields, self.data, "teste")

        self.assertEqual(descriptions["primaria"], "ATIVA VALOR P VALOR S")
        self.assertEqual(descriptions["secundaria"], "ATIVA VALOR P VALOR S")
        self.assertEqual(descriptions["sufixo"], "")

    def test_scope_rule_does_not_hide_target(self):
        rules = [self._rule("set_secondary", "alvo_primario")]
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=rules):
            visible = excel_bancos._visible_field_keys(self.fields, "teste", self.data)

        self.assertEqual(visible, {field["key"] for field in self.fields})

    def test_multiple_show_rules_for_same_target_are_alternatives(self):
        first_rule = self._rule("show", "alvo_secundario")
        second_rule = {
            **self._rule("show", "alvo_secundario"),
            "source_values": ["OUTRA OPCAO"],
        }
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=[first_rule, second_rule]):
            visible = excel_bancos._visible_field_keys(self.fields, "teste", self.data)

        self.assertIn("alvo_secundario", visible)

    def test_hide_rule_overrides_matching_show_rule(self):
        rules = [self._rule("show", "alvo_secundario"), self._rule("hide", "alvo_secundario")]
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=rules):
            visible = excel_bancos._visible_field_keys(self.fields, "teste", self.data)

        self.assertNotIn("alvo_secundario", visible)

    def test_dynamic_scope_keeps_original_workbook_column(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(
            [
                "DESCRICAO PRIMARIA",
                "DESCRICAO SECUNDARIA",
                "SUFIXO",
                "ORIGEM - PRIMARIO",
                "ALVO PRIMARIO - PRIMARIO",
                "ALVO SECUNDARIO - SECUNDARIO",
            ]
        )
        initial_columns = worksheet.max_column
        initial_mapping = excel_bancos._resolve_field_column_map(worksheet, self.fields, create_missing=True)

        rules = [self._rule("set_secondary", "alvo_primario")]
        with patch.object(excel_bancos, "_combined_conditional_rules", return_value=rules):
            excel_bancos.build_descriptions(self.fields, self.data, "teste")
        final_mapping = excel_bancos._resolve_field_column_map(worksheet, self.fields, create_missing=True)

        self.assertEqual(worksheet.max_column, initial_columns)
        self.assertEqual(final_mapping["alvo_primario"], initial_mapping["alvo_primario"])
        self.assertEqual(worksheet.cell(1, final_mapping["alvo_primario"][0]).value, "ALVO PRIMARIO - PRIMARIO")


class OptionIdentityTests(unittest.TestCase):
    def test_parenthesized_options_keep_their_content_for_duplicate_detection(self):
        self.assertNotEqual(
            excel_bancos.option_identity("2- (2REC / 1 REB)"),
            excel_bancos.option_identity("(2FIX / 1REB)"),
        )
        self.assertEqual(
            excel_bancos.option_identity("12- 2FIX / 1REB"),
            excel_bancos.option_identity("(2FIX / 1REB)"),
        )


class DistanciaPeTests(unittest.TestCase):
    def test_distancia_pe_is_always_ordered_by_vao_sequence(self):
        field = {
            "key": excel_bancos.DISTANCIA_PE_KEY,
            "label": "DISTÂNCIA PÉ",
            "scope": "secundaria",
            "selection_mode": excel_bancos.SELECTION_MODE_MULTIPLA,
            "description_order": 1,
            "options": [
                "2- SEGUNDO VAO 810 MM",
                "8- PRIMEIRO VAO 265 MM",
                "16- TERCEIRO VAO 1370 MM",
                "18- QUARTO VAO 1600 MM",
            ],
        }
        data = {
            excel_bancos.DISTANCIA_PE_KEY: [
                "16- TERCEIRO VAO 1370 MM",
                "2- SEGUNDO VAO 810 MM",
                "18- QUARTO VAO 1600 MM",
                "8- PRIMEIRO VAO 265 MM",
            ]
        }

        values = excel_bancos._serialize_field_values(field, data)
        saved = excel_bancos._format_field_saved_value(field, values)
        description = excel_bancos._format_field_description(field, values)

        self.assertEqual(
            values,
            [
                "8- PRIMEIRO VAO 265 MM",
                "2- SEGUNDO VAO 810 MM",
                "16- TERCEIRO VAO 1370 MM",
                "18- QUARTO VAO 1600 MM",
            ],
        )
        self.assertEqual(
            saved,
            "ORIENTADO A ESQ: 8- PRIMEIRO VAO 265 MM | 2- SEGUNDO VAO 810 MM | "
            "16- TERCEIRO VAO 1370 MM | 18- QUARTO VAO 1600 MM",
        )
        self.assertEqual(
            description,
            "ORIENTADO A ESQ: PRIMEIRO VAO 265 MM, SEGUNDO VAO 810 MM, "
            "TERCEIRO VAO 1370 MM, QUARTO VAO 1600 MM",
        )


class PnCodeTests(unittest.TestCase):
    def setUp(self):
        self.prefix_field = {
            "key": "prefixo",
            "label": "PREFIXO",
            "scope": "primaria",
            "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
            "description_order": 1,
            "options": ["CJ", "VIDRO", "PP", "JI CONFORT"],
        }

    def test_pn_prefix_uses_group_plus_category(self):
        category = {"label": "12 - VIDROS", "sheet_name": "12 - VIDROS"}

        self.assertEqual(excel_bancos.pn_code_prefix(category, [self.prefix_field], {"prefixo": "CJ"}), "3012")
        self.assertEqual(excel_bancos.pn_code_prefix(category, [self.prefix_field], {"prefixo": "VIDRO"}), "1012")

    def test_pn_prefix_supports_produto_processo_and_transformacao(self):
        piso = {"label": "14 - PISO", "sheet_name": "14 - PISO"}
        veiculo_pb = {"label": "34 - VEICULO P.B.", "sheet_name": "34 - VEICULO P.B."}

        self.assertEqual(excel_bancos.pn_code_prefix(piso, [self.prefix_field], {"prefixo": "PP"}), "2014")
        self.assertEqual(excel_bancos.pn_code_prefix(veiculo_pb, [self.prefix_field], {"prefixo": "JI CONFORT"}), "4034")

    def test_pn_prefix_uses_configurable_group(self):
        category = {"label": "12 - VIDROS", "sheet_name": "12 - VIDROS"}
        catalog = {
            "pn_groups": [
                {"code": "50", "label": "SERVICO", "prefixes": ["SERVICO", "INSTALACAO"]},
            ]
        }
        with patch.object(excel_bancos, "load_catalog", return_value=catalog):
            self.assertEqual(excel_bancos.pn_code_prefix(category, [self.prefix_field], {"prefixo": "SERVICO"}), "5012")

    def test_next_sku_scans_only_same_group_category_prefix(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["SKU"])
        worksheet.append(["SKU"])
        worksheet.append(["30120001"])
        worksheet.append(["10120001"])
        worksheet.append(["30120002"])

        category = {"label": "12 - VIDROS", "sheet_name": "12 - VIDROS"}

        self.assertEqual(
            excel_bancos._next_sequential_sku(worksheet, 1, category, [self.prefix_field], {"prefixo": "CJ"}),
            "30120003",
        )
        self.assertEqual(
            excel_bancos._next_sequential_sku(worksheet, 1, category, [self.prefix_field], {"prefixo": "VIDRO"}),
            "10120002",
        )


if __name__ == "__main__":
    unittest.main()

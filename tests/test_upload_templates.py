from io import BytesIO
import unittest

from openpyxl import load_workbook

import supabase_store
import supabase_suprimentos


class UploadTemplateTests(unittest.TestCase):
    def assert_template(self, content, sheet_name, expected_headers):
        workbook = load_workbook(BytesIO(content), read_only=True)
        try:
            self.assertEqual(workbook.active.title, sheet_name)
            headers = [cell.value for cell in next(workbook.active.iter_rows(min_row=1, max_row=1))]
            self.assertEqual(headers, expected_headers)
            self.assertEqual(workbook.active.max_row, 1)
            self.assertIn("INSTRUCOES", workbook.sheetnames)
        finally:
            workbook.close()

    def test_pessoas_template_matches_importer_headers(self):
        self.assert_template(
            supabase_suprimentos.template_pessoas_xlsx(),
            "PESSOAS",
            list(supabase_suprimentos.PESSOA_FIELDS),
        )

    def test_processos_template_matches_importer_headers(self):
        self.assert_template(
            supabase_suprimentos.template_processos_xlsx(),
            "PROCESSOS",
            ["conjunto", "processo", "atividade", "responsavel"],
        )

    def test_regras_template_matches_importer_headers(self):
        self.assert_template(
            supabase_suprimentos.template_regras_xlsx(),
            "REGRAS_ITEM",
            ["id_regra", "item_gatilho", "itens_opcoes", "quantidade", "quantidade_editavel"],
        )

    def test_relacoes_template_matches_importer_headers(self):
        self.assert_template(
            supabase_suprimentos.template_relacoes_xlsx(),
            "PROCESSO_ITEM",
            ["item_codigo", "processos"],
        )

    def test_bom_template_matches_importer_headers(self):
        self.assert_template(
            supabase_store.template_bom_xlsx(),
            "BOM",
            [
                "item_codigo",
                "item_descricao",
                "componente_codigo",
                "descricao",
                "unidade",
                "quantidade",
            ],
        )

    def test_bom_template_can_be_filled_and_parsed(self):
        workbook = load_workbook(BytesIO(supabase_store.template_bom_xlsx()))
        workbook.active.append(
            ["30180001", "CJ TESTE", "10180001", "COMPONENTE TESTE", "pc", 2]
        )
        output = BytesIO()
        workbook.save(output)
        workbook.close()

        result = supabase_store._parse_bom_workbook(output.getvalue(), "template_bom.xlsx")

        self.assertEqual(result["30180001"]["parent_description"], "CJ TESTE")
        self.assertEqual(result["30180001"]["components"][0]["codigo"], "10180001")
        self.assertEqual(result["30180001"]["components"][0]["quantidade"], 2)


if __name__ == "__main__":
    unittest.main()

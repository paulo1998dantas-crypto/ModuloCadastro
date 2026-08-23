import io
import unittest
from copy import deepcopy
from unittest.mock import patch

from openpyxl import load_workbook

import catalogo_regras_xlsx
import excel_bancos


def _catalog():
    return {
        "version": 2,
        "active_category": "teste",
        "pn_groups": [
            {"code": "20", "label": "PRODUTO / PROCESSO", "prefixes": ["PP"]},
            {"code": "30", "label": "CONJUNTO / KIT", "prefixes": ["CJ"]},
        ],
        "categories": [
            {
                "key": "teste",
                "label": "10 - TESTE",
                "sheet_name": "10 - TESTE",
                "field_overrides": {},
                "fields": [
                    {
                        "key": "origem",
                        "label": "ORIGEM",
                        "scope": "primaria",
                        "selection_mode": "unitaria",
                        "description_order": 1,
                        "required": True,
                        "free_text": False,
                        "options": ["1- ATIVO", "2- INATIVO"],
                    },
                    {
                        "key": "destino",
                        "label": "DESTINO",
                        "scope": "secundaria",
                        "selection_mode": "unitaria",
                        "description_order": 2,
                        "required": False,
                        "free_text": True,
                        "options": [],
                    },
                ],
                "conditional_rules": [
                    {
                        "key": "regra-teste",
                        "source_field_key": "origem",
                        "source_field_label": "ORIGEM",
                        "source_field_scope": "primaria",
                        "source_values": ["ATIVO"],
                        "target_field_key": "destino",
                        "target_field_label": "DESTINO",
                        "target_field_scope": "secundaria",
                        "action": "show",
                        "match_by": "option",
                    }
                ],
            }
        ],
    }


class CatalogoRegrasXlsxTests(unittest.TestCase):
    def _export(self, catalog=None):
        source = deepcopy(catalog or _catalog())
        with patch.object(excel_bancos, "load_catalog", return_value=source):
            return catalogo_regras_xlsx.export_catalog_workbook()

    def test_exporta_exatamente_as_duas_abas_e_colunas_do_contrato(self):
        content = self._export()
        workbook = load_workbook(io.BytesIO(content), data_only=True)
        try:
            self.assertEqual(
                workbook.sheetnames,
                [catalogo_regras_xlsx.SHEET_FIELDS, catalogo_regras_xlsx.SHEET_RULES],
            )
            self.assertEqual(
                [cell.value for cell in workbook[catalogo_regras_xlsx.SHEET_FIELDS][1]],
                catalogo_regras_xlsx.FIELD_HEADERS,
            )
            self.assertEqual(
                [cell.value for cell in workbook[catalogo_regras_xlsx.SHEET_RULES][1]],
                catalogo_regras_xlsx.RULE_HEADERS,
            )
            self.assertEqual(workbook[catalogo_regras_xlsx.SHEET_FIELDS]["H2"].value, "SIM")
            self.assertEqual(workbook[catalogo_regras_xlsx.SHEET_FIELDS]["I3"].value, "TEXTO_LIVRE")
        finally:
            workbook.close()

    def test_importacao_atualiza_campos_opcoes_e_regras_e_e_idempotente(self):
        content = self._export()
        workbook = load_workbook(io.BytesIO(content))
        fields = workbook[catalogo_regras_xlsx.SHEET_FIELDS]
        rules = workbook[catalogo_regras_xlsx.SHEET_RULES]
        fields["H2"] = "NAO"
        fields["J2"] = "1- ATIVO; 2- INATIVO; 3- PENDENTE"
        fields.append(
            [
                "UPSERT", "10 - TESTE", "teste", "NOVO CAMPO", "novo_campo",
                "SECUNDARIA", "MULTIPLA", "SIM", "LISTA", "1- A; 2- B", 3,
            ]
        )
        rules["E2"] = "HIDE"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        saved = []
        catalog = _catalog()
        with patch.object(excel_bancos, "load_catalog", side_effect=lambda: deepcopy(saved[-1] if saved else catalog)), patch.object(
            excel_bancos, "save_catalog", side_effect=lambda value: saved.append(deepcopy(value))
        ):
            first = catalogo_regras_xlsx.import_catalog_workbook(output.getvalue())
            second = catalogo_regras_xlsx.import_catalog_workbook(output.getvalue())

        self.assertEqual(first["fields_inserted"], 1)
        self.assertEqual(second["fields_inserted"], 0)
        category = saved[-1]["categories"][0]
        origem = next(field for field in category["fields"] if field["key"] == "origem")
        self.assertFalse(origem["required"])
        self.assertIn("3- PENDENTE", origem["options"])
        self.assertEqual(sum(field["key"] == "novo_campo" for field in category["fields"]), 1)
        self.assertEqual(category["conditional_rules"][0]["action"], "hide")

    def test_campo_virtual_e_gravado_como_sobrescrita_da_categoria(self):
        catalog = _catalog()
        category = catalog["categories"][0]
        category["key"] = excel_bancos.DEFAULT_CATEGORY_KEY
        category["label"] = excel_bancos.DEFAULT_CATEGORY_LABEL
        catalog["active_category"] = excel_bancos.DEFAULT_CATEGORY_KEY
        with patch.object(excel_bancos, "get_conditional_rules", return_value=[]):
            content = self._export(catalog)
        workbook = load_workbook(io.BytesIO(content))
        fields = workbook[catalogo_regras_xlsx.SHEET_FIELDS]
        key_column = catalogo_regras_xlsx.FIELD_HEADERS.index("CHAVE_CAMPO") + 1
        required_column = catalogo_regras_xlsx.FIELD_HEADERS.index("OBRIGATORIO") + 1
        row_number = next(
            row
            for row in range(2, fields.max_row + 1)
            if fields.cell(row, key_column).value == "cj_layout"
        )
        fields.cell(row_number, required_column).value = "SIM"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        saved = []
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)), patch.object(
            excel_bancos, "save_catalog", side_effect=lambda value: saved.append(deepcopy(value))
        ):
            catalogo_regras_xlsx.import_catalog_workbook(output.getvalue())

        overrides = saved[0]["categories"][0]["field_overrides"]
        self.assertTrue(overrides["cj_layout"]["required"])

    def test_erro_em_regra_nao_salva_parcialmente(self):
        content = self._export()
        workbook = load_workbook(io.BytesIO(content))
        rules = workbook[catalogo_regras_xlsx.SHEET_RULES]
        target_key_column = catalogo_regras_xlsx.RULE_HEADERS.index("CHAVE_CAMPO_DESTINO") + 1
        target_label_column = catalogo_regras_xlsx.RULE_HEADERS.index("CAMPO_DESTINO") + 1
        action_column = catalogo_regras_xlsx.RULE_HEADERS.index("ACAO") + 1
        rules.cell(2, target_key_column).value = "campo_inexistente"
        rules.cell(2, target_label_column).value = "CAMPO INEXISTENTE"
        rules.cell(2, action_column).value = "SET_PRIMARY"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with patch.object(excel_bancos, "load_catalog", return_value=_catalog()), patch.object(
            excel_bancos, "save_catalog"
        ) as save:
            with self.assertRaisesRegex(ValueError, "campo destino existente"):
                catalogo_regras_xlsx.import_catalog_workbook(output.getvalue())
        save.assert_not_called()

    def test_regra_por_grupo_e_exportada_e_reimportada(self):
        catalog = _catalog()
        catalog["categories"][0]["conditional_rules"].append(
            {
                "key": "regra-grupo",
                "source_type": "group",
                "source_field_key": excel_bancos.PN_GROUP_FORM_KEY,
                "source_field_label": "GRUPO DO SKU",
                "source_field_scope": "estrutura",
                "source_values": ["30"],
                "target_field_key": "destino",
                "target_field_label": "DESTINO",
                "target_field_scope": "secundaria",
                "action": "hide",
                "match_by": "option",
            }
        )
        content = self._export(catalog)
        saved = []
        with patch.object(excel_bancos, "load_catalog", return_value=deepcopy(catalog)), patch.object(
            excel_bancos, "save_catalog", side_effect=lambda value: saved.append(deepcopy(value))
        ):
            catalogo_regras_xlsx.import_catalog_workbook(content)

        rules = saved[0]["categories"][0]["conditional_rules"]
        group_rule = next(rule for rule in rules if rule["key"] == "regra-grupo")
        self.assertEqual(group_rule["source_type"], "group")
        self.assertEqual(group_rule["source_field_key"], excel_bancos.PN_GROUP_FORM_KEY)
        self.assertEqual(group_rule["source_values"], ["30"])


if __name__ == "__main__":
    unittest.main()

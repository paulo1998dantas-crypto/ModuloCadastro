import io
import unittest

from openpyxl import Workbook

import excel_bancos
import supabase_store
import technical_fields_reconciliation as reconciliation


CATEGORY = {"key": "cat_12_vidros", "label": "12 - VIDROS"}
PECAS_BCO_CATEGORY = {"key": "cat_22_pecas_bco", "label": "22 - PEÇAS BCO"}


def _fields():
    return [
        {
            "key": "fornecedor",
            "label": "FORNECEDOR",
            "scope": "primaria",
            "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
            "description_order": 1,
            "required": False,
            "free_text": False,
            "options": ["1- SALT"],
        },
        {
            "key": "especificidade",
            "label": "ESPECIFICIDADE",
            "scope": "secundaria",
            "selection_mode": excel_bancos.SELECTION_MODE_MULTIPLA,
            "description_order": 2,
            "required": False,
            "free_text": False,
            "options": ["1- N/A", "2- LATERAL"],
        },
    ]


def _rows():
    return [
        {
            "id": "item-1",
            "sku": "12180001",
            "form_values": {
                excel_bancos.PN_GROUP_FORM_KEY: ["10 - INSUMO"],
                "fornecedor": ["1- SALT"],
                "especificidade": ["1- N/A"],
            },
        },
        {
            "id": "item-2",
            "sku": "12180002",
            "form_values": {
                excel_bancos.PN_GROUP_FORM_KEY: ["10 - INSUMO"],
                "fornecedor": ["1- SALT"],
                "especificidade": ["2- LATERAL"],
            },
        },
    ]


def _workbook_bytes(rows, category=CATEGORY, fields=None):
    fields = fields or _fields()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CAMPOS_TECNICOS"
    worksheet.append(["COD", "CATEGORIA", "GRUPO", "FORNECEDOR", "ESPECIFICIDADE"])
    for row in rows:
        worksheet.append(row)
    reconciliation._write_template_metadata(workbook, category, fields)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


class TechnicalFieldsReconciliationTests(unittest.TestCase):
    def test_template_contains_all_active_skus_and_field_headers(self):
        content = reconciliation.build_template(CATEGORY, _fields(), _rows())
        loaded = reconciliation.load_rows(content, CATEGORY, _fields())
        self.assertEqual([row["sku"] for row in loaded], ["12180001", "12180002"])
        self.assertEqual(loaded[0]["values"]["fornecedor"], "SALT")

    def test_template_without_metadata_is_rejected_before_reading_rows(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CAMPOS_TECNICOS"
        worksheet.append(["COD", "CATEGORIA", "GRUPO", "FORNECEDOR", "ESPECIFICIDADE"])
        worksheet.append(["12180001", "12 - VIDROS", "10 - INSUMO", "SALT", "N/A"])
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "Template legado"):
            reconciliation.load_rows(output.getvalue(), CATEGORY, _fields())

    def test_template_with_stale_field_projection_is_rejected(self):
        content = reconciliation.build_template(CATEGORY, _fields(), _rows())
        workbook = reconciliation.load_workbook(io.BytesIO(content))
        metadata = workbook[reconciliation.META_SHEET]
        for row in metadata.iter_rows(min_col=1, max_col=2):
            if row[0].value == "field_fingerprint":
                row[1].value = "stale"
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "catálogo da categoria mudou"):
            reconciliation.load_rows(output.getvalue(), CATEGORY, _fields())

    def test_import_requires_exact_active_sku_set(self):
        content = _workbook_bytes([["12180001", "12 - VIDROS", "10 - INSUMO", "SALT", "N/A"]])
        with self.assertRaisesRegex(ValueError, "ativos ausentes"):
            reconciliation.prepare_reconciliation(
                content,
                CATEGORY,
                _fields(),
                _rows(),
                lambda row, values: {"id": row["id"], "form_values": values},
            )

    def test_unknown_option_rejects_complete_upload(self):
        content = _workbook_bytes(
            [
                ["12180001", "12 - VIDROS", "10 - INSUMO", "VTRX", "N/A | LATERAL"],
                ["12180002", "12 - VIDROS", "10 - INSUMO", "SALT", "LATERAL"],
            ]
        )
        with self.assertRaisesRegex(ValueError, "VTRX.*não existe no catálogo fechado"):
            reconciliation.missing_field_options(content, CATEGORY, _fields())

    def test_group_is_informative_and_preserved_from_registration(self):
        content = _workbook_bytes(
            [
                ["12180001", "12 - VIDROS", "30 - CONJUNTO", "SALT", "N/A"],
                ["12180002", "12 - VIDROS", "10 - INSUMO", "SALT", "LATERAL"],
            ]
        )
        result = reconciliation.prepare_reconciliation(
            content,
            CATEGORY,
            _fields(),
            _rows(),
            lambda row, values: {"id": row["id"], "form_values": values},
        )
        self.assertEqual(result["total"], 2)
        self.assertEqual(
            result["payloads"][0]["form_values"][excel_bancos.PN_GROUP_FORM_KEY],
            ["10 - INSUMO"],
        )

    def test_pecas_bco_converge_aliases_sem_criar_opcoes(self):
        fields = [
            {
                "key": "cor",
                "label": "COR",
                "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 1,
                "required": False,
                "free_text": False,
                "options": ["7- PRETO/CINZA"],
            },
            {
                "key": "fornecedor",
                "label": "FORNECEDOR",
                "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 2,
                "required": False,
                "free_text": False,
                "options": ["4- MC/CS", "8- ORI", "12- MC"],
            },
            {
                "key": "descritor_base",
                "label": "IDENTIFICACAO",
                "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 3,
                "required": False,
                "free_text": False,
                "options": ["7- PEGA MAO BCO"],
            },
        ]
        rows = [
            {
                "id": "peca-1",
                "sku": "10220001",
                "form_values": {excel_bancos.PN_GROUP_FORM_KEY: ["10 - INSUMO"]},
            }
        ]
        content = self._pecas_bco_workbook("PRETO C/ CINZA", "MC", "PEGA MAO")

        self.assertEqual(
            reconciliation.missing_field_options(content, PECAS_BCO_CATEGORY, fields),
            {},
        )
        result = reconciliation.prepare_reconciliation(
            content,
            PECAS_BCO_CATEGORY,
            fields,
            rows,
            lambda row, values: {"id": row["id"], "form_values": values},
        )
        self.assertEqual(result["payloads"][0]["form_values"]["cor"], ["7- PRETO/CINZA"])
        self.assertEqual(result["payloads"][0]["form_values"]["fornecedor"], ["12- MC"])
        self.assertEqual(
            reconciliation._matching_options(
                fields[1],
                "MC/CS",
                PECAS_BCO_CATEGORY["key"],
            ),
            ["4- MC/CS"],
        )
        self.assertEqual(
            reconciliation._matching_options(
                fields[1],
                "ORIGINAL",
                PECAS_BCO_CATEGORY["key"],
            ),
            ["8- ORI"],
        )
        self.assertEqual(
            result["payloads"][0]["form_values"]["descritor_base"],
            ["7- PEGA MAO BCO"],
        )

    def test_pecas_bco_rejeita_opcao_desconhecida_sem_ampliar_catalogo(self):
        fields = [
            {
                "key": "fornecedor",
                "label": "FORNECEDOR",
                "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 1,
                "required": False,
                "free_text": False,
                "options": ["4- MC/CS", "8- ORI"],
            },
        ]
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CAMPOS_TECNICOS"
        worksheet.append(["COD", "CATEGORIA", "GRUPO", "FORNECEDOR"])
        worksheet.append(["10220001", "22 - PEÇAS BCO", "10 - INSUMO", "FORNECEDOR NOVO"])
        reconciliation._write_template_metadata(workbook, PECAS_BCO_CATEGORY, fields)
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "Nenhuma opção foi criada"):
            reconciliation.missing_field_options(output.getvalue(), PECAS_BCO_CATEGORY, fields)

    @staticmethod
    def _pecas_bco_workbook(cor, fornecedor, identificacao):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "CAMPOS_TECNICOS"
        worksheet.append(["COD", "CATEGORIA", "GRUPO", "COR", "FORNECEDOR", "IDENTIFICACAO"])
        worksheet.append(["10220001", "22 - PEÇAS BCO", "10 - INSUMO", cor, fornecedor, identificacao])
        fields = [
            {
                "key": "cor", "label": "COR", "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 1, "required": False, "free_text": False,
                "options": ["7- PRETO/CINZA"],
            },
            {
                "key": "fornecedor", "label": "FORNECEDOR", "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 2, "required": False, "free_text": False,
                "options": ["4- MC/CS", "8- ORI", "12- MC"],
            },
            {
                "key": "descritor_base", "label": "IDENTIFICACAO", "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 3, "required": False, "free_text": False,
                "options": ["7- PEGA MAO BCO"],
            },
        ]
        reconciliation._write_template_metadata(workbook, PECAS_BCO_CATEGORY, fields)
        output = io.BytesIO()
        workbook.save(output)
        workbook.close()
        return output.getvalue()

    def test_supabase_payload_contains_required_identity_for_bulk_upsert(self):
        row = {
            "id": 504,
            "sku": "10180001",
            "unidade": "pc",
            "ativo": True,
            "form_values": {
                excel_bancos.PN_GROUP_FORM_KEY: ["10 - INSUMO"],
                "fornecedor": ["1- SALT"],
                "especificidade": ["1- N/A"],
            },
        }
        payload = supabase_store._technical_fields_reconciliation_payload(
            row,
            CATEGORY,
            _fields(),
            row["form_values"],
        )

        self.assertEqual(payload["id"], 504)
        self.assertEqual(payload["category_key"], CATEGORY["key"])
        self.assertEqual(payload["category_label"], CATEGORY["label"])
        self.assertEqual(payload["sku"], "10180001")
        self.assertEqual(payload["unidade"], "pc")
        self.assertTrue(payload["ativo"])

    def test_payload_preserva_grupo_mesmo_se_catalogo_contiver_chave_protegida(self):
        fields = [
            *_fields(),
            {
                "key": excel_bancos.PN_GROUP_FORM_KEY,
                "label": "GRUPO",
                "scope": "primaria",
                "selection_mode": excel_bancos.SELECTION_MODE_UNITARIA,
                "description_order": 99,
                "required": False,
                "free_text": False,
                "options": ["10 - INSUMO", "30 - CONJUNTO"],
            },
        ]
        row = {
            "id": 504,
            "sku": "30180001",
            "unidade": "cj",
            "ativo": True,
            "form_values": {
                excel_bancos.PN_GROUP_FORM_KEY: ["30"],
                "fornecedor": ["1- SALT"],
                "especificidade": ["1- N/A"],
            },
        }
        attempted = {
            **row["form_values"],
            excel_bancos.PN_GROUP_FORM_KEY: ["10 - INSUMO"],
        }

        payload = supabase_store._technical_fields_reconciliation_payload(
            row,
            CATEGORY,
            fields,
            attempted,
        )

        self.assertEqual(payload["form_values"][excel_bancos.PN_GROUP_FORM_KEY], ["30"])


if __name__ == "__main__":
    unittest.main()

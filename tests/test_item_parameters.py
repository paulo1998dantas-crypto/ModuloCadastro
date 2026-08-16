from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import Workbook

import leadtime_import
import supabase_store


class ItemParameterValidationTests(unittest.TestCase):
    def test_nonnegative_decimal_accepts_brazilian_value(self):
        self.assertEqual(
            "1234.5000",
            supabase_store._nonnegative_decimal(
                "R$ 1.234,50",
                "Preço",
                places="0.0001",
            ),
        )

    def test_nonnegative_decimal_rejects_negative_value(self):
        with self.assertRaises(supabase_store.SupabaseStoreError):
            supabase_store._nonnegative_decimal("-1", "Produção")

    def test_save_external_parameter_zeros_internal_times(self):
        registration = {"id": 81, "sku": "30200049"}
        values = {
            "origem_fabricacao": "EXTERNA",
            "fornecimento_dias": "5",
            "transporte_dias": "1,5",
            "recebimento_dias": "1",
            "inspecao_recebimento_dias": "0,5",
            "estocagem_dias": "1",
            "expedicao_dias": "2",
            "montagem_kit_dias": "3",
            "preco_compra": "125,90",
        }
        with patch.object(supabase_store, "_request", return_value=[]) as request:
            supabase_store.save_item_parameter(registration, values, "PAULO")

        payload = request.call_args.kwargs["payload"]
        self.assertEqual("EXTERNA", payload["origem_fabricacao"])
        self.assertEqual("1.500", payload["transporte_dias"])
        self.assertEqual("125.9000", payload["preco_compra"])
        self.assertEqual("0.000", payload["setup_dias"])
        self.assertEqual("0.000", payload["producao_dias"])
        self.assertEqual("0.000", payload["liberacao_dias"])

    def test_average_cost_uses_bounded_rpc_limit(self):
        response = [{"preco_medio": 12.5, "entradas_consideradas": 3}]
        with patch.object(supabase_store, "_request", return_value=response) as request:
            result = supabase_store.get_item_average_cost(" 10200001 ", limit=500)

        self.assertEqual(12.5, result["preco_medio"])
        request.assert_called_once_with(
            "POST",
            "rpc/cadastro_calcular_preco_medio",
            payload={"p_sku": "10200001", "p_limite": 100},
        )

    def test_reassign_is_backward_compatible_before_migration(self):
        error = supabase_store.SupabaseStoreError(
            'Erro Supabase 404: {"code":"PGRST205","message":"cadastro_item_parametros"}'
        )
        with patch.object(supabase_store, "_request", side_effect=error):
            self.assertFalse(supabase_store._reassign_item_parameter(1, 2, "20000001"))


class LeadTimeWorkbookTests(unittest.TestCase):
    def _workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = leadtime_import.SHEET_NAME
        headers = [
            "PN", "GRUPO", "DESCRIÇÃO PRIMÁRIA", "FORNEC", "TRANSP.",
            "RECBTO", "INSP.", "ESTOCAGEM", "EXPED.", "MONTAGEM",
            "PRODUÇÃO", "LIBERAÇÃO",
        ]
        for column, value in enumerate(headers, start=2):
            sheet.cell(leadtime_import.HEADER_ROW, column, value)

        sheet.cell(7, 2, 10100001)
        sheet.cell(7, 3, "10 - INSUMO")
        sheet.cell(7, 4, "ITEM EXTERNO")
        for column, value in enumerate([5, 1, 1, 0.5, 1, 2, 0], start=5):
            sheet.cell(7, column, value)

        sheet.cell(8, 2, 20100001)
        sheet.cell(8, 3, "20 - PRODUTO PROCESSO")
        sheet.cell(8, 4, "ITEM INTERNO")
        sheet.cell(8, 12, 3)
        sheet.cell(8, 13, 1)

        sheet.cell(9, 2, 30100001)
        sheet.cell(9, 3, "30 - CONJUNTO")
        sheet.cell(9, 4, "SEM TEMPO")
        workbook.save(path)

    def test_parse_workbook_separates_origins_and_does_not_infer_blank(self):
        with TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "leadtime.xlsx"
            self._workbook(path)
            rows, issues = leadtime_import.parse_workbook(path)

        self.assertEqual(["EXTERNA", "INTERNA"], [row.origin for row in rows])
        self.assertEqual("1.000", rows[0].values["transporte_dias"])
        self.assertEqual("3.000", rows[1].values["producao_dias"])
        self.assertEqual("SEM_PARAMETRO", issues[0]["status"])
        self.assertEqual("30100001", issues[0]["sku"])


if __name__ == "__main__":
    unittest.main()
